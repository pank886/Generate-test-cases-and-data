"""后台异步任务：文件处理、聊天生成、测试计划确认。

所有同步阻塞调用（LLM、文件 I/O、LangGraph）均通过 asyncio.to_thread()
卸载到独立线程池，保持 FastAPI 事件循环始终可响应轮询请求。
"""

import asyncio
import os
import json as _json
import threading
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime

from observability import set_trace_id, get_logger

logger = get_logger(__name__)


class _BoundedThreadPoolExecutor(ThreadPoolExecutor):
    """有界线程池：队列满时 submit 阻塞（默认 LinkedBlockingQueue 会无限制累积）。"""

    def __init__(self, max_workers: int = 10, max_queue: int = 30, **kwargs):
        super().__init__(max_workers=max_workers, **kwargs)
        self._sem = threading.BoundedSemaphore(max_queue)

    def submit(self, fn, *args, **kwargs):
        self._sem.acquire()
        future = super().submit(fn, *args, **kwargs)
        future.add_done_callback(lambda _: self._sem.release())
        return future


import config as _config
_MAX_WORKERS = _config.TASK_MAX_WORKERS
_executor = _BoundedThreadPoolExecutor(max_workers=_MAX_WORKERS, max_queue=_config.TASK_MAX_QUEUE)


# ========================================================================
# 文件处理（上传 → 向量库入库）
# ========================================================================

async def _process_file_bg(task_id: str, file_path: str, ext: str,
                            file_size: int, filename: str, file_type: str):
    """后台处理上传文件 -> 向量库入库。"""
    from web.state import _add_imported_file, _update_task

    set_trace_id(task_id)
    loop = asyncio.get_running_loop()

    def _progress(pct: int, msg: str):
        """跨线程安全回调：轻量，只做 run_coroutine_threadsafe 一件事。"""
        asyncio.run_coroutine_threadsafe(
            _update_task(task_id, progress=pct, message=msg),
            loop,
        )

    try:
        await _update_task(task_id, status="running", progress=5,
                           message="接收文件，开始处理...")

        if ext == ".zip":
            from ingest_v2 import process_axure_zip
            _progress(10, "解压 Axure 包，解析页面结构...")
            result = await asyncio.to_thread(
                process_axure_zip,
                file_path,
                progress_cb=lambda p, m: _progress(10 + int(p * 0.8), m),
            )
            count = result.get("chunks", 0)
            source = "Axure 原型"
        elif ext == ".md":
            _progress(20, "MD 文件已接收，等待选择提取方式...")
            # 不在这里跑 LLM——前端弹窗让用户选 代码提取 / LLM 提取
            resp = {
                "success": True,
                "message": "请选择接口提取方式",
                "file_path": file_path,
                "file_name": filename,
                "needs_extract_choice": True,  # 前端检测此标记 → 弹选择窗
            }
            await _update_task(task_id, status="completed", progress=100,
                               message="等待选择提取方式", result=resp)
            return

        else:
            from ingest_v2 import process_product_doc
            _progress(10, "读取文档，提取模块信息...")
            result = await asyncio.to_thread(
                process_product_doc,
                file_path,
                progress_cb=lambda p, m: _progress(10 + int(p * 0.8), m),
            )
            count = result.get("chunks", 0)
            source = {".docx": "Word 文档", ".pdf": "PDF 文档"}.get(ext, "文档")

        await _update_task(task_id, progress=90,
                           message=f"{source} 处理完成：{count} 个文本块")

        if count == 0:
            await _update_task(task_id, status="failed",
                               error="文件解析后无内容，请检查文件是否有效。")
            return

        module_name = result.get("module_name")
        doc_id = result.get("doc_id")

        file_info = {
            "name": filename,
            "size": f"{file_size / 1024:.1f} KB",
            "chunks": count,
            "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "type": file_type,
            "status": "ready",  # 数据已写入 SQLite + ChromaDB
        }
        try:
            await _add_imported_file(file_info)
        except Exception:
            logger.warning("内存状态更新失败（数据已持久化，下次启动自动恢复）: %s", filename, exc_info=True)

        logger.info("✅ %s 处理完成：%d 个文本块", source, count)

        try:
            _meta = {"chunks": count, "type": file_type,
                     "time": datetime.now().isoformat(),
                     "module": module_name or "", "doc_id": doc_id or ""}
            with open(file_path + ".meta.json", "w", encoding="utf-8") as _mf:
                _json.dump(_meta, _mf, ensure_ascii=False)
        except Exception:
            logger.warning("写入 meta.json 失败: %s", file_path, exc_info=True)

        resp = {
            "success": True,
            "message": f"已处理 {count} 个文本块",
            "file": file_info,
        }
        if module_name:
            resp["module_name"] = module_name
            resp["doc_id"] = doc_id
            resp["related_modules"] = result.get("related_modules", [])

        await _update_task(task_id, status="completed", progress=100,
                           message="处理完成", result=resp)

    except FileNotFoundError:
        await _update_task(task_id, status="failed", error="上传文件不存在")
    except Exception as e:
        logger.error("❌ 文件处理失败: %s", e)
        await _update_task(task_id, status="failed", error=str(e))
        if ext == ".md":
            try:
                os.remove(file_path)
            except Exception:
                logger.warning("清理失败文件失败: %s", file_path, exc_info=True)


# ========================================================================
# ========================================================================
# Phase C: 确认计划 → 生成 .py + .yaml
# ========================================================================

def _resolve_api_defs(excel_path: str) -> str | None:
    """M8 门禁升级：SQLite 该模块 API 文档非空 → 投影轻量索引。

    快照 api_defs.json 已取消（D3），显式入参已删（2026-08-11）。
    模块名 = Excel feature 列（Phase B 写入的 _feature = confirmed_module）。

    返回轻量索引 JSON `[{name, method, url}]`；SQLite 该模块无 API 文档 → None 阻断
    （调用方必须显式失败，严禁空定义续跑）。
    """
    module = _read_excel_module(excel_path)
    if not module:
        logger.error("无法从 Excel feature 列解析模块名: %s", excel_path)
        return None

    from database import get_session_ctx
    from database.operations import BindingOps
    with get_session_ctx() as session:
        docs = BindingOps.get_bound_docs(session, module)
        apis = [d for d in docs if d.doc_type == "api"]
        if not apis:
            logger.warning(
                "M8 阻断：模块 [%s] 无 API 文档（SQLite 未绑定/为空），禁止空定义续跑",
                module)
            return None
        index = [{"name": d.api_name or f"api_{i}",
                  "method": d.api_method or "?",
                  "url": d.api_url or ""}
                 for i, d in enumerate(apis)]
    logger.info("M8 通过：模块 [%s] 投影轻量索引 %d 个接口", module, len(index))
    return _json.dumps(index, ensure_ascii=False)


def _read_excel_module(excel_path: str) -> str:
    """读取 Excel 首行 feature 列（列索引 1）作为模块名。

    Phase B 写入时 feature 列 = _feature = confirmed_module（nodes.py:501-503）。
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        try:
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                return str(row[1] or "").strip()
        finally:
            wb.close()
    except Exception:
        logger.error("读取 Excel feature 列失败: %s", excel_path, exc_info=True)
    return ""


def _find_missing_dep_map_cases(rows: list, dep_map: dict) -> list:
    """返回 dep_map 遗漏的 Excel 用例行（case_id 不在 case_api_sequences，D5）。"""
    covered = {
        cid
        for s in (dep_map or {}).get("stories", [])
        for cid in (s.get("case_api_sequences") or {})
    }
    return [r for r in rows if r.get("case_id") and r["case_id"] not in covered]


def _find_stories_missing_anchors(rows: list, dep_map: dict) -> list:
    """返回有共享前置但 dep_map 里 story_pre/teardown 为空的 story（含前置文本，D5/A）。

    返回 [{story_name, preconditions}]，供 repair_stories 定向补 story_pre_api_sequence
    / teardown_api_sequence（有前置必有其前置序列与清理序列，绝不接受空数据）。
    """
    from collections import defaultdict
    story_pre = defaultdict(set)
    for r in rows:
        pre = r.get("preconditions") or ""
        if pre and pre != "无":
            for pid in pre.split(","):
                if pid.strip().startswith("PRE-"):
                    story_pre[r.get("story")].add(pid.strip())
    out = []
    for s in (dep_map or {}).get("stories", []):
        if not isinstance(s, dict):
            continue
        name = s.get("story_name")
        if name in story_pre:
            if not s.get("story_pre_api_sequence") or not s.get("teardown_api_sequence"):
                out.append({"story_name": name,
                            "preconditions": ", ".join(sorted(story_pre[name]))})
    return out


def _load_module_analysis(excel_path: str) -> str:
    """读模块的 Phase B 分析（ModuleAnalysis）作为补漏修复上下文（D5）。"""
    module = _read_excel_module(excel_path)
    if not module:
        return ""
    try:
        from database import get_session_ctx
        from database.operations import ModuleOps, AnalysisOps
        with get_session_ctx() as session:
            mod = ModuleOps.get_by_name(session, module)
            if not mod:
                return ""
            record = AnalysisOps.get_by_module_id(session, mod.id)
            if not record:
                return ""
            return record.analysis_json or record.api_analysis or ""
    except Exception:
        logger.warning("读取模块分析失败（D5 补漏跳过分析）: %s", module, exc_info=True)
        return ""


def _merge_dep_map(existing: dict, repair: dict) -> dict:
    """把补漏修复的 partial dep_map 合并进现有 dep_map（D5）。

    按 story_name 对齐；case 级三表（case_api_sequences / decision_map /
    internal_dependency）以 repair 覆盖/新增；story 级 story_pre_api_sequence /
    teardown_api_sequence 在 repair 提供非空时覆盖（补漏 story 锚）。
    """
    merged = dict(existing or {})
    stories = list(merged.get("stories") or [])
    by_name = {s.get("story_name"): s for s in stories if isinstance(s, dict)}
    for rs in (repair or {}).get("stories", []):
        if not isinstance(rs, dict):
            continue
        name = rs.get("story_name")
        if name in by_name:
            base = by_name[name]
            for key in ("case_api_sequences", "decision_map", "internal_dependency"):
                if rs.get(key):
                    base[key] = {**(base.get(key) or {}), **(rs[key])}
            if rs.get("story_pre_api_sequence"):
                base["story_pre_api_sequence"] = rs["story_pre_api_sequence"]
            if rs.get("teardown_api_sequence"):
                base["teardown_api_sequence"] = rs["teardown_api_sequence"]
        else:
            stories.append(rs)
            by_name[name] = rs
    merged["stories"] = stories
    return merged


async def _confirm_plan_bg(task_id: str, excel_path: str | None,
                          user_ctx: str = ""):
    """后台执行确认计划 -> 生成 .py + .yaml。

    user_ctx 由 /confirm-plan 端点传入；接口索引从 SQLite 投影（_resolve_api_defs）。
    """
    import glob
    import config

    from web.state import _phase_b_components, _update_task, _chroma_db

    set_trace_id(task_id)

    # 重建 LLM 客户端，避免复用上一个工作流残留的僵死连接池
    from agent_components.nodes import reload_llm
    reload_llm()

    try:
        if not excel_path:
            excel_files = glob.glob(
                os.path.join(config.TESTCASE_BASE, "**", "test_plan.xlsx"),
                recursive=True,
            )
            if excel_files:
                excel_path = max(excel_files, key=os.path.getmtime)

        if not excel_path:
            await _update_task(task_id, status="failed",
                               error="未找到测试计划 Excel 文件")
            return

        if not _phase_b_components:
            await _update_task(task_id, status="failed",
                               error="组件未初始化")
            return

        # 规则 M8：接口定义缺失必须显式阻断，禁止空定义盲写 YAML
        api_defs_json = _resolve_api_defs(excel_path)
        if api_defs_json is None:
            await _update_task(
                task_id, status="failed",
                error="未找到接口定义（SQLite 该模块无 API 文档），"
                      "请先完成 API 文档入库并绑定到模块，再重新确认计划",
            )
            return

        # ---- Phase C Step 0: 生成 dependency_map.json ----
        await _update_task(task_id, status="running", progress=15,
                           message="正在分析用例依赖关系...")
        output_dir = os.path.dirname(excel_path)
        try:
            # 从 ChromaDB 检索 product_docs（用 confirmed_module 作为查询 key）
            product_docs_json = "[]"
            try:
                if _chroma_db is not None:
                    from agent_components.dual_chroma import get_chroma_db
                    db = get_chroma_db()
                    docs = db.search_product_docs(
                        query=user_ctx,
                        k=config.RETRIEVAL_K,
                    )
                    if docs:
                        product_docs_json = _json.dumps(
                            [{"page_content": d.page_content, "metadata": d.metadata}
                             for d in docs],
                            ensure_ascii=False,
                        )
            except Exception:
                logger.warning("ChromaDB product_docs 检索失败，使用空文档继续", exc_info=True)

            # 模块树
            from database import get_session_ctx
            from database.operations import ModuleOps
            with get_session_ctx() as session:
                tree = ModuleOps.get_tree(session)
            module_tree_json = _json.dumps(tree, indent=2, ensure_ascii=False)

            # 异步生成 dep_map（LLM thinking 调用走线程池）
            dep_map_path = await asyncio.to_thread(
                _phase_b_components._generate_dependency_map,
                excel_path, output_dir, api_defs_json,
                module_tree_json, product_docs_json,
                "（Phase C Step 0 生成）", user_ctx,
            )
        except Exception as e:
            logger.warning("   ⚠️ dependency_map.json 生成失败（非致命，继续后续步骤）: %s", e)
            dep_map_path = None
        if dep_map_path:
            logger.info("   📄 dependency_map.json 已生成: %s", dep_map_path)

        # ---- Phase C Step 1: 加载 + 预校验 dep_map（失败不阻断） ----
        dep_map = None
        try:
            if dep_map_path:
                with open(dep_map_path, "r", encoding="utf-8") as f:
                    dep_map = _json.load(f)
                if dep_map.get("stories"):
                    logger.info("   📄 dependency_map.json 已加载: %d 个 story",
                                len(dep_map.get("stories", [])))
                else:
                    logger.warning("   ⚠️ dependency_map.json 无有效 story，跳过")
        except Exception as e:
            logger.warning("   ⚠️ dependency_map.json 解析失败（非致命）: %s", e)

        # ---- D5 覆盖校验 + 定向修复（dep_map 漏 case / 漏 story 锚 → 补漏，带 Phase B 分析）----
        if dep_map:
            try:
                rows = _phase_b_components._read_excel_rows(excel_path)
                missing_cases = _find_missing_dep_map_cases(rows, dep_map)
                missing_stories = _find_stories_missing_anchors(rows, dep_map)
                if missing_cases or missing_stories:
                    logger.warning(
                        "   ⚠️ dep_map 遗漏 %d 个用例 + %d 个 story 前置/清理锚"
                        "（D5 定向修复）", len(missing_cases), len(missing_stories))
                    if missing_cases:
                        logger.warning("      漏用例: %s",
                                       [r.get("case_id") for r in missing_cases])
                    if missing_stories:
                        logger.warning("      漏 story 锚: %s",
                                       [s["story_name"] for s in missing_stories])
                    analysis = _load_module_analysis(excel_path)
                    repair_path = await asyncio.to_thread(
                        _phase_b_components._generate_dependency_map,
                        excel_path, output_dir, api_defs_json,
                        module_tree_json, product_docs_json,
                        "（Phase C Step 0 补漏修复）", user_ctx,
                        repair_cases=missing_cases or None,
                        repair_stories=missing_stories or None,
                        analysis=analysis,
                    )
                    if repair_path:
                        with open(repair_path, "r", encoding="utf-8") as f:
                            repair_map = _json.load(f)
                        dep_map = _merge_dep_map(dep_map, repair_map)
                        with open(dep_map_path, "w", encoding="utf-8") as f:
                            _json.dump(dep_map, f, ensure_ascii=False, indent=2)
                        logger.info(
                            "   ✅ D5 补漏合并完成: 现 %d 个 story",
                            len(dep_map.get("stories", [])))
            except Exception as e:
                logger.warning(
                    "   ⚠️ D5 覆盖校验/补漏修复失败（非致命，缺失项将跳过）: %s", e)

        await _update_task(task_id, status="running", progress=20,
                           message="正在生成 .py 测试文件...")

        # LLM 调用 → 线程池
        py_result = await asyncio.to_thread(
            _phase_b_components._generate_py_file, excel_path,
        )

        await _update_task(task_id, progress=50,
                           message="正在生成 YAML 数据文件...")

        # LLM 调用 → 线程池（带心跳，YAML 生成耗时长，避免前端轮询超时）
        import asyncio as _asyncio
        import time as _time

        _heartbeat_stop = False

        async def _heartbeat():
            nonlocal _heartbeat_stop
            _t0 = _time.time()
            while not _heartbeat_stop:
                await _asyncio.sleep(10)
                if _heartbeat_stop:
                    break
                elapsed = int(_time.time() - _t0)
                await _update_task(
                    task_id, progress=55,
                    message=f"正在生成 YAML 数据文件...（{elapsed}s）",
                )

        hb_task = _asyncio.create_task(_heartbeat())
        try:
            yaml_result = await asyncio.to_thread(
                _phase_b_components._generate_all_yamls,
                excel_path, dep_map, user_ctx,
            )
        finally:
            _heartbeat_stop = True
            hb_task.cancel()
            try:
                await hb_task
            except _asyncio.CancelledError:
                pass

        msg = f".py: {py_result['py_file_name']}（{py_result['modules']}模块）"
        if yaml_result["total"] > 0:
            msg += f" | YAML: {yaml_result['success']}/{yaml_result['total']} 个"
            if yaml_result.get("skipped_api_missing"):
                msg += f"，跳过接口缺失 {yaml_result['skipped_api_missing']} 个"
            if yaml_result.get("repaired"):
                msg += f"（含自查修复 {yaml_result['repaired']} 个）"
            if yaml_result.get("failed"):
                msg += (f"，仍失败 {yaml_result['failed']} 个"
                        f"（详见 _generation_errors.json 与 logs/thinking_trace.log）")

        result = {
            "success": True,
            "message": msg,
            "py_file": py_result["py_file_name"],
            "py_path": py_result.get("py_path", ""),
            "yaml_success": yaml_result["success"],
            "yaml_total": yaml_result["total"],
            "yaml_repaired": yaml_result.get("repaired", 0),
            "yaml_failed": yaml_result.get("failed", 0),
            "yaml_skipped": yaml_result.get("skipped_api_missing", 0),
            "yaml_rounds": yaml_result.get("rounds", 0),
            "errors_file": yaml_result.get("errors_file"),
            "excel_path": excel_path,
            "output_dir": os.path.dirname(excel_path),
        }

        await _update_task(task_id, status="completed", progress=100,
                           message="文件生成完成", result=result)

    except Exception as e:
        logger.error("❌ 确认计划失败: %s", e)
        await _update_task(task_id, status="failed", error=str(e))


# ========================================================================
# Phase B: 多轮工作流恢复执行
# ========================================================================

async def _resume_workflow_bg(task_id: str, session_id: str, state: dict):
    """Phase B 后台恢复执行：从节点2开始，完成产品文档检索→关联模块→接口→测试点→Excel。

    使用 _phase_b_graph.astream() 逐节点上报进度，前端实时可见。
    """
    import os as _os
    import config
    from web.state import _phase_b_graph, _update_task

    set_trace_id(task_id)

    # 节点 → 进度映射
    try:
        await _update_task(task_id, status="running", progress=10,
                           message="正在检索产品文档...")

        # LangGraph 在独立线程中执行（同步节点：ChromaDB、LLM 等），
        # 主协程保持可响应，定期发心跳避免前端超时
        import asyncio as _asyncio
        import time as _time

        _heartbeat_stop = False

        async def _heartbeat():
            nonlocal _heartbeat_stop
            _t0 = _time.time()
            _step = 1
            _messages = [
                "正在检索产品文档...",
                "正在分析关联模块...",
                "正在检索接口定义...",
                "正在分析测试场景...",
                "正在生成测试计划...",
            ]
            while not _heartbeat_stop:
                await _asyncio.sleep(10)
                if _heartbeat_stop:
                    break
                elapsed = int(_time.time() - _t0)
                msg = _messages[min(_step, len(_messages) - 1)]
                _step += 1
                await _update_task(
                    task_id, progress=15,
                    message=f"{msg}（{elapsed}s）",
                )

        hb_task = _asyncio.create_task(_heartbeat())
        try:
            result = await asyncio.to_thread(_phase_b_graph.invoke, state)
        finally:
            _heartbeat_stop = True
            hb_task.cancel()
            try:
                await hb_task
            except _asyncio.CancelledError:
                pass

        # 检查 NO_DATA 中断
        if result.get("workflow_status") == "NO_DATA":
            await _update_task(task_id, status="failed", progress=100,
                               error=result.get("confirmation_question",
                                                "未找到产品文档，请先导入数据"))
            return

        await _update_task(task_id, progress=85,
                           message="生成完成，正在保存结果...")

        # 构建响应
        plan = result.get("excel_plan")
        case_count = len(plan.test_cases) if plan and hasattr(plan, "test_cases") else 0

        # 从 thinking_trace.log 检查是否有校验失败的行
        fail_warn = ""
        failed_tc_ids = []
        try:
            with open(_os.path.join(config.LOG_DIR, "thinking_trace.log"), "r", encoding="utf-8") as _lf:
                content = _lf.read()
                if "generate_excel_plan_FAILED" in content:
                    fail_warn = "（部分用例校验失败，详见 logs/thinking_trace.log）"
                    # 提取失败用例编号
                    import re
                    failed_tc_ids = list(set(re.findall(
                        r"\| (TC-\d+) \|", content)))
        except Exception:
            logger.warning("无法读取思考日志，跳过失败用例提取", exc_info=True)

        thinking_parts = [f"Excel 计划 {case_count} 条用例"]
        if failed_tc_ids:
            thinking_parts.append(
                f"⚠️ {len(failed_tc_ids)} 行校验失败需人工审查: {', '.join(sorted(failed_tc_ids))}"
            )
        resp = {
            "success": True,
            "thinking": thinking_parts,
            "reply": f"Excel 测试计划已生成：共 {case_count} 条用例{fail_warn}",
        }
        if result.get("excel_path"):
            resp["excel_path"] = result["excel_path"]
            resp["excel_name"] = _os.path.basename(result["excel_path"])
            resp["output_dir"] = result.get(
                "output_dir", _os.path.dirname(result["excel_path"]),
            )
        if result.get("requires_review"):
            resp["requires_review"] = True
            resp["error_info"] = result.get("error_info", [])

        await _update_task(task_id, status="completed", progress=100,
                           message="测试计划生成完成", result=resp)

    except Exception as e:
        logger.error("❌ Phase B 工作流执行失败: %s", e)
        await _update_task(task_id, status="failed", error=str(e))
    finally:
        from web.state import _workflow_sessions, _workflow_sessions_lock
        async with _workflow_sessions_lock:
            _workflow_sessions.pop(session_id, None)


# ========================================================================
# Phase A: 三步分析管线（2026-07-31 — 替代旧 _analyze_module_scenarios_bg）
# ========================================================================

async def _analyze_module_scenarios_3step_bg(task_id: str, module_name: str):
    """三步分析管线：产品→场景 / Axure→逻辑关系 / API→接口映射。

    Step 1: product 文档 → scenario_analysis
    Step 2: axure 文档 → ui_flow_analysis（无 Axure 跳过）
    Step 3: api 定义 → api_analysis（无 API 跳过）
    每步 thinking 模式，1 次 LLM 调用，输出自由文本。
    """
    from web.state import _update_task
    from agent_components.nodes import reload_llm, _get_llm
    from prompts.extraction_prompts import (
        analyze_product_scenarios_prompt,
        analyze_axure_ui_flow_prompt,
        analyze_api_mapping_prompt,
    )
    import config as _cfg

    set_trace_id(task_id)
    reload_llm()

    try:
        await _update_task(task_id, status="running", progress=5,
                           message=f"三步分析「{module_name}」...")

        # ── 1. 查 module_id + 加载绑定文档（单 session 取完所有数据）──
        from database import get_session_ctx
        from database.operations import ModuleOps, BindingOps
        from database.models import DocumentChunk, Document

        with get_session_ctx() as session:
            mod = ModuleOps.get_by_name(session, module_name)
            if not mod:
                await _update_task(task_id, status="failed",
                                   error=f"模块「{module_name}」不存在")
                return
            module_id = mod.id

            bound_docs = BindingOps.get_bound_docs(session, module_name)
            partners = BindingOps.get_partners(session, "module", module_name, "module")

            # ── 分离三种文档类型（session 内提取纯数据）──
            product_chunks: list[str] = []  # Step 1 输入
            axure_chunks: list[tuple[str, str]] = []  # Step 2 输入: (page_name, content)
            api_defs_raw: list[dict] = []  # Step 3 输入（session 内取出纯 dict）

            for doc in bound_docs:
                if doc.doc_type == "api":
                    if doc.api_url:
                        api_defs_raw.append({
                            "name": doc.api_name, "url": doc.api_url,
                            "method": doc.api_method,
                            "description": doc.api_description,
                            "headers": _json.loads(doc.api_headers or "[]"),
                            "parameters": _json.loads(doc.api_parameters or "[]"),
                            "returns": _json.loads(doc.api_returns or "[]"),
                        })
                else:
                    chunks = session.query(DocumentChunk).filter_by(
                        doc_id=doc.id).order_by(DocumentChunk.chunk_index).all()
                    for c in chunks:
                        content = c.content
                        pn = c.page_name or doc.file_name
                        if doc.doc_type == "axure":
                            axure_chunks.append((pn, content))
                        else:
                            product_chunks.append(content)

            # 跨模块文本
            cross_text = ", ".join(p[1] for p in partners) if partners else "无"

            # 模块树
            tree = ModuleOps.get_tree(session)
            module_tree_json = _json.dumps(tree, indent=2, ensure_ascii=False)

        # ── session 已关闭，后续只使用提取后的纯数据 ──
        api_defs_json = _json.dumps(api_defs_raw, indent=2, ensure_ascii=False)

        llm = _get_llm()
        think_kwargs = {}
        if _cfg.ENABLE_THINKING:
            think_kwargs["extra_body"] = {"thinking": {"type": "enabled"}}

        scenario_text = ""
        ui_flow_text = ""
        api_text = ""

        # ── 心跳工具 ──
        import asyncio as _asyncio
        import time as _time

        async def _heartbeat(msg: str, base_progress: int):
            _t0 = _time.time()
            while not _heartbeat_stop:
                await _asyncio.sleep(10)
                if _heartbeat_stop:
                    break
                elapsed = int(_time.time() - _t0)
                await _update_task(
                    task_id, progress=base_progress,
                    message=f"{msg}（{elapsed}s）",
                )

        # ════════════════════════════════════════════
        # Step 1: 产品文档 → 测试场景总结
        # ════════════════════════════════════════════
        if product_chunks:
            await _update_task(task_id, progress=15,
                               message=f"Step 1/3: 分析产品文档（{len(product_chunks)} 块）...")
            product_docs_text = "\n\n---\n\n".join(product_chunks)

            _heartbeat_stop = False
            hb_task = _asyncio.create_task(_heartbeat("Step 1/3: LLM 分析产品文档...", 20))
            try:
                prompt = analyze_product_scenarios_prompt()
                bound_llm = llm.bind(temperature=0.6, **think_kwargs)
                result = await asyncio.to_thread(
                    bound_llm.invoke,
                    prompt.format_messages(
                        module_name=module_name,
                        product_docs=product_docs_text,
                        cross_module_relations=cross_text,
                    ),
                )
                scenario_text = result.content if hasattr(result, "content") else str(result)
            finally:
                _heartbeat_stop = True
                hb_task.cancel()
                try: await hb_task
                except _asyncio.CancelledError: pass

            # 写入 Step 1 结果
            with get_session_ctx() as session:
                from database.operations.analysis import AnalysisOps
                AnalysisOps.upsert_3step(
                    session, module_id, module_name,
                    scenario_analysis=scenario_text,
                )
                session.commit()
            await _update_task(task_id, progress=40,
                               message="Step 1/3 完成：场景总结已保存")
        else:
            logger.info("三步分析: 无 product 文档，跳过 Step 1")

        # ════════════════════════════════════════════
        # Step 2: 场景 + Axure → 逻辑关系总结
        # ════════════════════════════════════════════
        if axure_chunks and scenario_text:
            await _update_task(task_id, progress=45,
                               message=f"Step 2/3: 分析 Axure 页面（{len(axure_chunks)} 块）...")
            axure_text = "\n\n---\n\n".join(
                f"[页面: {pn}] {content}" for pn, content in axure_chunks
            )

            _heartbeat_stop = False
            hb_task = _asyncio.create_task(_heartbeat("Step 2/3: LLM 分析 Axure 交互...", 50))
            try:
                prompt = analyze_axure_ui_flow_prompt()
                bound_llm = llm.bind(temperature=0.6, **think_kwargs)
                result = await asyncio.to_thread(
                    bound_llm.invoke,
                    prompt.format_messages(
                        module_name=module_name,
                        scenario_analysis=scenario_text,
                        axure_pages=axure_text,
                    ),
                )
                ui_flow_text = result.content if hasattr(result, "content") else str(result)
            finally:
                _heartbeat_stop = True
                hb_task.cancel()
                try: await hb_task
                except _asyncio.CancelledError: pass

            # 写入 Step 2 结果
            with get_session_ctx() as session:
                from database.operations.analysis import AnalysisOps
                AnalysisOps.upsert_3step(
                    session, module_id, module_name,
                    ui_flow_analysis=ui_flow_text,
                )
                session.commit()
            await _update_task(task_id, progress=65,
                               message="Step 2/3 完成：交互逻辑已保存")
        else:
            logger.info("三步分析: 无 Axure 文档或场景文本，跳过 Step 2")

        # ════════════════════════════════════════════
        # Step 3: 场景 + 逻辑关系 + API → 接口总结
        # ════════════════════════════════════════════
        if api_defs_raw:
            await _update_task(task_id, progress=70,
                               message=f"Step 3/3: 分析 API 映射（{len(api_defs_raw)} 个接口文档）...")

            _heartbeat_stop = False
            hb_task = _asyncio.create_task(_heartbeat("Step 3/3: LLM 分析接口映射...", 75))
            try:
                prompt = analyze_api_mapping_prompt()
                bound_llm = llm.bind(temperature=0.6, **think_kwargs)
                result = await asyncio.to_thread(
                    bound_llm.invoke,
                    prompt.format_messages(
                        module_name=module_name,
                        scenario_analysis=scenario_text or "（未分析）",
                        ui_flow_analysis=ui_flow_text or "（未分析）",
                        api_definitions=api_defs_json,
                        module_tree=module_tree_json,
                        cross_module_relations=cross_text,
                    ),
                )
                api_text = result.content if hasattr(result, "content") else str(result)
            finally:
                _heartbeat_stop = True
                hb_task.cancel()
                try: await hb_task
                except _asyncio.CancelledError: pass

            # 写入 Step 3 结果
            with get_session_ctx() as session:
                from database.operations.analysis import AnalysisOps
                AnalysisOps.upsert_3step(
                    session, module_id, module_name,
                    api_analysis=api_text,
                )
                session.commit()
            await _update_task(task_id, progress=90,
                               message="Step 3/3 完成：接口映射已保存")
        else:
            logger.info("三步分析: 无 API 文档，跳过 Step 3")

        # ── 完成 ──
        resp = {
            "success": True,
            "module_name": module_name,
            "analysis": {
                "has_scenario": bool(scenario_text),
                "has_ui_flow": bool(ui_flow_text),
                "has_api": bool(api_text),
                "summary": (
                    f"{'场景' if scenario_text else ''}"
                    f"{' · ' if scenario_text and ui_flow_text else ''}"
                    f"{'交互逻辑' if ui_flow_text else ''}"
                    f"{' · ' if (scenario_text or ui_flow_text) and api_text else ''}"
                    f"{'接口映射' if api_text else ''}"
                ),
            },
        }
        await _update_task(task_id, status="completed", progress=100,
                           message="三步分析完成", result=resp)

    except Exception as e:
        logger.error("❌ 三步分析失败: %s", e)
        await _update_task(task_id, status="failed", error=str(e))


# ========================================================================
# Phase A: 模块场景分析（旧版 — 已废弃，保留兼容未升级的前端调用）
# ========================================================================

async def _analyze_module_scenarios_bg(task_id: str, module_name: str):
    """[已废弃] 后台任务：分析模块场景 → 生成 module_analysis JSON → 写入 SQLite。

    请使用 _analyze_module_scenarios_3step_bg() 替代。
    """
    # 委托新版三步分析
    await _analyze_module_scenarios_3step_bg(task_id, module_name)


async def _commit_apis_bg(task_id: str, file_path: str, module_name: str,
                           apis: list, delete_original: bool):
    """后台任务：逐条向量化入库，前端轮询进度。"""
    from web.state import _add_imported_file, _update_task
    from ingest_v2 import commit_api_docs

    set_trace_id(task_id)
    loop = asyncio.get_running_loop()
    total = len(apis)

    def _progress(pct: int, msg: str):
        asyncio.run_coroutine_threadsafe(
            _update_task(task_id, progress=pct, message=msg),
            loop,
        )

    try:
        await _update_task(task_id, status="running", progress=0,
                           message=f"开始入库 {total} 个接口...")
        result = await asyncio.to_thread(
            commit_api_docs, file_path, module_name, apis,
            delete_original=delete_original,
            progress_cb=lambda p, m: _progress(p, m),
        )
        # 更新内存文件列表
        from datetime import datetime
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for api in apis:
            try:
                await _add_imported_file({
                    "name": f"{api.get('method', '?')} {api.get('url', '')}",
                    "size": "—", "chunks": 1, "time": now_str, "type": "api",
                })
            except Exception:
                pass
        await _update_task(task_id, status="completed", progress=100,
                           message=f"✅ {total} 个接口入库完成", result=result)
    except Exception as e:
        logger.error("❌ 接口入库失败: %s", e)
        await _update_task(task_id, status="failed", error=str(e))


# ========================================================================
# 补偿 Worker（独立轮询线程）—— 已迁移至 web/compensation.py，此处 re-export 保持兼容
# ========================================================================
from web.compensation import (
    _start_compensation_worker,
    _stop_compensation_worker,
    _compensation_loop,
    _process_pending_compensation,
    _compensate_simple_summary,
    _compensate_chroma_rebuild,
    _compensate_api_search_text,
)
