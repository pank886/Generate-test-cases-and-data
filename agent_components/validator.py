"""只读校验节点。纯 Python 代码校验，无 LLM 调用。

职责：
  - Excel 文件层校验（表头、枚举值、必填字段）
  - .py 文件层校验（语法、命名规范、路径格式）

校验失败返回具体错误列表，供自动修复循环使用。
"""

import ast
import os
from typing import List, Tuple


ValidationResult = Tuple[bool, List[str]]
""" (passed: bool, errors: List[str]) """


# ==================== Excel 校验 ====================

VALID_ENABLED = {"Y", "N"}
EXPECTED_HEADERS_SHEET1 = [
    "@allure.epic", "@allure.feature", "@allure.story", "@allure.title",
    "fixture等级", "用例编号", "前置步骤", "执行步骤", "预期结果",
]
EXPECTED_HEADERS_SHEET2 = [
    "前置编号", "前置名称", "详细步骤", "预期结果", "关联用例",
]


def validate_excel_file(excel_path: str) -> ValidationResult:
    """校验 Excel 测试计划文件（双 Sheet）。"""
    errors = []

    if not os.path.exists(excel_path):
        return False, ["文件不存在: " + excel_path]

    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
    except Exception as e:
        return False, [f"无法打开 Excel 文件: {e}"]

    # Sheet 1
    ws1 = wb.active
    if ws1 is None:
        return False, ["Excel 文件为空"]
    h1 = [cell.value for cell in ws1[1]]
    for i, expected in enumerate(EXPECTED_HEADERS_SHEET1):
        if i >= len(h1) or h1[i] != expected:
            errors.append(f"Sheet1 表头第{i+1}列应为'{expected}'，实际为'{h1[i] if i < len(h1) else '缺失'}'")
    for row_idx, row in enumerate(ws1.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is None:
            continue
        if not str(row[4] or "").strip():
            errors.append(f"Sheet1 第{row_idx}行: fixture等级为空")
        for ci, cn in [(0, "epic"), (1, "feature"), (2, "story"), (3, "title"),
                        (7, "执行步骤")]:
            v = row[ci]
            if v is None or (isinstance(v, str) and not v.strip()):
                errors.append(f"Sheet1 第{row_idx}行: {cn}为空")

    # Sheet 2
    if "共享前置" in wb.sheetnames:
        ws2 = wb["共享前置"]
        h2 = [cell.value for cell in ws2[1]]
        for i, expected in enumerate(EXPECTED_HEADERS_SHEET2):
            if i >= len(h2) or h2[i] != expected:
                errors.append(f"Sheet2 表头第{i+1}列应为'{expected}'，实际为'{h2[i] if i < len(h2) else '缺失'}'")
        for row_idx, row in enumerate(ws2.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] is None:
                continue
            for ci, cn in [(0, "前置编号"), (1, "前置名称"), (2, "详细步骤"), (3, "预期结果")]:
                v = row[ci]
                if v is None or (isinstance(v, str) and not v.strip()):
                    errors.append(f"Sheet2 第{row_idx}行: {cn}为空")

    wb.close()
    return len(errors) == 0, errors


