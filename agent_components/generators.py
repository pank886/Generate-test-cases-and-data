"""Phase C: PY/YAML 生成节点 Mixin"""
import os
import re
import json
from collections import defaultdict
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

import config
from observability import get_logger
from prompts.response_model import ClassCode, TestData, TranslationResult

logger = get_logger(__name__)


# ============================================================
# Phase C Step 2: URL 路径参数归一化 + 接口过滤
# ============================================================

_PARAM_RE = re.compile(r'\{[^}]+\}')
_DOUBLE_PARAM_RE = re.compile(r'\{\{[^}]+\}\}')


def normalize_url(url: str) -> str:
    """将 URL 中所有 {xxx} 或 {{xxx}} 替换为 {param}，用于路径参数名无关匹配。

    /order/query/{order_id}    →  /order/query/{param}
    /order/query/{{order_id}}  →  /order/query/{param}
    /user/{userId}/profile     →  /user/{param}/profile
    /order/create              →  /order/create  (不变)
    """
    url = url.strip().rstrip("/")
    if not url.startswith("/"):
        url = "/" + url
    # 先处理双花括号（LLM 有时会输出 {{param}}），再处理单花括号
    url = _DOUBLE_PARAM_RE.sub("{param}", url)
    return _PARAM_RE.sub("{param}", url)


def build_api_index(api_defs: list[dict]) -> dict[tuple, list[dict]]:
    """构建 api_defs 查找索引。

    Returns:
        {(method_upper, normalized_url): [api_def_dict, ...]}
    一个 key 可对应多个 API 定义（路径参数名不同导致归一化后碰撞）。
    """
    index = defaultdict(list)
    for api in api_defs:
        key = (api.get("method", "").strip().upper(), normalize_url(api.get("url", "")))
        index[key].append(api)
    return dict(index)


def filter_apis_by_urls(
    api_index: dict,
    url_set: set[tuple[str, str]],  # {(method, url)} from api_sequences
) -> list[dict]:
    """用 URL 集合过滤接口定义，结果去重。

    策略：「宁可多匹配，不可漏匹配」— 碰撞时多引入 1-2 个定义。
    """
    seen: set[tuple[str, str]] = set()
    result: list[dict] = []
    for method, url in url_set:
        key = (method.upper(), normalize_url(url))
        for api in api_index.get(key, []):
            uid = (api.get("name", ""), api.get("url", ""))
            if uid not in seen:
                seen.add(uid)
                result.append(api)
    return result


def _collect_story_urls(story: dict) -> set[tuple[str, str]]:
    """从单个 story 的 dep_map 中收集所有 (method, url) 对。

    合并 story_pre_api_sequence + 所有 case 的 case_api_sequences + teardown。
    """
    def _parse_entry(entry: str) -> tuple[str, str] | None:
        if isinstance(entry, str) and ":" in entry:
            _, rest = entry.split(":", 1)
            parts = rest.strip().split()
            if len(parts) >= 2:
                return (parts[0].upper(), parts[1])
        return None

    urls: set[tuple[str, str]] = set()
    for entry in story.get("story_pre_api_sequence", []):
        parsed = _parse_entry(entry)
        if parsed:
            urls.add(parsed)
    for case_seqs in story.get("case_api_sequences", {}).values():
        for entry in case_seqs:
            parsed = _parse_entry(entry)
            if parsed:
                urls.add(parsed)
    for entry in story.get("teardown_api_sequence", []):
        parsed = _parse_entry(entry)
        if parsed:
            urls.add(parsed)
    return urls


_JSON_FENCE_RE = re.compile(r'```(?:json)?\s*\n?(.*?)\n?```', re.DOTALL)
_NO_LANG_FENCE_RE = re.compile(r'```\s*(\{.*?\})\s*```', re.DOTALL)


def _extract_json_from_thinking(text: str) -> dict:
    """从 LLM 输出中提取 JSON 对象。多层降级，由快到慢。

    L1 直接解析优先（DeepSeek response_format: json_object 绝大多数情况输出纯 JSON），
    L2/L3 仅作安全网。
    """
    # L1: 直接解析（response_format: json_object 下的最快路径）
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # L2: ```json ... ``` 代码块
    m = re.search(r'```json\s*(.*?)\s*```', text, re.DOTALL)
    if m:
        try:
            return json.loads(m.group(1))
        except json.JSONDecodeError:
            pass

    # L2b: ``` ... ``` 无语言标记的代码块
    m = _NO_LANG_FENCE_RE.search(text)
    if m:
        try:
            return json.loads(m.group(1))
        except json.JSONDecodeError:
            pass

    # L3: 全文查找第一个 { 到最后一个 }
    start = text.find("{")
    end = text.rfind("}")
    if start != -1 and end != -1 and end > start:
        return json.loads(text[start:end + 1])

    raise json.JSONDecodeError("无法从 LLM 输出中提取 JSON 对象", text, 0)


def _audit_variable_reads_writes(output_base: str) -> list[dict]:
    """后置变量读写审计：扫描全部 YAML 文件，交叉校验 extract 写入与 get_extract_data 读取。

    Returns:
        警告列表，每条含 yaml_path / check / severity / current / expected / fix_hint
    """
    import glob as _glob

    warnings: list[dict] = []
    write_map: dict[str, list[str]] = {}  # variable_name → [yaml_path, ...]
    read_map: dict[str, list[str]] = {}   # variable_name → [yaml_path, ...]

    yaml_files = _glob.glob(os.path.join(output_base, "**", "*.yaml"), recursive=True)
    if not yaml_files:
        return warnings

    extract_key_re = re.compile(r'^\s+(extract|input_extract):', re.MULTILINE)
    extract_entry_re = re.compile(r'^\s+(\w+):\s*(.+)$')
    get_extract_re = re.compile(r'\$\{get_extract_data\((\w+)\)\}')

    for yf in yaml_files:
        try:
            with open(yf, "r", encoding="utf-8") as f:
                content = f.read()
        except Exception:
            continue

        # 收集写入集（extract / input_extract 块中的 key）
        in_extract_block = False
        for line in content.split("\n"):
            if extract_key_re.match(line):
                in_extract_block = True
                continue
            if in_extract_block:
                if line and not line[0].isspace() and not line.startswith(" "):
                    in_extract_block = False
                    continue
                m = extract_entry_re.match(line)
                if m:
                    var_name = m.group(1)
                    write_map.setdefault(var_name, []).append(yf)

        # 收集读取集（${get_extract_data(xxx)} 中的 xxx）
        for m in get_extract_re.finditer(content):
            var_name = m.group(1)
            read_map.setdefault(var_name, []).append(yf)

    # 交叉检查
    all_vars = set(write_map.keys()) | set(read_map.keys())
    for var in sorted(all_vars):
        writers = write_map.get(var, [])
        readers = read_map.get(var, [])
        if writers and not readers:
            warnings.append({
                "check": "extract_write_never_read",
                "yaml_path": writers[0],
                "severity": "P2",
                "current": f"extract: {var} 已写入但从未被 ${{get_extract_data({var})}} 读取",
                "expected": f"${var} 应有下游用例通过 ${{get_extract_data({var})}} 消费，"
                            f"或移除多余的 extract",
                "fix_hint": f"检查 {var} 是否确实需要提取；若不需要则删除 extract 条目",
                "related_files": writers,
            })
        if readers and not writers:
            warnings.append({
                "check": "get_extract_data_read_never_written",
                "yaml_path": readers[0],
                "severity": "P1",
                "current": f"${{get_extract_data({var})}} 读取但 {var} 从未被任何 extract 写入",
                "expected": f"需要在前置步骤中添加 extract: {var}: $.jsonpath",
                "fix_hint": f"在产出 {var} 的步骤 YAML 中添加 extract 块，"
                            f"或在 internal_dependency 中标注 extract_path",
                "related_files": readers,
            })

    return warnings


def _append_generation_errors(output_base: str, story_name: str,
                               errors: list[dict]) -> str:
    """追加写入 _generation_errors.json（每 story 一个 section，不覆盖）。

    Args:
        output_base: 输出目录
        story_name: story 名称（作为 section key）
        errors: 错误条目列表，每条含 case_id / error / round 等

    Returns:
        errors 文件路径
    """
    errors_path = os.path.join(output_base, "_generation_errors.json")
    existing: dict = {}
    if os.path.exists(errors_path):
        try:
            with open(errors_path, "r", encoding="utf-8") as f:
                existing = json.load(f)
        except (json.JSONDecodeError, OSError):
            existing = {}
    existing[story_name] = errors
    tmp = errors_path + ".tmp"
    try:
        os.remove(tmp)
    except OSError:
        pass
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)
    os.replace(tmp, errors_path)
    return errors_path


# ---- Phase C 修复循环辅助（错误分类关键词与校验器报错文案对齐）----

_ERROR_PATTERN_LABELS = [
    ("B1 双花括号 {{}}", ("双花括号",)),
    ("B2 占位符运算/拼接/未闭合", ("禁止运算/拼接", "未闭合或嵌套")),
    ("B3 非注册表占位符函数", ("未知占位符函数",)),
    ("B4 占位符实参不合规", ("实参个数", "第1个参数仅支持")),
    ("B5/B10 提取字段值须为字符串(无需提取应省略)", ("Input should be a valid string",)),
    ("B6/B7 空列表输出", ("at least 1 item", "too_short")),
    ("B9 json/params/data 并存", ("三选一",)),
]


def _summarize_error_patterns(failures: list) -> str:
    """按 B 类别聚合本轮错误计数（跨文件模式反馈，注入修复轮 prompt）。"""
    counts: dict = {}
    for f in failures:
        err = f.get("error", "")
        matched = False
        for label, keywords in _ERROR_PATTERN_LABELS:
            if any(kw in err for kw in keywords):
                counts[label] = counts.get(label, 0) + 1
                matched = True
        if not matched:
            counts["B8 结构解析失败(缺字段/类型错/JSON坏)"] = \
                counts.get("B8 结构解析失败(缺字段/类型错/JSON坏)", 0) + 1
    if not counts:
        return "（无统计）"
    return "\n".join(f"- {label}: {n} 处" for label, n in counts.items())


def _extract_completion_snippet(err_text: str, limit: int = 500) -> str:
    """从结构化输出异常文本中截取 LLM 原始 completion 片段（修复轮自查材料）。"""
    m = re.search(r"from completion (.+?)(?:\. Got:|$)", err_text, re.DOTALL)
    snippet = m.group(1) if m else err_text
    return snippet[:limit]


def _write_fail_detail(output_base: str, pid: str, case_id: str,
                       yaml_path: str, round_no: int, err_text: str,
                       raw_snippet: str) -> None:
    """单文件生成失败时，将原文和错误点写入详细日志。"""
    import os as _os
    log_path = _os.path.join(output_base, "_generation_error_details.log")
    # 提取 Pydantic 校验错误（去掉 LangChain 框架包装）
    _parsed_err = err_text
    _m = re.search(r"Got: (.+?)(?:\nFor troubleshooting|$)", err_text, re.DOTALL)
    if _m:
        _parsed_err = _m.group(1).strip()
    with open(log_path, "a", encoding="utf-8") as _f:
        _f.write(f"{'=' * 60}\n")
        _f.write(f"[{pid}] ROUND={round_no} | {case_id} | {yaml_path}\n")
        _f.write(f"{'=' * 60}\n")
        _f.write(f"--- 校验错误 ---\n{_parsed_err[:3000]}\n\n")
        _f.write(f"--- LLM 原始输出 (前 2000 字符) ---\n{raw_snippet[:2000]}\n\n")


def _format_post_issues_for_prompt(issues: list | None) -> str:
    """将后校验问题列表格式化为修复轮 prompt 文本。"""
    if not issues:
        return ""
    lines = ["⚠️ YAML 后校验发现问题（请逐条修正）：", ""]
    for i, iss in enumerate(issues, 1):
        lines.append(f"{i}. [{iss['check']}] {iss['yaml_path']}")
        lines.append(f"   当前: {iss['current']}")
        lines.append(f"   期望: {iss['expected']}")
        lines.append(f"   修复指引: {iss['fix_hint']}")
        lines.append("")
    return "\n".join(lines)


# ============================================================
# Phase C: JSON 骨架生成 + 修复诊断包
# ============================================================

def generate_json_skeleton(model_class: type) -> str:
    """从 Pydantic 模型生成 JSON 骨架，用于注入 prompt。

    递归规则：
      - object → {key: <recursive>, ...}
      - $ref → 解析到 $defs 中的对应定义
      - Dict[str, Any] 且含 json_schema_extra["example_keys"] → 展开示例键
      - Optional[object] / Optional[array] → 跳过不渲染（§5.10）
      - array  → [<recursive_element>]
      - string → ""
      - integer → 0
      - boolean → false
      - 其他   → null
    """
    try:
        schema = model_class.model_json_schema()
    except Exception:
        return "{}"

    _defs = schema.get("$defs", {})

    def _resolve(prop: dict) -> dict:
        """解析 $ref 引用，返回实际定义。"""
        ref = prop.get("$ref", "")
        if ref.startswith("#/$defs/"):
            name = ref[len("#/$defs/"):]
            return _defs.get(name, prop)
        # 处理 anyOf（Optional 字段用 anyOf: [{$ref: ...}, {type: null}]）
        any_of = prop.get("anyOf")
        if any_of:
            for opt in any_of:
                r = _resolve(opt)
                if r.get("type") != "null":
                    return r
        return prop

    def _is_optional(prop: dict) -> bool:
        """判断字段是否 Optional（anyOf 中含 null）。"""
        if "anyOf" in prop:
            return any(o.get("type") == "null" for o in prop["anyOf"])
        if isinstance(prop.get("type"), list):
            return "null" in prop["type"]
        return False

    def _build(prop: dict) -> Any:
        # Optional[object/list] 且无 example_keys → 跳过不渲染（§5.10）
        # example_keys 在原始 prop 上（anyOf 外层），不在 _resolve() 后的内层
        if _is_optional(prop):
            resolved = _resolve(prop)
            rtype = resolved.get("type")
            if isinstance(rtype, list):
                rtype = next((t for t in rtype if t != "null"), None)
            if rtype in ("object", "array"):
                # 有 example_keys 的 Optional[Dict] → 仍渲染为 {}（如 json/params）
                ex = prop.get("example_keys")
                if ex is None:
                    return None

        prop = _resolve(prop)
        ptype = prop.get("type")

        if isinstance(ptype, list):
            ptype = next((t for t in ptype if t != "null"), ptype[0] if ptype else None)

        if ptype == "object":
            props = prop.get("properties", {})
            if not props:
                # Dict[str, Any] → 查 example_keys（Pydantic 展平到顶层）
                # example value 决定渲染类型：str→"", dict→{}, list→[], 其他→null
                ex = prop.get("example_keys")
                if ex is not None:
                    result = {}
                    for k, v in ex.items():
                        if isinstance(v, dict):
                            result[k] = {}
                        elif isinstance(v, list):
                            result[k] = []
                        elif isinstance(v, str):
                            result[k] = ""
                        else:
                            result[k] = _build({"type": "string"})
                    return result
                return {}
            # 构建后剔除 None 值（Optional 字段已返回 None）
            result = {k: _build(v) for k, v in props.items()}
            return {k: v for k, v in result.items() if v is not None}

        if ptype == "array":
            items = prop.get("items", {})
            items = _resolve(items)
            return [_build(items)]

        if ptype == "string":
            return ""
        if ptype == "integer" or ptype == "number":
            return 0
        if ptype == "boolean":
            return False
        return None

    skeleton = _build(schema)
    return json.dumps(skeleton, indent=2, ensure_ascii=False)


class RepairNeeded(Exception):
    """YAML 生成校验失败时抛出的自定义异常，携带修复诊断包。

    repair_ctx 由 prepare_repair_context() 生成，包含:
      - failed_yaml: 解析后的 YAML 文本（JSON 合法时）或 raw_text（JSON 非法时）
      - error_roadmap: 精简错误路径摘要
      - raw_text: LLM 原始输出全文
    """

    def __init__(self, repair_ctx: dict):
        super().__init__(repair_ctx.get("error_roadmap", "校验失败"))
        self.repair_ctx = repair_ctx


def prepare_repair_context(raw_text: str, error: Exception) -> dict:
    """从 raw text + 异常构造修复诊断包。

    处理三种异常类型：
      - pydantic.ValidationError → 解析 JSON → YAML + 错误路径摘要
      - json.JSONDecodeError       → 直接用 raw_text 作为 failed_yaml
      - 其他 Exception             → {"failed_yaml": raw_text, "error_roadmap": str(error)}
    """
    from json import JSONDecodeError

    try:
        from pydantic import ValidationError
        _has_pydantic = True
    except ImportError:
        _has_pydantic = False

    if _has_pydantic and isinstance(error, ValidationError):
        try:
            parsed_dict = json.loads(raw_text)
            failed_yaml = yaml.dump(parsed_dict, allow_unicode=True, indent=2)
        except Exception:
            failed_yaml = raw_text

        error_lines = []
        for err in error.errors():
            path = " -> ".join(str(p) for p in err["loc"])
            error_lines.append(f"  - [{path}] {err['msg']}")
        error_roadmap = "\n".join(error_lines)

    elif isinstance(error, JSONDecodeError):
        failed_yaml = raw_text
        error_roadmap = (
            f"JSON 解析失败（第 {error.lineno} 行第 {error.colno} 列）: {error.msg}"
        )

    else:
        failed_yaml = raw_text
        error_roadmap = str(error)[:500]

    return {
        "failed_yaml": failed_yaml,
        "error_roadmap": error_roadmap,
        "raw_text": raw_text,
    }


def _error_mentions_placeholder(roadmap: str) -> bool:
    """错误是否涉及占位符/工厂方法。"""
    return any(kw in roadmap for kw in (
        "占位符", "get_extract_data", "${", "factory", "unknown function",
    ))


def _error_mentions_api(roadmap: str) -> bool:
    """错误是否涉及 API 匹配（url/method/参数）。"""
    return any(kw in roadmap for kw in (
        "url", "method", "api_name", "baseInfo", "header", "参数",
    ))


class GenerationMixin:
    """PY/YAML 测试文件生成节点"""

    # ==================== Phase C Step 0: dependency_map.json 生成 ====================

    def _generate_dependency_map(self, excel_path: str, output_dir: str,
                                  api_defs_json: str, module_tree_json: str,
                                  product_docs_json: str, label: str = "",
                                  user_ctx: str = "") -> str:
        """Phase C Step 0：调用 LLM（thinking 模式）生成 dependency_map.json。

        流程：
          1. 组装 prompt（Excel 全量用例 + api_defs + 产品文档 + 工厂方法 + 模块树）
          2. 调用 LLM thinking → 自由文本 + JSON 输出
          3. json.loads() 解析 → Pydantic 校验（DependencyMap）
          4. 校验失败 → 修复重试（DEPENDENCY_REPAIR_ATTEMPTS 次）
          5. 校验通过 → .tmp 原子写入 → os.replace 落盘

        Returns:
            dependency_map.json 的绝对路径
        """
        from prompts.extraction_prompts import generate_dependency_map_prompt, _MsgBuilder
        from prompts.response_model import DependencyMap
        from observability import log_thinking, get_thinking_logger
        from langchain_core.messages import SystemMessage, HumanMessage

        # 1. 读取 Excel 全量用例
        excel_rows = self._read_excel_rows(excel_path)
        shared_pres = self._read_shared_preconditions(excel_path)

        # 2. 工厂方法
        factory_methods_text = self._load_factory_methods()

        # 3. 系统消息来自 generate_dependency_map_prompt()（纯字符串），
        #    手动注入 {factory_methods}（绕过 LangChain 花括号校验）
        system_template = generate_dependency_map_prompt()
        system_msg_text = system_template.replace("{factory_methods}", factory_methods_text)

        # 4. LLM thinking 调用 + 修复重试
        dep_map_path = os.path.join(output_dir, "dependency_map.json")
        tlog = get_thinking_logger()
        max_attempts = max(1, config.DEPENDENCY_REPAIR_ATTEMPTS)

        last_error = None
        for attempt in range(1, max_attempts + 1):
            logger.info("   🧠 dep_map 生成 第 %d/%d 次...", attempt, max_attempts)

            llm_kwargs = {"extra_body": {"thinking": {"type": "enabled"}}}
            if config.THINKING_TIMEOUT:
                llm_kwargs["timeout"] = config.THINKING_TIMEOUT
            bound_llm = self.llm.bind(**llm_kwargs)

            # 使用 _MsgBuilder 构建人类消息（多段上下文拼接）
            user_ctx_text = user_ctx or label
            if attempt > 1 and last_error:
                user_ctx_text = (
                    str(user_ctx or label) + "\n\n"
                    "⚠️ 上一轮校验失败，错误信息：\n" + str(last_error) + "\n"
                    "请仔细对照 JSON Schema 修正。"
                )

            excel_builder = _MsgBuilder()
            if shared_pres:
                pre_lines = "\n".join(
                    f"{p['id']}: {p['name']}\n  步骤: {p['steps']}\n  预期: {p.get('expected', '')}"
                    for p in shared_pres
                )
                excel_builder.add("共享前置 (Sheet2)", pre_lines)
            case_lines = "\n".join(
                f"{r['case_id']} | story={r['story']} | title={r['title']}\n"
                f"  preconditions={r.get('preconditions', '')}\n"
                f"  steps={r['steps']}\n"
                f"  expected={r.get('expected', '')}"
                for r in excel_rows
            )
            excel_builder.add("测试用例 (Sheet1)", case_lines)
            excel_cases_text = excel_builder.build()

            human_msg = (_MsgBuilder()
                .add("用户需求", user_ctx_text)
                .add("模块树", module_tree_json)
                .add("接口定义", api_defs_json)
                .add("产品文档", product_docs_json)
                .add("Excel 测试计划（全量用例）", excel_cases_text)
                .build()
            )
            human_msg += "\n\n请分析以上信息，生成 dependency_map JSON："

            messages = [
                SystemMessage(content=system_msg_text),
                HumanMessage(content=human_msg),
            ]
            result = bound_llm.invoke(messages)
            raw_text = result.content if hasattr(result, "content") else str(result)

            # 落 thinking trace
            log_thinking(
                f"generate_dependency_map_attempt{attempt}",
                f"excel={os.path.basename(excel_path)}",
                raw_text,
                prompt_label="generate_dependency_map_prompt",
            )

            # 5. 从 thinking 输出中提取 JSON
            try:
                parsed = _extract_json_from_thinking(raw_text)
            except json.JSONDecodeError as e:
                last_error = f"JSON 解析失败（第{attempt}次）: {e}"
                logger.warning("   ⚠️ %s", last_error)
                continue

            # 6. Pydantic 校验
            try:
                dep_map = DependencyMap.model_validate(parsed)
            except Exception as e:
                last_error = f"Pydantic 校验失败（第{attempt}次）: {e}"
                logger.warning("   ⚠️ %s", last_error)
                continue

            # 7. 外部校验：case_id 是否在 Excel 中存在
            excel_case_ids = {r["case_id"] for r in excel_rows}
            missing_ids = []
            for story in dep_map.stories:
                for cid in story.case_api_sequences:
                    if cid not in excel_case_ids:
                        missing_ids.append(cid)
            if missing_ids:
                last_error = (
                    f"case_id 校验失败（第{attempt}次）: "
                    f"以下 case_id 在 Excel 中不存在: {missing_ids}"
                )
                logger.warning("   ⚠️ %s", last_error)
                continue

            # 8. 原子写入
            tmp_path = dep_map_path + ".tmp"
            try:
                os.remove(tmp_path)
            except OSError:
                pass
            with open(tmp_path, "w", encoding="utf-8") as f:
                json.dump(dep_map.model_dump(), f, ensure_ascii=False, indent=2)
            os.replace(tmp_path, dep_map_path)

            logger.info("   ✅ dependency_map.json 已生成: %d 个 story", len(dep_map.stories))
            return dep_map_path

        # 所有重试耗尽
        raise RuntimeError(
            "dependency_map.json 生成失败（" + str(max_attempts) + " 次尝试后仍失败）。"
            "最后错误: " + str(last_error)
        )

    @staticmethod
    def _load_dependency_map(excel_path: str) -> dict:
        """从 Excel 同级目录加载 dependency_map.json 并返回解析后的 dict。

        Raises:
            FileNotFoundError: 文件不存在
            json.JSONDecodeError: JSON 解析失败
            ValueError: stories 为空
        """
        output_dir = os.path.dirname(excel_path)
        dep_path = os.path.join(output_dir, "dependency_map.json")
        if not os.path.exists(dep_path):
            raise FileNotFoundError(f"dependency_map.json 不存在: {dep_path}")
        with open(dep_path, "r", encoding="utf-8") as f:
            dep_map = json.load(f)
        if not dep_map.get("stories"):
            raise ValueError("dependency_map.json 中 stories 为空")
        return dep_map

    # ==================== Phase C: 数据依赖分析（原有） ====================

    def _analyze_data_deps(self, case_steps: str, api_defs_json: str,
                           user_ctx: str) -> str:
        """数据依赖分析（thinking on，自由文本）。"""
        from prompts.extraction_prompts import analyze_data_deps_prompt

        from observability import log_phase_header
        log_phase_header("Phase C — 数据依赖分析")
        logger.info("\n🧠 分析数据依赖（深度思考）...")
        prompt = analyze_data_deps_prompt()
        llm_kwargs = {}
        if config.ENABLE_THINKING:
            llm_kwargs["extra_body"] = {"thinking": {"type": "enabled"}}
        else:
            llm_kwargs["extra_body"] = {"thinking": {"type": "disabled"}}
        bound_llm = self.llm.bind(**llm_kwargs)
        result = bound_llm.invoke(prompt.format_messages(
            api_definitions=api_defs_json,
            test_case_steps=case_steps,
            user_context=user_ctx,
        ))
        analysis = result.content if hasattr(result, "content") else str(result)
        logger.info(f"   => 数据依赖分析完成（{len(analysis)} 字符）")
        from observability import log_thinking
        log_thinking("analyze_data_deps", user_ctx, analysis, prompt_label="analyze_data_deps_prompt")
        return analysis

    def _format_data_plan(self, data_analysis: str, case_steps: str,
                          api_defs_json: str, user_ctx: str) -> dict:
        """格式化数据规划（thinking off + json_mode）。"""
        from prompts.extraction_prompts import generate_data_plan_prompt
        from prompts.response_model import DataPlan

        logger.info("\n--- 生成结构化数据规划 ---")
        prompt = generate_data_plan_prompt()
        result = self._invoke_structured(prompt, DataPlan,
            method="json_mode",
            data_analysis=data_analysis,
            api_definitions=api_defs_json,
            test_case_steps=case_steps,
            user_context=user_ctx,
        )
        if isinstance(result, list):
            result = DataPlan(steps=result, scenario_name="")
        logger.info(f"   => 数据规划完成: {len(result.steps)} 步")
        return {"data_plan": result.model_dump()}

    @staticmethod
    def _read_excel_rows(excel_path: str, enabled_only: bool = False) -> list[dict]:
        """读取 Excel V2 测试计划（9 列双 Sheet），返回 dict 列表。

        Sheet1 列: @allure.epic, @allure.feature, @allure.story, @allure.title,
                   fixture等级, 用例编号, 前置步骤, 执行步骤, 预期结果
        Sheet2: 共享前置（由 _read_shared_preconditions 独立读取）
        """
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        try:
            ws = wb.active  # Sheet1: 测试计划
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                rows.append({
                    "epic": row[0],          # @allure.epic
                    "feature": row[1],       # @allure.feature
                    "story": row[2],         # @allure.story
                    "title": row[3],         # @allure.title
                    "fixture_level": row[4], # fixture等级
                    "case_id": row[5],       # 用例编号 TC-xxx
                    "preconditions": row[6], # 前置步骤
                    "steps": row[7],         # 执行步骤
                    "expected": row[8],      # 预期结果
                })
            return rows
        finally:
            wb.close()

    @staticmethod
    def _read_shared_preconditions(excel_path: str) -> list[dict]:
        """读取 Excel V2 Sheet2（共享前置），返回 dict 列表。

        Sheet2 列: 前置编号, 前置名称, 详细步骤, 预期结果, 关联用例
        """
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        try:
            if "共享前置" not in wb.sheetnames:
                return []
            ws = wb["共享前置"]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                rows.append({
                    "id": row[0],           # 前置编号 PRE-xxx
                    "name": row[1],          # 前置名称
                    "steps": row[2],         # 详细步骤
                    "expected": row[3],      # 预期结果
                    "linked_cases": row[4],  # 关联用例（逗号分隔）
                })
            return rows
        finally:
            wb.close()

    # ==================== C4: 英文翻译 + C4-1: 幂等性保障 ====================

    @staticmethod
    def _sanitize_en(name: str) -> str:
        """LLM 输出后强制清洗，确保合法 Python identifier。"""
        sanitized = re.sub(r'[^a-zA-Z0-9_]', '', name.replace(' ', '_'))
        if not sanitized or sanitized[0].isdigit():
            sanitized = '_' + sanitized
        return sanitized

    @staticmethod
    def _load_translation_cache(excel_path: str) -> dict:
        """从 Excel 同级目录读取翻译缓存。"""
        cache_path = os.path.join(os.path.dirname(excel_path), "translation_cache.json")
        if os.path.exists(cache_path):
            try:
                with open(cache_path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                logger.warning("读取翻译缓存失败: %s", cache_path, exc_info=True)
        return {}

    @staticmethod
    def _save_translation_cache(excel_path: str, cache: dict) -> None:
        """保存翻译缓存到 Excel 同级目录。"""
        cache_path = os.path.join(os.path.dirname(excel_path), "translation_cache.json")
        try:
            with open(cache_path, "w", encoding="utf-8") as f:
                json.dump(cache, f, ensure_ascii=False, indent=2)
        except Exception:
            logger.warning("保存翻译缓存失败: %s", cache_path, exc_info=True)

    @staticmethod
    def _pinyin_fallback(text: str) -> str:
        """拼音首字母缩写 Fallback（LLM 翻译失败时使用）。"""
        try:
            from pypinyin import lazy_pinyin
            return ''.join(w[0].upper() for w in lazy_pinyin(text) if w)
        except ImportError:
            # pypinyin 未安装时用简单 hash 兜底
            import hashlib
            return 'M' + hashlib.md5(text.encode()).hexdigest()[:7].upper()

    def _translate_to_en(self, excel_path: str, rows: list[dict]) -> dict:
        """批量翻译 feature/story/title 为英文，带缓存 + sanitize + 降级。

        Returns:
            {"feature_en": {中文: 英文}, "story_en": {...}, "title_en": {...}}
        """
        # 收集待翻译文本
        features = list(dict.fromkeys(r["feature"] for r in rows if r.get("feature")))
        stories = list(dict.fromkeys(r["story"] for r in rows if r.get("story")))
        titles = list(dict.fromkeys(r["title"] for r in rows if r.get("title")))

        # 查缓存
        cache = self._load_translation_cache(excel_path)
        cache_fe = cache.get("feature_en", {})
        cache_st = cache.get("story_en", {})
        cache_ti = cache.get("title_en", {})

        uncached_fe = [f for f in features if f not in cache_fe]
        uncached_st = [s for s in stories if s not in cache_st]
        uncached_ti = [t for t in titles if t not in cache_ti]

        all_uncached = uncached_fe + uncached_st + uncached_ti

        if all_uncached:
            logger.info("\n🌐 翻译 %d 条中文标识符...", len(all_uncached))
            try:
                from prompts.extraction_prompts import translate_to_en_prompt
                prompt = translate_to_en_prompt()
                result = self._invoke_structured(prompt, TranslationResult,
                    method="json_mode",
                    features=json.dumps(uncached_fe, ensure_ascii=False),
                    stories=json.dumps(uncached_st, ensure_ascii=False),
                    titles=json.dumps(uncached_ti, ensure_ascii=False),
                )
            except Exception as e:
                logger.warning("LLM 翻译失败，全部使用拼音 Fallback: %s", e)
                result = None

            if result and isinstance(result, TranslationResult):
                for cn, en in result.feature_en.items():
                    cache_fe[cn] = self._sanitize_en(en)
                for cn, en in result.story_en.items():
                    cache_st[cn] = self._sanitize_en(en)
                for cn, en in result.title_en.items():
                    cache_ti[cn] = self._sanitize_en(en)

            # 拼音 Fallback：LLM 未返回或翻译缺失的条目
            for cn in uncached_fe:
                if cn not in cache_fe:
                    cache_fe[cn] = self._sanitize_en(self._pinyin_fallback(cn))
                    logger.warning("拼音 Fallback: feature '%s' → '%s'", cn, cache_fe[cn])
            for cn in uncached_st:
                if cn not in cache_st:
                    cache_st[cn] = self._sanitize_en(self._pinyin_fallback(cn))
                    logger.warning("拼音 Fallback: story '%s' → '%s'", cn, cache_st[cn])
            for cn in uncached_ti:
                if cn not in cache_ti:
                    cache_ti[cn] = self._sanitize_en(self._pinyin_fallback(cn))
                    logger.warning("拼音 Fallback: title '%s' → '%s'", cn, cache_ti[cn])

            # 存缓存
            cache["feature_en"] = cache_fe
            cache["story_en"] = cache_st
            cache["title_en"] = cache_ti
            self._save_translation_cache(excel_path, cache)

        return {
            "feature_en": cache_fe,
            "story_en": cache_st,
            "title_en": cache_ti,
        }

    # ==================== C6-1: 断言关键词解析 ====================

    class AssertionParseError(ValueError):
        """断言格式校验异常。"""

    _ASSERTION_PATTERN = re.compile(r'\[(eq|contains|ne|db)\]', re.IGNORECASE)
    _ASSERTION_INVALID_SPACE = re.compile(
        r'\[\s+(eq|contains|ne|db)\s*\]|\[\s*(eq|contains|ne|db)\s+\]',
        re.IGNORECASE,
    )  # 仅当关键词两侧至少有一处空格时命中

    @classmethod
    def _parse_assertion(cls, expected_text: str) -> tuple[str, str]:
        """从预期结果文本解析断言关键词。返回 (keyword_lower, rest_of_text)。

        Raises:
            AssertionParseError: 格式非法时抛出。
        """
        if re.search(r'\[\[|\]\]', expected_text):
            raise cls.AssertionParseError(f"断言格式非法（双层括号）: {expected_text[:60]}")
        if cls._ASSERTION_INVALID_SPACE.search(expected_text):
            raise cls.AssertionParseError(f"断言关键词含空格: {expected_text[:60]}")
        m = cls._ASSERTION_PATTERN.search(expected_text)
        if not m:
            raise cls.AssertionParseError(f"未找到断言关键词 [eq/contains/ne/db]: {expected_text[:60]}")
        keyword = m.group(1).lower()
        rest = expected_text[m.end():].strip()
        # 同一步骤含多个关键词 → 非法（框架每步只认一个断言类型）
        if cls._ASSERTION_PATTERN.search(rest):
            raise cls.AssertionParseError(
                f"同一步骤含多个断言关键词，每步只能有一个: {expected_text[:60]}")
        return keyword, rest

    def _generate_py_file(self, excel_path: str, project_name: str = None,
                          dep_map: dict | None = None) -> dict:
        """Phase C V2: 按 feature 生成 .py 文件，fixture + parametrize 结构。

        同一 feature → 一个 .py 文件
        同一 story   → 一个 class（含 fixture + test functions）

        dep_map 不为 None 时，检查 teardown_api_sequence 决定是否生成 teardown 引用。
        """
        logger.info("\n🐍 正在生成 Python 测试文件...")

        if not excel_path:
            logger.info("   ⚠️ 无 Excel 路径，跳过 .py 生成")
            return {"py_path": "", "py_file_name": "", "modules": 0, "cases": 0}

        from collections import defaultdict
        expanded_rows = self._read_excel_rows(excel_path)

        if not expanded_rows:
            raise ValueError("Excel 中无数据")

        # C4: 英文翻译
        translations = self._translate_to_en(excel_path, expanded_rows)
        feature_en_map = translations["feature_en"]
        story_en_map = translations["story_en"]
        title_en_map = translations["title_en"]

        # C5: 读取共享前置（Sheet2）
        shared_pres = self._read_shared_preconditions(excel_path)
        pre_by_id = {p["id"]: p for p in shared_pres}

        # V2: 从 dep_map 提取需要 teardown 的 story 名集合
        stories_with_teardown: set[str] = set()
        if dep_map:
            for s in dep_map.get("stories", []):
                if s.get("teardown_api_sequence"):
                    stories_with_teardown.add(s.get("story_name", ""))

        # 按 feature → story → cases 分组
        features = defaultdict(lambda: defaultdict(list))
        for r in expanded_rows:
            features[r["feature"]][r["story"]].append(r)

        import_header = (
            "import pytest\n"
            "import allure\n"
            "from common.readyaml import ReadYamlData, get_testcase_yaml\n"
            "from common.sendrequests import SendRequests\n"
            "from common.recordlog import logs\n"
            "from base.apiutil import RequestsBase\n"
        )

        output_base = os.path.dirname(excel_path)
        total_modules = 0
        total_cases = 0
        py_files = []

        for feature_cn, stories in features.items():
            feature_en = feature_en_map.get(feature_cn, self._sanitize_en(self._pinyin_fallback(feature_cn)))
            feature_dir = os.path.join(output_base, feature_en)
            os.makedirs(feature_dir, exist_ok=True)

            # __init__.py
            init_path = os.path.join(feature_dir, "__init__.py")
            if not os.path.exists(init_path):
                with open(init_path, "w", encoding="utf-8") as f:
                    f.write("# auto-generated\n")

            class_blocks = []
            for story_cn, cases in stories.items():
                story_en = story_en_map.get(story_cn, self._sanitize_en(self._pinyin_fallback(story_cn)))
                class_slug = re.sub(r'(?<!^)(?=[A-Z])', '_', story_en).lower()
                total_modules += 1

                # 收集该 story 的共享前置引用
                pre_ids = set()
                for c in cases:
                    pre_str = c.get("preconditions", "")
                    if pre_str and pre_str != "无":
                        for pid in pre_str.split(","):
                            pid = pid.strip()
                            if pid.startswith("PRE-"):
                                pre_ids.add(pid)

                # 生成 fixture
                fixture_code = ""
                if pre_ids:
                    has_teardown = dep_map is None or story_cn in stories_with_teardown
                    if has_teardown:
                        fixture_code = (
                            f'\n@pytest.fixture(scope="class")\n'
                            f'def setup_{class_slug}():\n'
                            f'    read = ReadYamlData()\n'
                            f'    base = RequestsBase()\n'
                            f'    base.specification_yaml(get_testcase_yaml(\n'
                            f'        \'./testcase/{feature_en}/setup_data/setup_{class_slug}.yaml\'))\n'
                            f'    yield\n'
                            f'    base.specification_yaml(get_testcase_yaml(\n'
                            f'        \'./testcase/{feature_en}/setup_data/teardown_{class_slug}.yaml\'))\n'
                        )
                    else:
                        # dep_map 判定无需清理 → fixture 不引用 teardown
                        fixture_code = (
                            f'\n@pytest.fixture(scope="class")\n'
                            f'def setup_{class_slug}():\n'
                            f'    read = ReadYamlData()\n'
                            f'    base = RequestsBase()\n'
                            f'    base.specification_yaml(get_testcase_yaml(\n'
                            f'        \'./testcase/{feature_en}/setup_data/setup_{class_slug}.yaml\'))\n'
                            f'    yield\n'
                        )
                else:
                    fixture_code = (
                        f'\n@pytest.fixture(scope="class")\n'
                        f'def setup_{class_slug}():\n'
                        f'    pass\n'
                        f'    yield\n'
                    )

                # 生成 test functions — run_blocks 加载单个 YAML（含所有 step）
                func_lines = []
                for i, c in enumerate(cases, 1):
                    title_cn = c["title"]
                    func_en = title_en_map.get(
                        title_cn,
                        "test_" + self._sanitize_en(self._pinyin_fallback(title_cn))
                    )
                    if not func_en.startswith("test_"):
                        func_en = "test_" + func_en
                    total_cases += 1

                    func_lines.append(
                        f'    @allure.title(\'{title_cn}\')\n'
                        f'    @pytest.mark.order({i})\n'
                        f'    def {func_en}(self):\n'
                        f'        RequestsBase().run_blocks(\n'
                        f'            \'./testcase/{feature_en}/{func_en}/test_data.yaml\')\n'
                    )

                # 组装 class
                usefixtures = f'\n@pytest.mark.usefixtures("setup_{class_slug}")' if pre_ids else ''
                class_code = (
                    f'{fixture_code}\n'
                    f'@allure.story(\'{story_cn}\')\n'
                    f'@pytest.mark.danyuan'
                    f'{usefixtures}\n'
                    f'class Test{story_en}:\n'
                    + '\n'.join(func_lines)
                )
                class_blocks.append(class_code)

            # 写 .py 文件
            file_name = f"test_{feature_en}.py"
            full_content = import_header + "\n" + "\n".join(class_blocks)
            py_path = os.path.join(feature_dir, file_name)
            tmp_path = py_path + ".tmp"
            try:
                os.remove(tmp_path)
            except OSError:
                pass
            with open(tmp_path, "w", encoding="utf-8", newline="\r\n") as f:
                f.write(full_content)
            os.replace(tmp_path, py_path)
            py_files.append(py_path)
            logger.info(f"   📄 {file_name} ({len(stories)} classes, {sum(len(v) for v in stories.values())} cases)")

        logger.info(f"   📦 {len(py_files)} 个 .py 文件, {total_modules} 个 class, {total_cases} 条用例")

        result = {
            "py_path": py_files[0] if py_files else "",
            "py_file_name": ", ".join(os.path.basename(p) for p in py_files),
            "modules": total_modules,
            "cases": total_cases,
        }
        self._log_node_output("generate_py_file", result)
        return result

    def _generate_one_yaml(self, row: dict, api_defs_json: str, user_ctx: str,
                           output_path: str, skeleton_text: str,
                           repair_ctx: dict | None = None) -> str:
        """Phase C YAML 生成：thinking 分析 → json_mode 填表 → Pydantic 校验。

        thinking / json_mode 共用同一个 skeleton 字符串，由 _generate_all_yamls 入口生成一次。

        校验失败 → prepare_repair_context() 构造诊断包 → 抛 RepairNeeded，
        由 _run_yaml_rounds 捕获后进入修复轮。
        """
        from prompts.extraction_prompts import (
            analyze_yaml_data_prompt, format_yaml_data_prompt, repair_yaml_data_prompt,
        )
        from prompts.response_model import TestData as TestDataModel
        from observability import log_thinking

        factory_methods_text = self._load_factory_methods()
        test_case_logic = f"执行步骤: {row['steps']}\n预期结果: {row.get('expected', '')}"

        # 注入 decision_context（V2 Prefetch 流水线产出的精炼决策）
        decision_context = row.get("_decision_context")
        if decision_context:
            test_case_logic += (
                f"\n\n### 精炼后的赋值指令（decision_context）\n"
                f"{json.dumps(decision_context, ensure_ascii=False, indent=2)}\n"
                f"请根据以上指令填写 params 和 assertions，"
                "${{}} 占位符原样复制到 YAML 中。"
            )

        # 注入完整 api_sequence（V2 Prefetch 流水线拼接的前置 + 用例 API 序列）
        api_sequence = row.get("_api_sequence")
        if api_sequence:
            test_case_logic += (
                f"\n\n### 完整 API 调用序列\n"
                + "\n".join(api_sequence)
            )
        case_label = (
            f"{row.get('case_id') or os.path.basename(os.path.dirname(output_path))}"
            f" | {os.path.basename(os.path.dirname(output_path))}/{os.path.basename(output_path)}"
        )

        # === 阶段 1：thinking 分析（首轮=需求分析 / 修复轮=带诊断包自查） ===
        if repair_ctx:
            think_prompt = repair_yaml_data_prompt()
            prompt_vars = dict(
                skeleton=repair_ctx.get("skeleton", skeleton_text),
                failed_yaml=repair_ctx.get("failed_yaml", ""),
                error_roadmap=repair_ctx.get("error_roadmap", ""),
                data_factory_methods_section=repair_ctx.get("data_factory_methods_section", ""),
                api_definitions_section=repair_ctx.get("api_definitions_section", ""),
            )
            node_label = f"repair_yaml_data_ROUND{repair_ctx.get('round_no', 2)}"
            prompt_label = "repair_yaml_data_prompt"
        else:
            think_prompt = analyze_yaml_data_prompt()
            prompt_vars = dict(
                skeleton=skeleton_text,
                api_definitions=api_defs_json,
                test_case_logic=test_case_logic,
                user_context=user_ctx,
                data_factory_methods=factory_methods_text,
            )
            node_label = "analyze_yaml_data"
            prompt_label = "analyze_yaml_data_prompt"

        llm_kwargs = {"extra_body": {"thinking": {"type": "enabled"}}}
        bound_llm = self.llm.bind(**llm_kwargs)
        analysis_result = bound_llm.invoke(think_prompt.format_messages(**prompt_vars))
        analysis = analysis_result.content if hasattr(analysis_result, "content") else str(analysis_result)

        log_thinking(node_label, case_label, analysis, prompt_label=prompt_label)

        # === 阶段 2：json_mode 填表 → 提取 JSON → Pydantic 校验 ===
        format_prompt = format_yaml_data_prompt()
        format_vars = dict(
            skeleton=skeleton_text,
            data_analysis=analysis,
            api_definitions=api_defs_json,
            test_case_logic=test_case_logic,
            user_context=user_ctx,
            data_factory_methods=factory_methods_text,
        )
        llm_with_json = self.llm.bind(response_format={"type": "json_object"})
        raw_result = llm_with_json.invoke(format_prompt.format_messages(**format_vars))
        raw_text = raw_result.content if hasattr(raw_result, "content") else str(raw_result)

        # 提取 JSON + Pydantic 校验
        try:
            parsed = _extract_json_from_thinking(raw_text)
            test_data = TestDataModel.model_validate(parsed)
        except Exception as e:
            repair_ctx = prepare_repair_context(raw_text, e)
            raise RepairNeeded(repair_ctx) from e

        yaml_text = yaml.dump(
            [step.model_dump(exclude_none=True, by_alias=True) for step in test_data.data],
            allow_unicode=True, indent=2, default_flow_style=False,
        )
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        tmp_path = output_path + ".tmp"
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        with open(tmp_path, "w", encoding="utf-8") as f:
            f.write(yaml_text)
        os.replace(tmp_path, output_path)
        return output_path

    def _thinking_per_story(self, story: dict, filtered_apis_json: str,
                            factory_methods: str) -> dict:
        """Phase C 单 story thinking 调用：消费 dep_map 的 decision_map，结合工厂字典和
        去重 API 定义，精炼为最终赋值方案（refined_decision_map，内存传递）。

        Args:
            story: dep_map 中的一个 story 对象（含 decision_map / internal_dependency 等）
            filtered_apis_json: 该 story 涉及的接口定义 JSON（已去重）
            factory_methods: 数据工厂方法清单文本

        Returns:
            refined_decision_map: {case_id: {"steps": [refined DecisionStep dicts]}}
        """
        from observability import log_thinking

        story_name = story.get("story_name", "unknown")
        decision_map = story.get("decision_map", {})
        internal_dep = story.get("internal_dependency", {})
        cross_module = story.get("cross_module_dependency", {})

        # 组装 thinking prompt
        system_text = (
            "你是测试数据精炼专家。根据 dep_map 的初步 decision_map、去重的接口定义和"
            "数据工厂方法清单，对每条用例的参数赋值方案进行精炼。\n\n"
            "### 你的任务\n"
            "1. 校验 decision_map 中每个 params 字段的赋值是否正确\n"
            "2. 补充漏填的必填字段（从接口定义 parameters 中获取）\n"
            "3. 修正错误的工厂方法引用，替换为正确的 ${函数名(参数)} 语法\n"
            "4. 校验 assertions 的字段是否来自接口 returns\n"
            "5. 确保 ${get_extract_data(xxx)} 中的变量名 xxx 存在于 internal_dependency 的 output_var\n\n"
            "### 数据工厂方法（只能从此清单选择）\n"
            f"{factory_methods}\n\n"
            "### 输出格式\n"
            "输出 JSON 对象，key 为 case_id，value 为该用例精炼后的 decision：\n"
            '{"TC-001": {"steps": [{...}]}, "TC-002": {...}}\n\n'
            "params 中：静态值直接写，动态值用 ${} 字符串。\n"
            "禁止编造工厂清单中不存在的函数名。禁止 Markdown。"
        )

        human_parts = [
            f"### Story: {story_name}",
            f"### 接口定义（仅该 story 涉及）\n{filtered_apis_json}",
            f"### 初步 decision_map\n{json.dumps(decision_map, ensure_ascii=False, indent=2)}",
            f"### internal_dependency\n{json.dumps(internal_dep, ensure_ascii=False, indent=2)}",
        ]
        if cross_module:
            human_parts.append(
                f"### cross_module_dependency\n{json.dumps(cross_module, ensure_ascii=False, indent=2)}"
            )
        human_text = "\n\n".join(human_parts)

        llm_kwargs = {"extra_body": {"thinking": {"type": "enabled"}}}
        if config.THINKING_TIMEOUT:
            llm_kwargs["timeout"] = config.THINKING_TIMEOUT
        bound_llm = self.llm.bind(**llm_kwargs)

        from langchain_core.messages import SystemMessage, HumanMessage
        messages = [
            SystemMessage(content=system_text),
            HumanMessage(content=human_text + "\n\n请精炼 decision_map："),
        ]
        result = bound_llm.invoke(messages)
        raw_text = result.content if hasattr(result, "content") else str(result)

        log_thinking(
            f"thinking_per_story_{story_name}",
            f"cases={len(decision_map)}",
            raw_text,
            prompt_label="thinking_per_story",
        )

        # 从 thinking 输出提取 JSON
        refined = _extract_json_from_thinking(raw_text)

        # 合并原始 decision_map 中未被精炼覆盖的 key（LLM 可能省略未改动的条目）
        for cid in decision_map:
            if cid not in refined:
                refined[cid] = decision_map[cid]

        return refined

    def _generate_all_yamls(self, excel_path: str, api_defs_json: str,
                            user_ctx: str, dep_map: dict | None = None) -> dict:
        """Phase C YAML 生成 — 支持两种模式：

        V2 (dep_map 非 None): Prefetch 流水线 — Queue(1) producer-consumer，
          thinking 提前预取下一个 story，LLM 调用 ≤ 1 thinking + N json_mode。

        V1 (dep_map=None): 传统模式 — 直接从 Excel 按 feature→story 分组生成。
          保留用于向后兼容和 dep_map 生成失败时的降级路径。

        目录结构:
          testcase/<feature_en>/
            setup_data/setup_<class_slug>.yaml
            setup_data/teardown_<class_slug>.yaml
            <func_en>/test_data.yaml
        """
        logger.info("\n🔢 正在生成 YAML 测试数据...")

        _empty = {"total": 0, "success": 0, "failed": 0,
                  "repaired": 0, "rounds": 0, "errors_file": None}

        if not excel_path:
            logger.info("   ⚠️ 无 Excel 路径，跳过 YAML 生成")
            return dict(_empty)

        output_base = os.path.dirname(excel_path)

        # 新一轮全量生成开始：清理上次残留的错误清单
        errors_path = os.path.join(output_base, "_generation_errors.json")
        try:
            os.remove(errors_path)
        except OSError:
            pass

        raw_rows = self._read_excel_rows(excel_path)
        if not raw_rows:
            logger.info("   ⚠️ Excel 中无数据，跳过 YAML 生成")
            return dict(_empty)

        translations = self._translate_to_en(excel_path, raw_rows)
        feature_en_map = translations["feature_en"]
        story_en_map = translations["story_en"]
        title_en_map = translations["title_en"]
        shared_pres = self._read_shared_preconditions(excel_path)

        # C6-1: 断言校验（V1/V2 共用）
        assertion_errors = []
        for r in raw_rows:
            expected = r.get("expected", "")
            if not expected:
                continue
            for step_idx, step_text in enumerate(expected.split("\n"), 1):
                step_text = step_text.strip()
                if not step_text:
                    continue
                try:
                    self._parse_assertion(step_text)
                except self.AssertionParseError as e:
                    assertion_errors.append(
                        f"{r.get('case_id', '?')} step{step_idx}: {e}"
                    )
        if assertion_errors:
            logger.warning("   ⚠️ 断言格式校验失败 %d 条（不阻断，继续生成）:", len(assertion_errors))
            for err in assertion_errors[:10]:
                logger.warning("     %s", err)

        # ---- JSON 骨架生成（入口一次，沿链路传递）----
        from prompts.response_model import TestData as TestDataModel
        skeleton_text = generate_json_skeleton(TestDataModel)
        logger.info("   📐 JSON 骨架已生成（%d 字符）", len(skeleton_text))

        # ---- 选择 V2 或 V1 路径 ----
        if dep_map and dep_map.get("stories"):
            result = self._generate_yamls_v2(
                dep_map, raw_rows, api_defs_json, user_ctx, output_base, skeleton_text,
                feature_en_map, story_en_map, title_en_map, shared_pres,
            )
        else:
            result = self._generate_yamls_v1(
                raw_rows, api_defs_json, user_ctx, output_base, skeleton_text,
                feature_en_map, story_en_map, title_en_map, shared_pres,
            )

        # --- YAML 后校验（V1/V2 共用）---
        from agent_components.post_validator import YamlPostValidator
        validator = YamlPostValidator()
        post_issues = validator.validate_all(output_base)
        _post_issues_path = os.path.join(output_base, "_post_validation_issues.json")
        if post_issues:
            import json as _json
            with open(_post_issues_path, "w", encoding="utf-8") as _f:
                _json.dump(post_issues, _f, ensure_ascii=False, indent=2)
            _fixable = [i for i in post_issues if i.get("severity") in ("P0", "P1")]
            _all_tasks = getattr(self, "_all_yaml_tasks", [])
            if _fixable and _all_tasks and result["rounds"] < config.YAML_REPAIR_ROUNDS:
                _affected_paths = {i["yaml_path"] for i in _fixable}
                _affected_tasks = [
                    (row, path) for row, path in _all_tasks
                    if os.path.abspath(path) in {os.path.abspath(p) for p in _affected_paths}
                ]
                if _affected_tasks:
                    logger.info(f"   🔧 后校验发现 {len(_fixable)} 个 P0/P1 问题，"
                                f"追加一轮修复（{len(_affected_tasks)} 个文件）")
                    _post_result = self._run_yaml_rounds(
                        _affected_tasks, api_defs_json, user_ctx, output_base,
                        skeleton_text, post_check_issues=_fixable, repair_rounds=1,
                    )
                    result["success"] = result["success"] - len(_affected_tasks) + _post_result["success"]
                    result["failed"] = _post_result["failed"]
                    result["repaired"] += _post_result["repaired"]
                    result["rounds"] += _post_result["rounds"]
            _p2_count = len(post_issues) - len(_fixable)
            if _p2_count:
                logger.info(f"   📝 后校验发现 {_p2_count} 个 P2 问题（仅告警，见 {_post_issues_path}）")
        else:
            try:
                os.remove(_post_issues_path)
            except OSError:
                pass

        # --- 变量读写审计（§4.6，纯代码后置扫描）---
        audit_warnings = _audit_variable_reads_writes(output_base)
        if audit_warnings:
            result["warnings"] = audit_warnings
            p1_count = sum(1 for w in audit_warnings if w.get("severity") == "P1")
            p2_count = sum(1 for w in audit_warnings if w.get("severity") == "P2")
            logger.warning("   🔍 变量读写审计: %d 个警告（P1=%d, P2=%d）",
                           len(audit_warnings), p1_count, p2_count)
            for w in audit_warnings[:5]:
                logger.warning("     [%s] %s: %s", w.get("severity", "?"),
                               w.get("check", ""), w.get("current", "")[:120])
            # 写入 thinking_trace.log
            from observability import log_thinking
            log_thinking(
                "variable_read_write_audit",
                f"{len(audit_warnings)} warnings",
                json.dumps(audit_warnings, ensure_ascii=False, indent=2),
                prompt_label="post_audit",
            )
            # 同步写入 _generation_errors.json（追加 _audit section，不覆盖已有 story 错误）
            _append_generation_errors(output_base, "_audit", audit_warnings)

        self._log_node_output("generate_all_yamls", result)
        return result

    # ---- V2: Prefetch 流水线 ----

    def _generate_yamls_v2(self, dep_map: dict, raw_rows: list, api_defs_json: str,
                            user_ctx: str, output_base: str, skeleton_text: str,
                            feature_en_map: dict, story_en_map: dict,
                            title_en_map: dict, shared_pres: list) -> dict:
        """V2 Prefetch 流水线：Queue(1) producer-consumer 模式。

        producer 线程串行产出 thinking 结果（每 story 一个 refined_decision_map），
        consumer 主线程消费并运行 json_mode 并发生成。
        """
        from queue import Queue
        import threading

        stories = dep_map.get("stories", [])
        api_index = build_api_index(json.loads(api_defs_json))
        factory_methods_text = self._load_factory_methods()

        # 构建 case_id → raw_row 索引
        case_row_map: dict[str, dict] = {r["case_id"]: r for r in raw_rows}

        ready_queue: Queue = Queue(maxsize=1)
        all_results: list[dict] = []
        self._all_yaml_tasks: list = []

        def thinking_producer():
            """线程：串行产出 thinking 结果，给 consumer 消费。"""
            for story in stories:
                story_name = story.get("story_name", "unknown")
                try:
                    urls = _collect_story_urls(story)
                    filtered = filter_apis_by_urls(api_index, urls)
                    if not filtered:
                        logger.warning("   ⚠️ story '%s': URL 匹配 0 个接口，跳过", story_name)
                        ready_queue.put((story, [], {}))
                        continue
                    filtered_json = json.dumps(filtered, ensure_ascii=False)
                    refined = self._thinking_per_story(story, filtered_json, factory_methods_text)
                    ready_queue.put((story, filtered, refined))
                except Exception as e:
                    logger.error("   ❌ story '%s' thinking 失败: %s", story_name, e)
                    ready_queue.put((story, [], {"_error": str(e)}))
            ready_queue.put(None)  # 哨兵

        producer = threading.Thread(target=thinking_producer, daemon=True)
        producer.start()

        total_success = 0
        total_failed = 0
        total_repaired = 0
        total_rounds = 0

        while True:
            item = ready_queue.get()
            if item is None:
                break
            story, filtered_apis, refined = item
            story_name = story.get("story_name", "unknown")

            if refined.get("_error"):
                total_failed += len(story.get("case_api_sequences", {}))
                _append_generation_errors(output_base, story_name,
                    [{"case_id": cid, "error": "THINKING_FAILED: " + str(refined.get("_error", ""))}
                     for cid in story.get("case_api_sequences", {})])
                continue

            # 构建该 story 的 yaml_tasks
            story_tasks = self._build_story_yaml_tasks(
                story, refined, case_row_map, output_base,
                feature_en_map, story_en_map, title_en_map, shared_pres,
            )
            self._all_yaml_tasks.extend(story_tasks)

            if not story_tasks:
                logger.info("   📭 story '%s': 无用例任务", story_name)
                continue

            filtered_json = json.dumps(filtered_apis, ensure_ascii=False) if filtered_apis else api_defs_json

            logger.info("   🚀 story '%s': %d 个 YAML 任务（含 setup/teardown）", story_name, len(story_tasks))
            s_result = self._run_yaml_rounds(story_tasks, filtered_json, user_ctx, output_base,
                                             skeleton_text, write_errors_file=False)

            total_success += s_result["success"]
            total_failed += s_result["failed"]
            total_repaired += s_result.get("repaired", 0)
            total_rounds = max(total_rounds, s_result.get("rounds", 0))

            # V2: 按 story 追加写入 _generation_errors.json（避免后一 story 覆盖前一 story）
            if s_result.get("error_payloads"):
                _append_generation_errors(output_base, story_name, s_result["error_payloads"])

            all_results.append(s_result)

        producer.join()

        total_tasks = sum(r["total"] for r in all_results)
        errors_file = os.path.join(output_base, "_generation_errors.json")
        if not os.path.exists(errors_file):
            errors_file = None
        else:
            logger.info("   📋 _generation_errors.json: %d 个 story 有终态失败",
                         len([r for r in all_results if r.get("error_payloads")]))

        result = {
            "total": total_tasks,
            "success": total_success,
            "failed": total_failed,
            "repaired": total_repaired,
            "rounds": total_rounds,
            "errors_file": errors_file,
        }
        return result

    def _build_story_yaml_tasks(self, story: dict, refined: dict,
                                 case_row_map: dict, output_base: str,
                                 feature_en_map: dict, story_en_map: dict,
                                 title_en_map: dict, shared_pres: list) -> list:
        """为一个 story 构建 yaml_tasks 列表。

        Args:
            story: dep_map 中的 story 对象
            refined: 精炼后的 decision_map
            case_row_map: case_id → Excel 行映射
        Returns:
            [(row_dict, yaml_path), ...]
        """
        story_name = story.get("story_name", "")
        case_api_seqs = story.get("case_api_sequences", {})

        # 找到该 story 对应的 feature_en / story_en
        feature_cn = ""
        for cid in case_api_seqs:
            row = case_row_map.get(cid)
            if row:
                feature_cn = row.get("feature", "")
                break
        feature_en = feature_en_map.get(feature_cn, self._sanitize_en(self._pinyin_fallback(feature_cn)))
        story_en = story_en_map.get(story_name, self._sanitize_en(self._pinyin_fallback(story_name)))
        class_slug = re.sub(r'(?<!^)(?=[A-Z])', '_', story_en).lower()

        # 收集该 story 涉及的前置条件
        pre_ids = set()
        story_case_rows = []
        for cid in case_api_seqs:
            row = case_row_map.get(cid)
            if row:
                story_case_rows.append(row)
                pre_str = row.get("preconditions", "")
                if pre_str and pre_str != "无":
                    for pid in pre_str.split(","):
                        pid = pid.strip()
                        if pid.startswith("PRE-"):
                            pre_ids.add(pid)

        tasks = []

        # setup_data YAML
        setup_dir = os.path.join(output_base, feature_en, "setup_data")
        os.makedirs(setup_dir, exist_ok=True)

        # 用 dep_map 的 story_pre_api_sequence + teardown_api_sequence 生成 setup/teardown
        pre_api_seq = story.get("story_pre_api_sequence", [])
        teardown_api_seq = story.get("teardown_api_sequence", [])

        if pre_api_seq or pre_ids:
            # setup YAML: 从 story_pre_api_sequence 构建
            setup_parts = []
            for entry in pre_api_seq:
                setup_parts.append(f"# {entry}")
            if pre_ids:
                for pid in sorted(pre_ids):
                    pre = next((p for p in shared_pres if p["id"] == pid), None)
                    if pre:
                        setup_parts.append(f"# {pid}: {pre['name']}\n{pre['steps']}")
            setup_text = "\n".join(setup_parts) if setup_parts else "# 无前置步骤"
            setup_yaml = os.path.join(setup_dir, f"setup_{class_slug}.yaml")
            tasks.append((
                {"steps": setup_text, "expected": "", "case_id": f"setup_{class_slug}"},
                setup_yaml,
            ))

        if teardown_api_seq:
            teardown_text = "\n".join(f"# {entry}" for entry in teardown_api_seq)
            teardown_yaml = os.path.join(setup_dir, f"teardown_{class_slug}.yaml")
            tasks.append((
                {"steps": teardown_text, "expected": "", "case_id": f"teardown_{class_slug}"},
                teardown_yaml,
            ))

        # func YAML（每个 TC 一个目录）
        for cid in case_api_seqs:
            row = case_row_map.get(cid)
            if not row:
                logger.warning("   ⚠️ case_id '%s' 在 Excel 中不存在，跳过", cid)
                continue

            title_cn = row["title"]
            func_en = title_en_map.get(
                title_cn,
                "test_" + self._sanitize_en(self._pinyin_fallback(title_cn))
            )
            if not func_en.startswith("test_"):
                func_en = "test_" + func_en
            func_dir = os.path.join(output_base, feature_en, func_en)
            os.makedirs(func_dir, exist_ok=True)

            # 注入 decision_context 到 row（供 _generate_one_yaml 使用）
            row_with_ctx = dict(row)
            if cid in refined:
                row_with_ctx["_decision_context"] = refined[cid]
            # 拼接完整 api_sequence
            full_seq = story.get("story_pre_api_sequence", []) + case_api_seqs.get(cid, [])
            row_with_ctx["_api_sequence"] = full_seq

            yaml_path = os.path.join(func_dir, "test_data.yaml")
            tasks.append((row_with_ctx, yaml_path))

        return tasks

    # ---- V1: 传统模式（降级路径） ----

    def _generate_yamls_v1(self, raw_rows: list, api_defs_json: str,
                            user_ctx: str, output_base: str, skeleton_text: str,
                            feature_en_map: dict, story_en_map: dict,
                            title_en_map: dict, shared_pres: list) -> dict:
        """V1 传统模式：直接从 Excel 按 feature→story 分组生成 YAML（无 dep_map）。"""
        from collections import defaultdict as _defaultdict

        feature_story_map = _defaultdict(lambda: _defaultdict(list))
        for r in raw_rows:
            feature_story_map[r["feature"]][r["story"]].append(r)

        yaml_tasks = []
        for feature_cn, stories in feature_story_map.items():
            feature_en = feature_en_map.get(feature_cn,
                self._sanitize_en(self._pinyin_fallback(feature_cn)))
            for story_cn, cases in stories.items():
                story_en = story_en_map.get(story_cn,
                    self._sanitize_en(self._pinyin_fallback(story_cn)))
                class_slug = re.sub(r'(?<!^)(?=[A-Z])', '_', story_en).lower()

                pre_ids = set()
                for c in cases:
                    pre_str = c.get("preconditions", "")
                    if pre_str and pre_str != "无":
                        for pid in pre_str.split(","):
                            pid = pid.strip()
                            if pid.startswith("PRE-"):
                                pre_ids.add(pid)

                setup_dir = os.path.join(output_base, feature_en, "setup_data")
                os.makedirs(setup_dir, exist_ok=True)

                if pre_ids:
                    setup_lines = []
                    teardown_lines = []
                    for pid in sorted(pre_ids):
                        pre = next((p for p in shared_pres if p["id"] == pid), None)
                        if pre:
                            setup_lines.append(f"# {pid}: {pre['name']}\n{pre['steps']}")
                            teardown_lines.append(
                                f"# 清理 {pid}: {pre['name']}\n"
                                f"根据 {pid} 的创建步骤逆向操作：{pre['steps'][:200]}"
                            )
                    setup_text = "\n".join(setup_lines)
                    teardown_text = "\n".join(teardown_lines)
                    setup_yaml = os.path.join(setup_dir, f"setup_{class_slug}.yaml")
                    teardown_yaml = os.path.join(setup_dir, f"teardown_{class_slug}.yaml")
                    yaml_tasks.append((
                        {"steps": setup_text, "expected": "",
                         "case_id": f"setup_{class_slug}"}, setup_yaml))
                    yaml_tasks.append((
                        {"steps": teardown_text, "expected": "",
                         "case_id": f"teardown_{class_slug}"}, teardown_yaml))

                for c in cases:
                    title_cn = c["title"]
                    func_en = title_en_map.get(
                        title_cn,
                        "test_" + self._sanitize_en(self._pinyin_fallback(title_cn))
                    )
                    if not func_en.startswith("test_"):
                        func_en = "test_" + func_en
                    func_dir = os.path.join(output_base, feature_en, func_en)
                    os.makedirs(func_dir, exist_ok=True)
                    yaml_path = os.path.join(func_dir, "test_data.yaml")
                    yaml_tasks.append((c, yaml_path))

        self._all_yaml_tasks = yaml_tasks

        total = len(yaml_tasks)
        if not total:
            logger.info("   ⚠️ 没有需要生成的 YAML")
            return {"total": 0, "success": 0, "failed": 0,
                    "repaired": 0, "rounds": 0, "errors_file": None}

        logger.info(f"   📋 共需生成 {total} 个 YAML 文件（含 setup/teardown），"
                    f"并发 {config.YAML_CONCURRENCY} 个线程，"
                    f"修复轮上限 {config.YAML_REPAIR_ROUNDS}")

        return self._run_yaml_rounds(yaml_tasks, api_defs_json, user_ctx, output_base, skeleton_text)

    def _run_yaml_rounds(self, yaml_tasks: list, api_defs_json: str, user_ctx: str,
                         output_base: str, skeleton_text: str,
                         gen_func=None, repair_rounds: int = None,
                         post_check_issues: list | None = None,
                         write_errors_file: bool = True,
                         circuit_breaker_threshold: float | None = None) -> dict:
        """YAML 生成轮次循环。

        第 1 轮全量并发生成；失败项登记占位（不写盘）→ 轮末汇总错误模式 →
        修复轮携带诊断包 + 骨架 + 条件上下文送思考节点自查重生成；
        超过修复轮上限仍失败 → 终态：计 failed + 写 _generation_errors.json，
        不写任何占位假文件。

        Args:
            skeleton_text: JSON 骨架文本，由 _generate_all_yamls 入口生成一次后传入
            gen_func: 可注入的单文件生成函数（单元测试用），签名同 _generate_one_yaml
            repair_rounds: 修复轮数覆盖（默认 config.YAML_REPAIR_ROUNDS）
            post_check_issues: YAML 后校验发现的问题列表（直接注入修复轮）
            write_errors_file: V2 流水线中设为 False，由调用方通过 _append_generation_errors
                              按 story 追加，避免后一 story 覆盖前一 story 的终态错误
        """
        from observability import log_phase_header, log_thinking, get_thinking_logger
        from web.tasks import _BoundedThreadPoolExecutor
        from concurrent.futures import as_completed
        from prompts.response_model import ValidationInterceptor

        ValidationInterceptor.reset()

        gen = gen_func or (lambda row, api, ctx, path, rctx=None:
                           self._generate_one_yaml(row, api, ctx, path, skeleton_text, rctx))
        max_repair = config.YAML_REPAIR_ROUNDS if repair_rounds is None else repair_rounds
        tlog = get_thinking_logger()

        total = len(yaml_tasks)
        success = 0
        repaired = 0
        rounds_run = 0
        fail_seq = 0
        registry: list = []      # 最近一轮的失败登记（循环结束即终态失败清单）
        pending = [(row, path, None) for row, path in yaml_tasks]

        for round_no in range(1, max_repair + 2):   # 1=全量轮, 2..=修复轮
            if not pending:
                break
            rounds_run = round_no
            label = "第1轮(全量)" if round_no == 1 else f"修复轮{round_no}"
            log_phase_header(f"Phase C — YAML 生成 {label} ({len(pending)} 个)")
            logger.info(f"   🔄 {label}: {len(pending)} 个任务")

            failures: list = []
            batch = len(pending)
            with _BoundedThreadPoolExecutor(
                    max_workers=config.YAML_CONCURRENCY,
                    max_queue=config.YAML_CONCURRENCY * 2) as executor:
                future_map = {
                    executor.submit(gen, row, api_defs_json, user_ctx, path, rctx):
                        (row, path)
                    for row, path, rctx in pending
                }
                done = 0
                for future in as_completed(future_map):
                    row, path = future_map[future]
                    done += 1
                    try:
                        future.result()
                        success += 1
                        if round_no > 1:
                            repaired += 1
                        if done % 20 == 0:
                            logger.info(f"      [{done}/{batch}] ...")
                    except RepairNeeded as e:
                        fail_seq += 1
                        pid = f"GEN-FAIL-R{round_no}-{fail_seq:03d}"
                        rel_path = os.path.relpath(path, output_base).replace("\\", "/")
                        case_id = str(row.get("case_id")
                                      or os.path.basename(os.path.dirname(path)))
                        repair_ctx = e.repair_ctx
                        failures.append({
                            "placeholder_id": pid,
                            "case_id": case_id,
                            "yaml_path": rel_path,
                            "round": round_no,
                            "error": repair_ctx.get("error_roadmap", ""),
                            "raw_output_snippet": repair_ctx.get("raw_text", ""),
                            "repair_ctx": repair_ctx,
                            "row": row,
                            "path": path,
                        })
                    except Exception as e:
                        fail_seq += 1
                        pid = f"GEN-FAIL-R{round_no}-{fail_seq:03d}"
                        err_text = str(e)
                        rel_path = os.path.relpath(path, output_base).replace("\\", "/")
                        case_id = str(row.get("case_id")
                                      or os.path.basename(os.path.dirname(path)))
                        raw_snippet = _extract_completion_snippet(err_text)
                        failures.append({
                            "placeholder_id": pid,
                            "case_id": case_id,
                            "yaml_path": rel_path,
                            "round": round_no,
                            "error": err_text[:2000],
                            "raw_output_snippet": raw_snippet,
                            "row": row,
                            "path": path,
                        })
                        logger.info(f"      [{done}/{batch}] ❌ "
                                    f"{os.path.basename(path)} ({pid})")
                        # 失败标记落 thinking_trace.log（与 generate_excel_plan_FAILED 同风格）
                        log_thinking(
                            "generate_yaml_FAILED",
                            f"| {case_id} | {rel_path} | {pid} |",
                            err_text[:1500],
                            prompt_label="format_yaml_data_prompt",
                        )
                        # 详细错误日志：原文 + 错误点，写入输出目录
                        _write_fail_detail(output_base, pid, case_id, rel_path,
                                           round_no, err_text, raw_snippet)

            ok = batch - len(failures)
            tlog.info(f"ROUND{round_no}: {ok}/{batch} 通过, {len(failures)} 登记")
            logger.info(f"   ✅ {label}: {ok}/{batch} 通过, {len(failures)} 失败登记")

            # === 全局熔断：首轮失败率超阈值 → 终止（防止 prompt/骨架缺陷导致 token 失控）===
            if round_no == 1 and batch > 0:
                failure_rate = len(failures) / batch
                threshold = (circuit_breaker_threshold
                             if circuit_breaker_threshold is not None
                             else getattr(config, "YAML_FAILURE_CIRCUIT_BREAKER", 0.5))
                if failure_rate > threshold:
                    _sample_errors = [f["error"][:120] for f in failures[:5]]
                    raise RuntimeError(
                        f"YAML 生成熔断：首轮失败率 {failure_rate:.0%}（{len(failures)}/{batch}）"
                        f"超过阈值 {threshold:.0%}。"
                        f"可能原因：骨架结构缺陷或 prompt 规则与 Pydantic 校验不匹配。"
                        f"错误样本: {_sample_errors}"
                    )

            registry = failures
            if not failures or round_no >= max_repair + 1:
                break

            # 组装修复轮：诊断包 + 骨架 + 条件注入上下文
            pattern = _summarize_error_patterns(failures)
            pending = []
            for f in failures:
                repair_ctx = f.get("repair_ctx", {})
                if repair_ctx:
                    # RepairNeeded 路径：使用诊断包
                    road = repair_ctx.get("error_roadmap", "")
                    factory_section = ""
                    api_section = ""
                    if _error_mentions_placeholder(road):
                        factory_section = (
                            "### 数据工厂方法（对照修正占位符）\n"
                            + self._load_factory_methods() + "\n\n"
                        )
                    if _error_mentions_api(road):
                        api_section = (
                            "### 接口定义（对照修正 url/method/参数）\n"
                            + api_defs_json + "\n\n"
                        )
                    rctx = {
                        "skeleton": skeleton_text,
                        "failed_yaml": repair_ctx.get("failed_yaml", ""),
                        "error_roadmap": road,
                        "data_factory_methods_section": factory_section,
                        "api_definitions_section": api_section,
                        "api_definitions": api_defs_json,
                        "test_case_logic": "",
                        "user_context": user_ctx,
                        "data_factory_methods": self._load_factory_methods(),
                        "error_pattern_summary": pattern,
                        "prior_output": repair_ctx.get("raw_text", ""),
                        "error_detail": road,
                        "round_no": round_no + 1,
                        "post_check_issues": "",
                    }
                else:
                    # 传统路径（非 RepairNeeded 异常）：保留旧逻辑兼容
                    rctx = {
                        "prior_output": f["raw_output_snippet"],
                        "error_detail": f["error"],
                        "error_pattern_summary": pattern,
                        "round_no": round_no + 1,
                        "post_check_issues": _format_post_issues_for_prompt(post_check_issues)
                            if post_check_issues else "",
                        "data_factory_methods": self._load_factory_methods(),
                        "api_definitions": api_defs_json,
                        "test_case_logic": "",
                        "user_context": user_ctx,
                        "skeleton": skeleton_text,
                        "failed_yaml": f["raw_output_snippet"],
                        "error_roadmap": f["error"],
                        "data_factory_methods_section": "",
                        "api_definitions_section": "",
                    }
                pending.append((f["row"], f["path"], rctx))

        failed = len(registry)
        errors_file = None
        error_payloads: list[dict] = []
        if registry:
            error_payloads = [{
                "placeholder_id": r["placeholder_id"],
                "case_id": r["case_id"],
                "yaml_path": r["yaml_path"],
                "rounds_attempted": rounds_run,
                "error": r["error"],
                "raw_output_snippet": r["raw_output_snippet"],
            } for r in registry]
            if write_errors_file:
                errors_file = os.path.join(output_base, "_generation_errors.json")
                with open(errors_file, "w", encoding="utf-8") as f:
                    json.dump(error_payloads, f, ensure_ascii=False, indent=2)
                tlog.info(f"FINAL_FAILED: {failed} 个 → {errors_file}")
                logger.warning("   ⚠️ 终态失败 %d 个（不写占位文件），详见 %s",
                               failed, errors_file)
            else:
                logger.info("   📋 终态失败 %d 个（由调用方写入 _generation_errors.json）", failed)

        logger.info(f"   ✅ 完成: {success}/{total}，修复 {repaired}，"
                    f"仍失败 {failed}，轮次 {rounds_run}")

        # 写入 Schema 校验拦截报告（独立于 _generation_errors.json，用于提示词优化）
        ValidationInterceptor.write_report("logs")

        return {"total": total, "success": success, "failed": failed,
                "repaired": repaired, "rounds": rounds_run,
                "errors_file": errors_file,
                "error_payloads": error_payloads}
