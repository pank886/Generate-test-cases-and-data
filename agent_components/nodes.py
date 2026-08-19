"""LangGraph 各个节点方法"""
import json
import os
from collections import defaultdict
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import config
from observability import get_logger
from agent_components.dual_chroma import get_chroma_db
from agent_components.state import State
from agent_components.llm_client import (
    reload_llm,
    _get_llm,
    load_factory_methods,
    invoke_think,
    invoke_structured,
)
from agent_components.graph_logging import (
    split_thinking_sections,
    serialize_for_log,
    cleanup_logs,
    log_node_output,
)
from agent_components.prompt_builder import prepare_plan_prompt_vars
from prompts.response_model import (
    ProperResponse,
    TestData,
    ExcelPlan,
    ExcelRow,
    ExcelPlanV2,
    SharedPrecondition,
    TestCaseRow,
    IntentConfirmation,
)
from prompts.definitions import PromptFactory
from agent_components.retrievers import RetrievalMixin
from agent_components.generators import GenerationMixin

logger = get_logger(__name__)

# 方法特性配置表（声明式，集中管理 method 与 thinking 的兼容性）
METHOD_FEATURES = {
    "function_calling": {"supports_thinking": False},
    "json_mode": {"supports_thinking": False},
    "json_schema": {"supports_thinking": False},
    "free_text": {"supports_thinking": True},
}


def _quality_gate_decision(n_pass: int, n_total: int, regen_attempted: bool) -> str | None:
    """质量门禁决策：首轮通过率 < 50% 时的处置。

    Returns:
        "regen" —— 首轮不达标，触发全量重新生成（真实重跑 thinking 节点）
        "abort" —— 重试后仍不达标，终止生成并报错
        None    —— 通过（达标或无需判断）
    """
    if n_total <= 0 or n_pass >= n_total / 2:
        return None
    return "abort" if regen_attempted else "regen"


class ChatTestAgentGraph(RetrievalMixin, GenerationMixin):
    """智能测试助手——LangGraph 节点方法的容器类

    Excel 计划生成 + 核心工具方法（本文件）
    Phase B 检索节点 → RetrievalMixin (retrievers.py)
    Phase C PY/YAML 生成节点 → GenerationMixin (generators.py)
    """

    def __init__(self):
        self.llm = _get_llm()

        self.prompt_factory = PromptFactory()

        self.dual_chroma = get_chroma_db()

        # 工作流日志累积器（同一次运行的所有节点共用一份文件）
        # 原初始化误放在 _finalize_excel_plan 的 return 之后（死代码），导致 _log_node_output 访问时报
        # AttributeError: _run_timestamp；2026-08-03 移入 __init__ 修复。
        self._run_data: dict = {}
        self._run_timestamp: Optional[str] = None

    # ==================== 图内节点方法 ====================

    def _generate_excel_plan_thinking(self, state: State, gen_warning: str = ""):
        """【新】thinking+json_mode 一步生成 Excel 计划。

        合并原 analyze_test_points_raw + generate_excel_plan 两步：
        thinking 分析 → json_object 直接输出 ExcelPlanV2。
        失败时由 graph builder 降级到旧两步流程。

        Args:
            gen_warning: 质量门禁重试时的系统警告文本（注入 prompt 开头，
                首轮为空串，保持正常生成提示词不变）。
        """
        from observability import log_phase_header
        from prompts.response_model import ExcelPlanV2
        from agent_components.retrievers import _build_full_api_defs_text
        from database import get_session_ctx
        from database.operations import ModuleOps, BindingOps
        from database.operations.analysis import AnalysisOps
        from agent_components.api_annotations import ApiAnnotationRegistry

        log_phase_header("Phase B — thinking+json 一步生成 Excel 计划")
        logger.info("\n🧠 一步生成 Excel 测试计划（thinking + json_object）...")

        # ── 1. 准备数据（同 _analyze_test_points_raw）──
        confirmed_module = state.get("confirmed_module", "")
        analysis_text = ""
        with get_session_ctx() as session:
            mod = ModuleOps.get_by_name(session, confirmed_module)
            if mod:
                record = AnalysisOps.get_by_module_id(session, mod.id)
                if record:
                    parts = []
                    for key, label in [("scenario_analysis", "### 测试场景分析\n"),
                                       ("ui_flow_analysis", "### 页面交互逻辑\n"),
                                       ("api_analysis", "### 接口映射分析\n")]:
                        val = getattr(record, key, None)
                        if val:
                            parts.append(label + val)
                    if parts:
                        analysis_text = "\n\n".join(parts)
                    elif record.analysis_json:
                        analysis_text = record.analysis_json

            # 模块树
            tree = ModuleOps.get_tree(session)
            module_tree = json.dumps(tree, indent=2, ensure_ascii=False) if tree else "[]"

        # API 定义（供 LLM 分析：只给概要 name/method/url/description，
        # 避免全量参数/返回值撑爆 context 导致 thinking+json_object 返回空 content；
        # 详细参数分析由 module_analysis 预分析提供，与旧两步流程 _generate_excel_plan_node 一致）
        apis_text = json.dumps(
            [{"name": d.get("name", "?"), "method": d.get("method", "GET"),
              "url": d.get("url", ""), "description": d.get("description", "")}
             for d in (state.get("api_definitions") or [])],
            indent=2, ensure_ascii=False)

        # Phase C 快照用完整 API（新结构：header 映射 + body/return 六字段数组）
        api_full_for_snapshot = [
            {
                "name": d.get("name", "?"), "url": d.get("url", ""),
                "method": d.get("method", "GET"), "description": d.get("description", ""),
                "header": d.get("header", {}),
                "body": d.get("body", d.get("parameters", [])),
                "return": d.get("return", d.get("returns", [])),
                "annotations": {},
            }
            for d in (state.get("api_definitions") or [])
        ]
        for api in api_full_for_snapshot:
            ApiAnnotationRegistry.apply_all(api)

        # 关联模块
        related_text = ", ".join(state.get("related_modules", [])) or "无"
        user_ctx = state.get("original_input", "")

        # ── 2. 构造 prompt ──
        prompt = self.prompt_factory.generate_excel_plan_thinking()

        # ── 3. thinking + json_object 调用（空 content 有限重试，走公共方法）──
        #    2026-08-03 P2：deepseek-v4-flash 偶发返回空 content，复用同一份输入重试
        #    config.MAX_RETRIES 次（默认 2），仍失败抛异常 → 处理节点 requires_review。
        _think_llm = self.llm.bind(
            temperature=0.4,
            response_format={"type": "json_object"},
            extra_body={"thinking": {"type": "enabled"}},
        )
        _messages = prompt.format_messages(
            json_schema=json.dumps(
                ExcelPlanV2.model_json_schema(), ensure_ascii=False, indent=2),
            module_analysis=analysis_text or "（无预分析，请根据接口定义自行分析）",
            api_definitions=apis_text,
            related_docs=related_text,
            user_context=user_ctx,
            db_schema=config.DB_SCHEMA,  # 数据库表结构（占位，为空禁 [db]，2026-08-04 问题 2）
            gen_warning=gen_warning,  # 质量门禁重试注入警告；首轮空串保持提示词不变
        )
        _text = self._invoke_think(_think_llm, _messages, label="generate_plan_thinking")
        # 2026-08-12 修复：一步生成节点自 4a61792 起漏写 thinking 日志，恢复 thinking_trace.log 记录
        from observability import log_thinking
        log_thinking("generate_plan_thinking", state.get("original_input", ""), _text,
                     prompt_label="generate_excel_plan_thinking")

        # ── 4. 解析 ──
        plan = ExcelPlanV2.model_validate(json.loads(_text))
        logger.info(f"   => 一步生成完成: {len(plan.shared_preconditions)} 前置, {len(plan.test_cases)} 用例")

        # ── 5. 只生成不落盘：标注数据源，交由 generate_excel_plan 处理节点 校验/修复/落盘 ──
        return {
            "excel_plan": plan,
            "plan_source": "thinking",
            "api_full_for_snapshot": api_full_for_snapshot,
            "module_tree_json": module_tree,
        }

    def _generate_excel_plan_node(self, state: State):
        """生成 Excel 测试计划 V2（双 Sheet：测试计划 + 共享前置）。"""
        logger.info("\n📊 正在生成 Excel 测试计划...")

        from prompts.extraction_prompts import repair_excel_plan_prompt
        from agent_components.validator import validate_excel_file
        from agent_components.plan_validator import ExcelPlanValidator

        prompt = self.prompt_factory.generate_excel_plan_node()
        # Phase B prompt：只传接口概要（name/method/url/description），避免全量 JSON 撑爆 context
        api_summaries = [
            {"name": d.get("name", "?"), "method": d.get("method", "GET"),
             "url": d.get("url", ""), "description": d.get("description", "")}
            for d in (state.get("api_definitions") or [])
        ]
        all_apis_json = json.dumps(api_summaries, indent=2, ensure_ascii=False)
        # Phase C 快照：保留完整定义（新结构 header/body/return），存为 api_defs.json
        api_full_for_snapshot = [
            {
                "name": d.get("name", "?"), "url": d.get("url", ""),
                "method": d.get("method", "GET"), "description": d.get("description", ""),
                "header": d.get("header", {}),
                "body": d.get("body", d.get("parameters", [])),
                "return": d.get("return", d.get("returns", [])),
                "annotations": {},
            }
            for d in (state.get("api_definitions") or [])
        ]
        # 接口异常标识自动检测
        from agent_components.api_annotations import ApiAnnotationRegistry
        for api in api_full_for_snapshot:
            ApiAnnotationRegistry.apply_all(api)
        from database import get_session_ctx
        from database.operations import ModuleOps
        with get_session_ctx() as session:
            tree = ModuleOps.get_tree(session)
        module_tree_json = json.dumps(tree, indent=2, ensure_ascii=False)
        test_analysis = state.get("test_point_analysis") or "（无）"
        _sections = self._split_thinking_sections(test_analysis)
        prompt_vars = {
            "module_tree": module_tree_json,
            "analysis_section": _sections["analysis"],
            "shared_pre_section": _sections["preconditions"],
            "cases_section": _sections["cases"],
            "all_apis_info": all_apis_json,
            "user_context": state["original_input"],
        }

        # ── 数据源检测（2026-08 生成/处理解耦，方案3）：
        #    generate_excel_plan 为纯处理节点，只消费上游生成的 plan（thinking / 未来旧链路）。
        #    无外部 plan（thinking 失败）→ requires_review，不降级自生成。
        #    旧节点自生成兜底方案见 changelog/2026-08-02_old_generation_fallback.md，暂未启用。
        incoming_plan = state.get("excel_plan")
        if incoming_plan is None:
            return {
                "excel_plan": None, "excel_path": "", "output_dir": None,
                "requires_review": True,
                "error_info": ["生成环节未产出 plan（thinking 失败），且旧节点自生成兜底未启用"],
                "response_obj": ProperResponse(
                    proper_thinking=[], worth_to_remember=False,
                    final_response="测试计划生成失败：生成环节未产出计划，请重试",
                ),
            }
        # thinking 已构造的快照/模块树优先，避免重复查询与不一致
        api_full_for_snapshot = state.get("api_full_for_snapshot") or api_full_for_snapshot
        module_tree_json = state.get("module_tree_json") or module_tree_json

        # URL 有效性校验用真实接口路径模板（含 {xxx} 路径参数），2026-08-03 建议 3
        api_urls = [
            str(d.get("url", "")).strip()
            for d in (state.get("api_definitions") or [])
            if d.get("url")
        ]

        output_dir = None
        plan = None
        failed_details: list[tuple[int, dict, list[str]]] = []
        all_confirmed: list = []
        all_shared_pres: list = []  # 首轮共享前置，重试时复用
        failed_ids: set = set()  # 失败行 TC ID 集合，重试时只接受这些 ID 的修复
        _gen_attempt = 1          # 生成轮次（1=首轮，2=质量门禁全量重试轮）
        _gen_warning = ""         # 质量不达标时的警告信息（注入重生成 prompt）

        for attempt in range(config.EXCEL_REPAIR_ATTEMPTS):
            if attempt == 0 or _gen_warning:
                # === 首轮 / 质量门禁全量重试 ===
                if attempt == 0:
                    # 消费上游 thinking 生成的 plan（纯处理，不生成）
                    plan = incoming_plan
                    incoming_plan = None  # 仅首轮消费
                else:
                    # 质量门禁重试轮：真实全量重生成（thinking 节点，注入警告），非局部修复
                    logger.warning(
                        f"   ⚠️ 质量门禁：首轮通过率 < 50%，触发第 {_gen_attempt} 次全量重新生成...")
                    try:
                        plan = self._generate_excel_plan_thinking(
                            state, gen_warning=_gen_warning)["excel_plan"]
                    except Exception as e:
                        logger.error("   ❌ 质量门禁重试：全量重新生成异常: %s", e, exc_info=True)
                        raise RuntimeError(
                            f"Excel 生成质量门禁重试失败：全量重新生成异常（{e}），已终止") from e
                    _gen_warning = ""  # 已消费；重试轮之后仍失败走局部修复轮
                all_shared_pres = plan.shared_preconditions

                # 首轮校验全部用例（校验收敛：ExcelPlanValidator，含 9 类错误聚合 + URL 有效性 + db 拦截）
                _vr = ExcelPlanValidator.validate(plan, test_analysis,
                                                  api_urls=api_urls,
                                                  db_schema=config.DB_SCHEMA)
                _new_failed = _vr.failed_details
                all_confirmed = _vr.all_confirmed

                # 质量门禁：首轮通过率 < 50% → 全量重新生成一次；
                # 重试后仍 < 50% → 记录日志，终止生成并报错
                n_total = len(plan.test_cases)
                n_pass = len(all_confirmed)
                _gate = _quality_gate_decision(n_pass, n_total, regen_attempted=(_gen_attempt >= 2))
                if _gate == "regen":
                    _gen_attempt += 1
                    logger.warning(
                        f"   ⚠️ 质量门禁：首轮通过率 {n_pass}/{n_total} < 50%，"
                        f"触发全量重新生成"
                    )
                    _gen_warning = (
                        f"### ⚠️ 系统警告：上一轮生成质量未达标，本轮请重新生成\n"
                        f"上一轮 {n_total} 条用例中仅有 {n_pass} 条通过校验，"
                        f"通过率 {n_pass}/{n_total}，未达到 50% 的最低要求，"
                        "说明存在较多格式或字段不合规的问题。\n"
                        "请重新审视并严格遵循上述所有格式规则，确保本轮生成的用例全部合规。\n\n"
                    )
                    failed_details = []
                    continue
                if _gate == "abort":
                    logger.error(
                        f"   ❌ 质量门禁：全量重试后通过率仍 {n_pass}/{n_total} < 50%，终止生成")
                    raise RuntimeError(
                        f"Excel 生成质量不达标：全量重试后通过率仍 {n_pass}/{n_total} < 50%，已终止")

                # 记录失败行 ID（重试时只有匹配这些 ID 的修复才被接受）
                failed_ids = {f[1].get("id", "") for f in _new_failed}
                failed_details = _new_failed
                logger.warning(
                    f"   ⚠️ 校验: {n_pass} 用例通过, "
                    f"{len(failed_details)} 失败 (第{_gen_attempt}次)"
                )
                if not failed_details:
                    break
            else:
                # 重试：LLM 获得完整上下文修复，代码侧根据 failed_ids 裁剪输出
                attempt_label = f"第{attempt+1}次重试"
                failed_tc_list = []
                for f_idx, f_dict, f_errs in failed_details:
                    failed_tc_list.append(
                        f"TC ID: {f_dict.get('id','?')}\n"
                        f"  子模块: {f_dict.get('story','?')}\n"
                        f"  标题: {f_dict.get('title','?')}\n"
                        f"  步骤: {f_dict.get('steps','?')}\n"
                        f"  预期: {f_dict.get('expected','?')}\n"
                        f"  错误: {'; '.join(f_errs)}"
                    )
                failed_tc_text = "\n---\n".join(failed_tc_list)
                # 修复节点入参统一（2026-08 D1/D2）：
                #   ① 共享数据（与生成节点一致，_prepare_plan_prompt_vars）
                #   ② 错误用例 failed_test_cases
                #   ③ 拦截原因 block_reasons（validator 聚合，同类一条）
                _shared_vars = self._prepare_plan_prompt_vars(state)
                _block_reasons_text = "\n".join(
                    ExcelPlanValidator.aggregate_block_reasons(failed_details))
                repair_prompt = repair_excel_plan_prompt()
                plan = self._invoke_structured(repair_prompt, ExcelPlanV2,
                    method="json_mode",
                    json_schema=json.dumps(
                        ExcelPlanV2.model_json_schema(), ensure_ascii=False, indent=2),
                    failed_test_cases=failed_tc_text,
                    block_reasons=_block_reasons_text,
                    module_tree=_shared_vars["module_tree"],
                    analysis_section=_shared_vars["analysis_section"],
                    shared_pre_section=_shared_vars["shared_pre_section"],
                    cases_section=_shared_vars["cases_section"],
                    all_apis_info=_shared_vars["all_apis_info"],
                    db_schema=_shared_vars["db_schema"],
                )
                if isinstance(plan, list):
                    plan = ExcelPlanV2(shared_preconditions=[], test_cases=plan)

                # 重试校验：只接受 ID 匹配失败行的修复
                pre_ids_all = {p.id for p in all_shared_pres}
                # 修复轮共享前置按 ID 合并回 all_shared_pres（如 URL 修正落地，2026-08-03 建议 3）
                for p in plan.shared_preconditions:
                    pre_ids_all.add(p.id)
                    _pre_idx = next(
                        (i for i, x in enumerate(all_shared_pres) if x.id == p.id), None)
                    if _pre_idx is not None:
                        all_shared_pres[_pre_idx] = p
                    else:
                        all_shared_pres.append(p)

                _new_failed = []
                fixed_ids = set()
                _already_confirmed = {tc.id for tc in all_confirmed}
                _seen_in_retry = set()  # 防止同一批次内 LLM 输出重复 TC ID
                for tc in plan.test_cases:
                    # 拒绝不在失败 ID 集合中的行（LLM 幻觉出新的用例）
                    if tc.id not in failed_ids:
                        logger.warning(f"   ⚠️ 重试返回了不在失败列表中的 TC {tc.id}，已丢弃")
                        continue
                    # 拒绝重复输出已通过校验的用例
                    if tc.id in _already_confirmed:
                        logger.warning(f"   ⚠️ 重试返回了已通过的 TC {tc.id}，已丢弃")
                        continue
                    # 拒绝同一批次内的重复（LLM 在单次输出中生成多个相同 ID）
                    if tc.id in _seen_in_retry:
                        logger.warning(f"   ⚠️ 重试批次内重复 TC {tc.id}，已丢弃")
                        continue
                    _seen_in_retry.add(tc.id)
                    # 单用例校验（校验收敛：ExcelPlanValidator.check_case，含 URL 有效性 + db 拦截）
                    errs = ExcelPlanValidator.check_case(tc, pre_ids_all,
                                                         api_urls=api_urls,
                                                         db_schema=config.DB_SCHEMA)
                    if errs:
                        _new_failed.append((0, tc.model_dump(), errs))
                    else:
                        all_confirmed.append(tc)
                        fixed_ids.add(tc.id)

                # 仍未修复的失败行：保留在 failed_details 中
                #   - 已修复的 TC 移除
                #   - PRE 条目（共享前置 URL 错误）：修复轮已输出修正版并按 ID 合并，
                #     若该前置步骤 URL 现已命中真实接口 → 视为修复，移除
                _still_failed = []
                for f_idx, f_dict, f_errs in failed_details:
                    if f_dict.get("id", "") in fixed_ids:
                        continue
                    if f_dict.get("id", "").startswith("PRE-") and api_urls:
                        if not ExcelPlanValidator.check_urls(f_dict.get("steps", ""), api_urls):
                            continue
                    _still_failed.append((f_idx, f_dict, f_errs))
                failed_details = _still_failed + _new_failed
                # 去重：同一 ID 在旧版和新版同时存在时保留最新版
                _seen_ids = {}
                for _item in failed_details:
                    _seen_ids[_item[1].get("id", "")] = _item
                if len(_seen_ids) < len(failed_details):
                    logger.warning(f"   ⚠️ 修复轮去重: {len(failed_details)} → {len(_seen_ids)}")
                failed_details = list(_seen_ids.values())
                logger.warning(
                    f"   ⚠️ 校验: {len(all_confirmed)} 用例通过（本次修复 {len(plan.test_cases)} 行, "
                    f"接受 {len(fixed_ids)}, 仍失败 {len(failed_details)}）, "
                    f"({attempt_label})"
                )
                if not failed_details:
                    break

        # 最终校验 + 引用完整性
        if failed_details:
            pre_ids = {p.id for p in all_shared_pres}
            valid_cases = [tc for tc in all_confirmed]
            for f_idx, f_dict, f_errs in failed_details:
                if f_dict.get("id", "").startswith("PRE-"):
                    # 共享前置失败行（URL 未修复）不参与用例落盘；仍计入 fail_warn 提示人工审查
                    continue
                orphan = [p for p in (f_dict.get("preconditions") or []) if p not in pre_ids]
                if orphan:
                    continue
                valid_cases.append(TestCaseRow(
                    id=f_dict.get("id","?"), story=f_dict.get("story",""),
                    title=f_dict.get("title","?"),
                    preconditions=f_dict.get("preconditions") or [],
                    steps=f_dict.get("steps",""), expected=f_dict.get("expected",""),
                    mutates_data=f_dict.get("mutates_data", False),
                    is_negative_test=f_dict.get("is_negative_test", False),
                ))
        else:
            valid_cases = all_confirmed

        # === 最终安全阀：valid_cases 按 ID 去重（防御多路径聚合的重复） ===
        _seen_vc = set()
        _deduped = []
        _dup_count = 0
        for tc in valid_cases:
            if tc.id in _seen_vc:
                _dup_count += 1
                continue
            _seen_vc.add(tc.id)
            _deduped.append(tc)
        if _dup_count:
            logger.warning(
                f"   ⚠️ 最终去重安全阀触发: 移除 {_dup_count} 条重复用例 "
                f"（{len(valid_cases)} → {len(_deduped)}）")
        valid_cases = _deduped

        if failed_details:
            from observability import log_thinking
            error_parts = []
            for f_idx, f_dict, f_errs in failed_details:
                error_parts.append(
                    f"第{f_idx}行 | {f_dict.get('id','?')} | " + "; ".join(f_errs))
            fail_text = "\n".join(error_parts)
            log_thinking("generate_excel_plan_FAILED",
                         state.get("original_input", "?"),
                         f"校验失败 {len(failed_details)} 行\n"
                         f"--- 通过: {len(valid_cases)} 行 ---\n"
                         f"--- 失败详情 ---\n{fail_text}",
                         prompt_label="generate_excel_plan_node")

        if not valid_cases:
            return {
                "excel_plan": None, "excel_path": "",
                "output_dir": output_dir, "error_info": ["所有行均未通过校验"],
                "response_obj": ProperResponse(
                    proper_thinking=[], worth_to_remember=False,
                    final_response="Excel 测试计划生成失败：所有用例均未通过校验，请重试",
                ),
            }

        # 模块树路径
        tree_module = state.get("confirmed_module") or ""
        def _find_node_path(nodes, target, parts=None):
            if parts is None: parts = []
            for n in (nodes or []):
                if n.get("name") == target and n.get("name") != "全部模块":
                    return parts + [n.get("name")]
                found = _find_node_path(n.get("children") or [], target,
                                         parts + ([n.get("name")] if n.get("name") != "全部模块" else []))
                if found: return found
            return None
        path_parts = _find_node_path(tree, tree_module)
        if path_parts:
            dir_prefix = os.path.join(config.TESTCASE_BASE, *path_parts)
            _project = path_parts[0]
            _feature = path_parts[-1]
        else:
            _project = tree_module or valid_cases[0].story
            _feature = tree_module or valid_cases[0].story
            dir_prefix = os.path.join(config.TESTCASE_BASE, _project, _feature)

        output_dir = state.get("output_dir")
        if not output_dir:
            base_path = dir_prefix
            def _is_dir_empty(d):
                if not os.path.exists(d): return True
                try: return not any(os.path.isfile(os.path.join(d,f)) for f in os.listdir(d))
                except OSError: return True
            if os.path.exists(base_path) and not _is_dir_empty(base_path):
                for n in range(2, 1000):
                    alt = f"{dir_prefix}_{n}"
                    if not os.path.exists(alt) or _is_dir_empty(alt):
                        output_dir = alt; break
                else:
                    from datetime import datetime
                    output_dir = f"{dir_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}"
            else:
                output_dir = base_path
        os.makedirs(output_dir, exist_ok=True)
        excel_path = os.path.join(output_dir, "test_plan.xlsx")

        # Phase B 资源冲突消解（纯代码，LLM 输出 → Excel 写入之间）
        # 使用包含全部 confirmed 用例的临时 plan，确保所有轮次的 TC 都经过消解
        if all_shared_pres:
            _full_plan = ExcelPlanV2(
                shared_preconditions=list(all_shared_pres),
                test_cases=list(valid_cases),
            )
            self._resolve_resource_conflicts(_full_plan, all_shared_pres)

        # 写双 Sheet
        n_confirmed = len(valid_cases)
        wb = Workbook()
        hf = Font(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
        tb = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
        wa = Alignment(wrap_text=True, vertical="center")

        # Sheet 1: 测试计划（9列）
        ws1 = wb.active
        ws1.title = "测试计划"
        h1 = ["@allure.epic", "@allure.feature", "@allure.story", "@allure.title",
              "fixture等级", "用例编号", "前置步骤", "执行步骤", "预期结果"]
        for col, h in enumerate(h1, 1):
            c = ws1.cell(row=1, column=col, value=h)
            c.font, c.fill, c.border, c.alignment = hf, hfill, tb, Alignment(horizontal="center", vertical="center")
        for i, tc in enumerate(valid_cases, 2):
            vals = [_project, _feature, tc.story, tc.title, "danyuan", tc.id,
                    ", ".join(tc.preconditions) if tc.preconditions else "无",
                    tc.steps, tc.expected]
            for col, val in enumerate(vals, 1):
                c = ws1.cell(row=i, column=col, value=val); c.border, c.alignment = tb, wa

        # Sheet 2: 共享前置
        ws2 = wb.create_sheet("共享前置")
        h2 = ["前置编号", "前置名称", "详细步骤", "预期结果", "关联用例"]
        for col, h in enumerate(h2, 1):
            c = ws2.cell(row=1, column=col, value=h)
            c.font, c.fill, c.border, c.alignment = hf, hfill, tb, Alignment(horizontal="center", vertical="center")
        pre_to_cases = {}
        for tc in valid_cases:
            for pid in tc.preconditions:
                pre_to_cases.setdefault(pid, []).append(tc.id)
        # 去重: PRE 关联用例列表按首次出现顺序去重
        for pid in pre_to_cases:
            _seen_linked = set()
            _deduped_linked = []
            for cid in pre_to_cases[pid]:
                if cid not in _seen_linked:
                    _seen_linked.add(cid)
                    _deduped_linked.append(cid)
            pre_to_cases[pid] = _deduped_linked
        # 去重: shared_preconditions 按 ID 去重
        _seen_pre = set()
        _deduped_pres = []
        for pre in all_shared_pres:
            if pre.id not in _seen_pre:
                _seen_pre.add(pre.id)
                _deduped_pres.append(pre)
        if len(_deduped_pres) < len(all_shared_pres):
            logger.warning(f"   ⚠️ 共享前置去重: {len(all_shared_pres)} → {len(_deduped_pres)}")
        all_shared_pres = _deduped_pres
        for i, pre in enumerate(all_shared_pres, 2):
            linked = ", ".join(pre_to_cases.get(pre.id, []))
            vals = [pre.id, pre.name, pre.steps, pre.expected, linked or "（无引用）"]
            for col, val in enumerate(vals, 1):
                c = ws2.cell(row=i, column=col, value=val); c.border, c.alignment = tb, wa

        for ws in (ws1, ws2):
            for ci, h in enumerate([c.value for c in ws[1]], 1):
                mx = max((len(str(ws.cell(r, ci).value or "")) for r in range(2, ws.max_row + 1)), default=0)
                ws.column_dimensions[get_column_letter(ci)].width = max(len(str(h)) + 2, min(mx + 2, 55))
        wb.save(excel_path); wb.close()

        # 接口定义快照与 Excel 同目录落盘 —— Phase C 生成 YAML 的数据来源。
        # 规则 M8：接口定义靠产物传递（快照随计划走），禁止依赖内存态跨阶段交接。
        api_defs_path = os.path.join(output_dir, "api_defs.json")
        try:
            with open(api_defs_path, "w", encoding="utf-8") as f:
                json.dump(api_full_for_snapshot, f, ensure_ascii=False, indent=2)
            logger.info(f"   📄 接口定义快照已保存: {api_defs_path} ({len(api_full_for_snapshot)} 个接口)")
        except OSError:
            logger.error("接口定义快照写入失败（Phase C 确认时将按 M8 阻断）: %s",
                         api_defs_path, exc_info=True)

        fail_warn = f"（{len(failed_details)} 行未通过校验，需人工审查）" if failed_details else ""
        n_modules = len(set(tc.story for tc in valid_cases))
        n_pres = len(all_shared_pres)
        logger.info(f"   📄 Excel 已保存: {excel_path} ({n_confirmed}条/{n_modules}模块, {n_pres}共享前置){fail_warn}")

        self._log_node_output("generate_excel_plan",
                              {"excel_plan": {
                                  "shared_preconditions": [p.model_dump() for p in all_shared_pres],
                                  "test_cases": [tc.model_dump() for tc in valid_cases],
                              }, "excel_path": excel_path, "output_dir": output_dir})
        file_ok, file_errors = validate_excel_file(excel_path)
        if not file_ok:
            logger.warning(f"   ⚠️ 文件校验失败: {len(file_errors)} 个错误")
            from observability import log_thinking
            log_thinking("generate_excel_plan_FILE_FAIL",
                         state.get("original_input", "?"),
                         f"文件层校验失败（{n_confirmed} 行通过）\n文件: {excel_path}\n错误: {'; '.join(file_errors)}",
                         prompt_label="generate_excel_plan_node")
        return {
            "excel_plan": plan, "excel_path": excel_path, "output_dir": output_dir,
            "response_obj": ProperResponse(
                proper_thinking=[f"已提取 {len(api_full_for_snapshot)} 个接口，分析 {n_confirmed} 条用例"],
                final_response=f"Excel 测试计划已生成：共 {n_confirmed} 条用例{fail_warn}",
                worth_to_remember=False,
            ),
        }

    # ==================== Phase B 资源冲突消解 ====================

    def _resolve_resource_conflicts(self, plan: ExcelPlanV2,
                                     shared_pres: list = None) -> None:
        """资源冲突消解：检测同一 PRE 被多个正向写操作用例引用时，克隆隔离。

        纯代码节点，不调用 LLM。嵌入在 _generate_excel_plan_node 内部，
        shared_pres 为初始轮保存的共享前置列表（修复轮 plan 可能为空）。
        LLM 生成 → 校验 → 消解 → 写 Excel 的流程中执行。

        算法:
          1. 关键词兜底 LLM 漏标（mutates_data 未标但 steps 含写操作关键词）
          2. 构建 PRE → 正向写操作用例列表
          3. 同一 PRE 被 ≥2 个正向写操作用例引用 → 克隆隔离
        """
        if not plan or not plan.test_cases:
            return

        # 1. 代码兜底 LLM 漏标
        for tc in plan.test_cases:
            if not tc.preconditions or tc.mutates_data:
                continue
            if any(kw in tc.steps for kw in config.RESOURCE_MUTATE_KEYWORDS):
                tc.mutates_data = True
                logger.debug(
                    "消解器兜底: %s 未标 mutates_data，但步骤含写操作关键词，已自动标记",
                    tc.id,
                )

        # 2. 构建 PRE → 正向写操作用例列表
        pre_refs: dict[str, list] = defaultdict(list)
        for tc in plan.test_cases:
            if not tc.mutates_data or tc.is_negative_test:
                continue
            for pid in tc.preconditions:
                pre_refs[pid].append(tc)

        # 3. 检测冲突 → 克隆隔离
        isolation_count = 0
        for pre_id, ref_list in pre_refs.items():
            if len(ref_list) <= 1:
                continue
            _pre_list = shared_pres if shared_pres else plan.shared_preconditions
            original = next((p for p in _pre_list if p.id == pre_id), None)
            if original is None:
                logger.warning("消解器: PRE %s 被 %d 个用例引用但未在 shared_preconditions 中找到，跳过",
                               pre_id, len(ref_list))
                continue
            # 第一个用例保持引用原始 PRE，其余克隆隔离
            for tc in ref_list[1:]:
                clone_id = f"{pre_id}_isolated_{tc.id}"
                _pre_list.append(SharedPrecondition(
                    id=clone_id,
                    name=f"{original.name}（{tc.id}专用）",
                    steps=original.steps,
                    expected=original.expected,
                    cloned_from=pre_id,
                ))
                tc.preconditions = [
                    clone_id if p == pre_id else p for p in tc.preconditions
                ]
                isolation_count += 1
                logger.info("消解器: %s → %s（%s 隔离）", pre_id, clone_id, tc.id)

        if isolation_count:
            logger.info("消解器完成: %d 个 PRE 被隔离，共 %d 条用例受影响",
                        len([p for p, r in pre_refs.items() if len(r) > 1]),
                        isolation_count)

    # ==================== 日志辅助方法（实现已外移，此处薄转发保持签名） ====================

    @staticmethod
    def _split_thinking_sections(text: str) -> dict:
        """将 thinking 分析输出按三个段落拆分（实现见 graph_logging.py）。"""
        return split_thinking_sections(text)

    def _prepare_plan_prompt_vars(self, state: State) -> dict:
        """生成/修复共用的 prompt 变量构造（实现见 prompt_builder.py）。"""
        return prepare_plan_prompt_vars(self, state)

    @staticmethod
    def _serialize_for_log(obj):
        """递归序列化对象为 JSON 可序列化格式（实现见 graph_logging.py）。"""
        return serialize_for_log(obj)

    def _log_node_output(self, node_name: str, output: dict):
        """将节点产出物累积到运行日志（实现见 graph_logging.py）。"""
        return log_node_output(self, node_name, output)

    @staticmethod
    def _cleanup_logs(log_dir: str, max_pairs: int = 15):
        """保留最多 max_pairs 组工作流日志（实现见 graph_logging.py）。"""
        return cleanup_logs(log_dir, max_pairs)

    @staticmethod
    def _load_factory_methods() -> str:
        """数据工厂方法清单（prompt 注入文本），实现见 llm_client.py。"""
        return load_factory_methods()

    def _invoke_think(self, bound_llm, messages, max_retries: int | None = None,
                      label: str = "LLM") -> str:
        """通用 thinking 调用（空 content 复用输入重试），实现见 llm_client.py。"""
        return invoke_think(bound_llm, messages, max_retries=max_retries, label=label)

    def _invoke_structured(self, prompt, model_class,
                           max_retries: int = config.MAX_RETRIES,
                           method: str = "function_calling",
                           thinking: bool = False,
                           temperature: float | None = None,
                           log_label: str = "",
                           pre_validate=None,
                           **kwargs):
        """调用 LLM 并校验结构化输出（实现见 llm_client.py）。"""
        return invoke_structured(
            self.llm, prompt, model_class, METHOD_FEATURES,
            max_retries=max_retries, method=method, thinking=thinking,
            temperature=temperature, log_label=log_label,
            pre_validate=pre_validate, **kwargs,
        )
