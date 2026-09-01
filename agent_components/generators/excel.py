"""Phase C: Excel 测试计划读取 + dependency_map 生成 Mixin

拆分自 generators/__init__.py（2026-08-07 大文件拆分）。

依赖宿主类提供: self.llm / self._invoke_think / self._load_factory_methods
"""
import json
import os
import re

import infrastructure.config as config
from infrastructure.observability import get_logger

logger = get_logger(__name__)


class ExcelMixin:
    """Excel 用例读取 + Phase C Step 0 依赖映射生成"""

    def _generate_dependency_map(self, excel_path: str, output_dir: str,
                                  api_defs_json: str, module_tree_json: str,
                                  product_docs_json: str, context_note: str,
                                  user_ctx: str, repair_cases: list | None = None,
                                  repair_stories: list | None = None,
                                  analysis: str = "") -> str:
        """Phase C Step 0: 生成 dependency_map.json（D5 支持补漏修复）。

        调用 LLM（thinking 模式）分析 Excel 测试计划 + 接口定义 + 模块树，
        输出结构化依赖映射表，经 Pydantic 校验后原子写入。

        补漏模式（repair_cases / repair_stories 任一非空）：
          用 repair_dependency_map_prompt，注入 analysis（Phase B 模块分析）辅助；
          补 repair_cases 的 case 级三表，或补 repair_stories 的 story_pre/teardown；
          输出 partial 供上层 merge。

        Args:
            excel_path: test_plan.xlsx 路径
            output_dir: 输出目录
            api_defs_json: 接口定义 JSON 字符串
            module_tree_json: 模块树 JSON 字符串
            product_docs_json: 产品文档 JSON 字符串
            context_note: 上下文备注（注入 prompt）
            user_ctx: 用户原始意图/上下文
            repair_cases: 补漏模式下待补的用例行列表（含 case_id），None=首次全量生成
            repair_stories: 补漏模式下待补 pre/teardown 的 story 列表
                            [{story_name, preconditions}]，None=首次全量生成
            analysis: 补漏模式下注入的 Phase B 模块分析文本

        Returns:
            dependency_map.json 的绝对路径
        """
        from infrastructure.observability import log_phase_header, log_thinking
        from prompts.extraction_prompts import (
            generate_dependency_map_prompt, repair_dependency_map_prompt)
        from prompts.response_model import DependencyMap

        is_repair = bool(repair_cases) or bool(repair_stories)
        log_phase_header("Phase C Step 0 — 生成依赖映射表"
                         + ("（补漏修复）" if is_repair else ""))
        logger.info("\n🗺️  生成 dependency_map.json ..."
                    + ("（补漏修复）" if is_repair else ""))

        # json_schema 注入（与 pydantic 模型同源，避免 prompt 骨架漂移）
        json_schema_text = json.dumps(
            DependencyMap.model_json_schema(), ensure_ascii=False, indent=2)

        # 读取 Excel 用例行（补漏模式 repair 走 repair_cases/repair_stories 直传，不需过滤）
        rows = self._read_excel_rows(excel_path)
        if not is_repair and not rows:
            raise ValueError("Excel 中无用例数据，无法生成 dependency_map")

        excel_rows_json = json.dumps([
            {k: v for k, v in r.items()}
            for r in rows
        ], ensure_ascii=False, indent=2)

        factory_methods_text = self._load_factory_methods()

        prompt = (repair_dependency_map_prompt() if is_repair
                  else generate_dependency_map_prompt())
        llm_kwargs = {"extra_body": {"thinking": {"type": "enabled"}}}
        bound_llm = self.llm.bind(**llm_kwargs)

        last_error = None
        for attempt in range(1 + config.YAML_REPAIR_ROUNDS):
            try:
                if is_repair:
                    _prompt_vars = dict(
                        data_factory_methods=factory_methods_text,
                        all_apis_info=api_defs_json,
                        module_tree=module_tree_json,
                        product_docs=product_docs_json,
                        context_note=context_note,
                        user_context=user_ctx,
                        json_schema=json_schema_text,
                        repair_cases=json.dumps(repair_cases or [],
                                                ensure_ascii=False, indent=2),
                        repair_stories=json.dumps(repair_stories or [],
                                                  ensure_ascii=False, indent=2),
                        analysis=analysis,
                    )
                else:
                    _prompt_vars = dict(
                        data_factory_methods=factory_methods_text,
                        all_apis_info=api_defs_json,
                        excel_rows=excel_rows_json,
                        module_tree=module_tree_json,
                        product_docs=product_docs_json,
                        context_note=context_note,
                        user_context=user_ctx,
                        json_schema=json_schema_text,
                    )
                # 空 content 走公共方法 _invoke_think（max_retries=0：外层已有重试循环，避免双重重试）
                raw_text = self._invoke_think(
                    bound_llm,
                    prompt.format_messages(**_prompt_vars),
                    max_retries=0,
                    label="generate_dependency_map_repair" if is_repair
                          else "generate_dependency_map",
                )

                # 提取 JSON（LLM 可能在 JSON 外面包了 markdown 代码块）
                json_text = raw_text.strip()
                if json_text.startswith("```"):
                    # 去掉 ```json ... ``` 包裹
                    json_text = re.sub(r"^```(?:json)?\s*\n?", "", json_text)
                    json_text = re.sub(r"\n?```\s*$", "", json_text)

                parsed = json.loads(json_text)
                dep_map = DependencyMap(**parsed)

                log_thinking(
                    "generate_dependency_map_repair" if is_repair else "generate_dependency_map",
                    f"{len(dep_map.stories)} stories",
                    json.dumps(parsed, ensure_ascii=False, indent=2)[:8000],
                    prompt_label="repair_dependency_map_prompt" if is_repair
                                 else "generate_dependency_map_prompt")

                # 原子写入
                dep_map_path = os.path.join(output_dir, "dependency_map.json")
                tmp_path = dep_map_path + ".tmp"
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass
                with open(tmp_path, "w", encoding="utf-8") as f:
                    json.dump(dep_map.model_dump(exclude_none=True),
                              f, ensure_ascii=False, indent=2)
                os.replace(tmp_path, dep_map_path)

                logger.info(f"   ✅ dependency_map.json 已生成: {len(dep_map.stories)} stories"
                            + ("（补漏修复）" if is_repair else ""))
                return dep_map_path

            except (json.JSONDecodeError, Exception) as e:
                last_error = e
                err_text = str(e)
                if attempt < config.YAML_REPAIR_ROUNDS:
                    logger.warning(
                        f"   ⚠️ dependency_map 生成/校验失败，第 {attempt + 1} 次重试: "
                        f"{type(e).__name__}: {err_text[:200]}"
                    )
                else:
                    logger.error(f"   ❌ dependency_map 生成失败（已重试 {attempt} 次）")
                    log_thinking("generate_dependency_map_FAILED",
                                 user_ctx,
                                 f"最终失败: {err_text[:2000]}",
                                 prompt_label="generate_dependency_map_prompt")

        raise RuntimeError(
            f"dependency_map.json 生成失败（已重试 {config.YAML_REPAIR_ROUNDS} 次）: {last_error}"
        )

    @staticmethod
    def _read_excel_rows(excel_path: str, enabled_only: bool = False) -> list[dict]:
        """读取 Excel V2 测试计划（9 列双 Sheet），返回 dict 列表。

        Sheet1 列: @allure.epic, @allure.feature, @allure.story, @allure.title,
                   fixture等级, 用例编号, 前置步骤, 执行步骤, 预期结果
        Sheet2: 共享前置（由 _read_shared_preconditions 独立读取）
        """
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        try:
            ws = wb.active  # Sheet1: 测试计划
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                rows.append({
                    "epic": row[0],          # @allure.epic
                    "feature": row[1],       # @allure.feature
                    "story": row[2],         # @allure.story
                    "title": row[3],         # @allure.title
                    "fixture_level": row[4], # fixture等级
                    "case_id": row[5],       # 用例编号 TC-xxx
                    "preconditions": row[6], # 前置步骤
                    "steps": row[7],         # 执行步骤
                    "expected": row[8],      # 预期结果
                })
            return rows
        finally:
            wb.close()

    @staticmethod
    def _read_shared_preconditions(excel_path: str) -> list[dict]:
        """读取 Excel V2 Sheet2（共享前置），返回 dict 列表。

        Sheet2 列: 前置编号, 前置名称, 详细步骤, 预期结果, 关联用例
        """
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        try:
            if "共享前置" not in wb.sheetnames:
                return []
            ws = wb["共享前置"]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                rows.append({
                    "id": row[0],           # 前置编号 PRE-xxx
                    "name": row[1],          # 前置名称
                    "steps": row[2],         # 详细步骤
                    "expected": row[3],      # 预期结果
                    "linked_cases": row[4],  # 关联用例（逗号分隔）
                })
            return rows
        finally:
            wb.close()
