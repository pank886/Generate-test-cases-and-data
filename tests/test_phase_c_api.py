"""Phase C (/confirm-plan) API 集成测试 — 基于 testcase/园区基线/健身房_4 真实测试计划。

验证范围:
  1. 前置资产完整性 — test_plan.xlsx / translation_cache.json 可读且结构正确
  2. /confirm-plan — 任务提交、异常路径（Excel 不存在时后台任务优雅失败）
  3. /task/{id} — Phase C 后台任务轮询
  4. 端到端: confirm-plan → poll → 校验生成的 .py + YAML 产物质量
     （针对历史缺陷: YAML 生成结果为空列表 `[]`、缺失 baseInfo/testCase 结构、
       .py 引用的 YAML 文件实际未生成 等）

运行方式:
  # 确保服务已启动（python web_app.py），然后:
  # 快速校验（不触发 LLM 生成，秒级）:
  pytest tests/test_phase_c_api.py -v -m "not slow"

  # 完整端到端（真实 LLM 生成 63 个 YAML，约 7-20 分钟）:
  pytest tests/test_phase_c_api.py -v -m slow

  # 轮询超时可通过环境变量覆盖（默认 1500 秒）:
  PHASE_C_POLL_TIMEOUT=2400 pytest tests/test_phase_c_api.py -v -m slow

设计要点:
  - 全部使用 httpx 发送 HTTP 请求，与前端 confirmPlan() 行为一致
    （仅传 excel_path，api_defs_json / user_ctx 留空）
  - 产物校验基于磁盘文件 + .py→YAML 引用完整性，不依赖内部实现
  - 端到端执行前自动删除上次生成的产物目录（如 Gym/），保证校验的是
    本次生成结果；输入资产 test_plan.xlsx / translation_cache.json 保留
"""

import json
import os
import re
import shutil
import time

import pytest
import httpx
import yaml

# ============================================================
# 常量：测试资产路径（testcase/园区基线/健身房_4）
# ============================================================

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PLAN_DIR = os.path.join(_PROJECT_ROOT, "testcase", "园区基线", "健身房_4")
EXCEL_PATH = os.path.join(PLAN_DIR, "test_plan.xlsx")
TRANSLATION_CACHE = os.path.join(PLAN_DIR, "translation_cache.json")

# Phase C 全量生成轮询参数（上次全量 63 个 YAML 实测约 6.5 分钟）
PHASE_C_POLL_INTERVAL = 5.0
PHASE_C_POLL_TIMEOUT = float(os.environ.get("PHASE_C_POLL_TIMEOUT", "1500"))


@pytest.fixture(scope="session")
def base_url(request):
    return request.config.getoption("--base-url")


@pytest.fixture(scope="session")
def client(base_url):
    """共享 httpx 客户端（整个 session 复用连接池）。"""
    with httpx.Client(base_url=base_url, timeout=30.0) as c:
        yield c


# ============================================================
# 辅助函数
# ============================================================

def poll_task(client: httpx.Client, task_id: str,
              timeout: float = PHASE_C_POLL_TIMEOUT,
              interval: float = PHASE_C_POLL_INTERVAL) -> dict:
    """轮询 GET /task/{task_id} 直到 completed/failed，返回完整 task 对象。"""
    deadline = time.time() + timeout
    last_msg = ""
    while time.time() < deadline:
        resp = client.get(f"/task/{task_id}")
        if resp.status_code == 404:
            pytest.fail(f"task_id={task_id} 返回 404，任务可能已过期")
        data = resp.json()
        assert data.get("success"), f"轮询失败: {data}"
        task = data["task"]
        status = task.get("status", "unknown")
        if status in ("completed", "failed"):
            return task
        msg = f"{status}/{task.get('progress')}%/{task.get('message')}"
        if msg != last_msg:
            print(f"  [task] {msg}")
            last_msg = msg
        time.sleep(interval)
    pytest.fail(f"轮询超时（>{timeout}s），最近状态: {last_msg}")


def _load_feature_dirs() -> list[str]:
    """从翻译缓存读取 feature_en，映射到磁盘上的 feature 输出目录。"""
    with open(TRANSLATION_CACHE, encoding="utf-8") as f:
        cache = json.load(f)
    return [os.path.join(PLAN_DIR, en) for en in cache.get("feature_en", {}).values()]


def _scan_placeholder_issues(node, path: str) -> list[str]:
    """占位符防线检查（与 response_model.validate_placeholders 同源读注册表）。

    理论上 B 类问题不会落盘（失败不写文件），落盘即防线告警。
    """
    from data_factory.registry import get_validation_rules
    rules = get_validation_rules()
    issues: list[str] = []
    ph_re = re.compile(r"\$\{([^{}]*)\}")
    call_re = re.compile(r"^([A-Za-z_]\w*)\(([^()]*)\)$")

    def _walk(n, p):
        if isinstance(n, str):
            if "{{" in n or "}}" in n:
                issues.append(f"{p}: 含 {{{{}}}} 双花括号: {n[:60]}")
            matches = list(ph_re.finditer(n))
            if n.count("${") > len(matches):
                issues.append(f"{p}: 占位符未闭合: {n[:60]}")
            for m in matches:
                call = call_re.match(m.group(1).strip())
                if not call:
                    issues.append(f"{p}: 占位符格式非法: {m.group(0)[:60]}")
                    continue
                func, args_str = call.group(1), call.group(2)
                rule = rules.get(func)
                if rule is None:
                    issues.append(f"{p}: 非注册表函数 '{func}'")
                    continue
                args = ([a.strip() for a in args_str.split(",")]
                        if args_str.strip() else [])
                mn, mx = rule.get("min_args"), rule.get("max_args")
                if (mn is not None and len(args) < mn) or (mx is not None and len(args) > mx):
                    issues.append(f"{p}: {func} 实参个数 {len(args)} 超出 [{mn},{mx}]")
                enum0 = rule.get("arg0_enum")
                if enum0 and args and args[0].strip("'\"").lower() not in {
                        str(e).lower() for e in enum0}:
                    issues.append(f"{p}: {func} 首参 '{args[0]}' 不在枚举 {enum0}")
        elif isinstance(n, dict):
            for k, v in n.items():
                _walk(v, f"{p}.{k}")
        elif isinstance(n, list):
            for i, v in enumerate(n):
                _walk(v, f"{p}[{i}]")

    _walk(node, path)
    return issues


def _mergeable_validation_count(validation: list) -> int:
    """模拟 response_model 的同类型断言合并，返回合并后的条目数。"""
    merged: dict = {}
    count = 0
    for item in validation:
        if not (isinstance(item, dict) and len(item) == 1):
            count += 1
            continue
        vtype, payload = next(iter(item.items()))
        if not isinstance(payload, dict):
            count += 1
            continue
        bucket = merged.get(vtype)
        if bucket is None:
            merged[vtype] = dict(payload)
            count += 1
        elif any(k in bucket and bucket[k] != v for k, v in payload.items()):
            count += 1  # 真实冲突允许独立
        else:
            bucket.update(payload)
    return count


def _validate_test_data_yaml(path: str) -> list[str]:
    """校验单个 test_data.yaml 的结构质量，返回问题列表（空 = 通过）。

    针对历史缺陷:
      - 文件内容为空列表 `[]`（LLM 输出为空但仍写盘）
      - block 缺 baseInfo（api_name/url/method 必填）
      - method 未小写 / url 带域名
      - testCase 为空 / 缺 case_name / 缺 validation → 无断言的伪用例
    产物评审新增（2026-07-18 用户反馈）:
      - baseInfo 缺 header
      - 空对象占位字段（有 json 仍带 params: {} 等）
      - 可合并的同类型断言被拆成多条
    """
    issues = []
    try:
        with open(path, encoding="utf-8") as f:
            data = yaml.safe_load(f)
    except yaml.YAMLError as e:
        return [f"YAML 解析失败: {e}"]

    if not isinstance(data, list) or not data:
        return [f"应为非空列表，实际: {data!r}"]

    for i, block in enumerate(data):
        if not isinstance(block, dict):
            issues.append(f"block[{i}] 非 dict: {block!r}")
            continue
        base = block.get("baseInfo")
        cases = block.get("testCase")
        block_cases = [c for c in cases if isinstance(c, dict)] if isinstance(cases, list) else []
        if not isinstance(base, dict):
            issues.append(f"block[{i}] 缺 baseInfo")
        else:
            if not base.get("api_name"):
                issues.append(f"block[{i}].baseInfo 缺 api_name")
            url = base.get("url")
            if not url:
                issues.append(f"block[{i}].baseInfo 缺 url")
            elif str(url).startswith(("http://", "https://")):
                issues.append(f"block[{i}].baseInfo url 含域名: {url}")
            method = base.get("method")
            if not method:
                issues.append(f"block[{i}].baseInfo 缺 method")
            elif str(method) != str(method).lower():
                issues.append(f"block[{i}].baseInfo method 未小写: {method}")
            # header 规则：json 体（非上传）必须有 header；data 体必须表单 Content-Type；
            # 仅 params / 上传可省略（公共头 yq-app-code/token 由框架常量注入，不在产物中）
            has_json_body = any(c.get("json") is not None for c in block_cases)
            has_upload = any(isinstance(c.get("json"), dict) and "file" in c["json"]
                             for c in block_cases)
            has_form_body = any(c.get("data") is not None for c in block_cases)
            header = base.get("header")
            ct = ""
            if isinstance(header, dict):
                ct = str(next((v for k, v in header.items()
                               if str(k).lower() == "content-type"), ""))
            if has_json_body and not has_upload and not isinstance(header, dict):
                issues.append(f"block[{i}].baseInfo 缺 header")
            if has_form_body and "x-www-form-urlencoded" not in ct.lower():
                issues.append(
                    f"block[{i}] data 表单体缺 x-www-form-urlencoded Content-Type")
        if not isinstance(cases, list) or not cases:
            issues.append(f"block[{i}].testCase 为空")
            continue
        for j, c in enumerate(cases):
            if not isinstance(c, dict):
                issues.append(f"block[{i}].testCase[{j}] 非 dict")
                continue
            if not c.get("case_name"):
                issues.append(f"block[{i}].testCase[{j}] 缺 case_name")
            validation = c.get("validation")
            if not validation:
                issues.append(f"block[{i}].testCase[{j}] 缺 validation（无断言）")
            # 问题3: 空对象占位（params 仅在已有请求体时视为占位）
            has_body = c.get("json") is not None or c.get("data") is not None
            if c.get("params") == {} and has_body:
                issues.append(f"block[{i}].testCase[{j}].params 为空占位 {{}}")
            for field in ("extract", "input_extract", "extract_list"):
                if c.get(field) == {}:
                    issues.append(f"block[{i}].testCase[{j}].{field} 为空占位 {{}}")
            # 问题2: 可合并的同类型断言被拆分
            if isinstance(validation, list) and len(validation) > 1:
                if _mergeable_validation_count(validation) < len(validation):
                    issues.append(f"block[{i}].testCase[{j}] 同类型断言未合并: {validation}")
        # B1-B4 防线: 占位符注册表检查（失败本不应落盘）
        issues.extend(_scan_placeholder_issues(block, f"block[{i}]"))
    return issues


def _validate_setup_yaml(path: str) -> list[str]:
    """校验 setup/teardown YAML：可解析且非空 + 占位符防线（宽松标准）。"""
    try:
        with open(path, encoding="utf-8") as f:
            data = yaml.safe_load(f)
    except yaml.YAMLError as e:
        return [f"YAML 解析失败: {e}"]
    if data is None or data == [] or data == {}:
        return [f"内容为空: {data!r}"]
    return _scan_placeholder_issues(data, "root")


# ============================================================
# 清理 fixture：每次执行前删除上次生成的产物
# ============================================================

@pytest.fixture()
def clean_previous_output():
    """E2E 执行前删除上次生成的全部产物（feature 输出目录，如 Gym/）。

    保留输入资产: test_plan.xlsx、translation_cache.json。
    目的:
      1. 残留的旧文件会让产物校验误判为"本次生成成功"
      2. 验证 Phase C 从零生成的完整能力（目录创建、.py、全部 YAML）
    """
    removed = []
    plan_root = os.path.realpath(PLAN_DIR)
    for fdir in _load_feature_dirs():
        real = os.path.realpath(fdir)
        # 防御：只删除 PLAN_DIR 内部的目录
        if not real.startswith(plan_root + os.sep):
            pytest.fail(f"清理目标越界，拒绝删除: {real}")
        if os.path.isdir(real):
            shutil.rmtree(real)
            removed.append(fdir)
    # 上次运行的终态错误清单一并清理
    stale_errors = os.path.join(PLAN_DIR, "_generation_errors.json")
    if os.path.exists(stale_errors):
        os.remove(stale_errors)
        removed.append(stale_errors)
    # 清理后确认：feature 目录必须不存在，产物校验才能证明是本次生成
    leftovers = [d for d in _load_feature_dirs() if os.path.exists(d)]
    assert not leftovers, f"清理失败，残留目录: {leftovers}"
    if removed:
        print(f"\n  [cleanup] 已删除上次产物: {[os.path.basename(d) for d in removed]}")
    yield removed


# ============================================================
# 前置资产完整性（快速，不依赖服务端）
# ============================================================

class TestPlanAssets:
    """健身房_4 测试资产完整性校验 — 运行 Phase C 前的前置检查。"""

    def test_excel_exists_and_valid(self):
        """test_plan.xlsx 存在，双 Sheet（测试计划 9 列 + 共享前置 5 列）。"""
        assert os.path.exists(EXCEL_PATH), f"缺少测试计划: {EXCEL_PATH}"
        from openpyxl import load_workbook
        wb = load_workbook(EXCEL_PATH, read_only=True)
        try:
            assert "测试计划" in wb.sheetnames, f"缺少 Sheet '测试计划': {wb.sheetnames}"
            assert "共享前置" in wb.sheetnames, f"缺少 Sheet '共享前置': {wb.sheetnames}"
            ws = wb["测试计划"]
            header = [c.value for c in next(ws.iter_rows(max_row=1))]
            assert header[:9] == [
                "@allure.epic", "@allure.feature", "@allure.story", "@allure.title",
                "fixture等级", "用例编号", "前置步骤", "执行步骤", "预期结果",
            ], f"9 列表头不匹配: {header}"
            assert ws.max_row > 1, "测试计划无数据行"
        finally:
            wb.close()

    def test_translation_cache_valid(self):
        """translation_cache.json 存在且含三层映射（保证文件名幂等）。"""
        assert os.path.exists(TRANSLATION_CACHE), f"缺少翻译缓存: {TRANSLATION_CACHE}"
        with open(TRANSLATION_CACHE, encoding="utf-8") as f:
            cache = json.load(f)
        for key in ("feature_en", "story_en", "title_en"):
            assert isinstance(cache.get(key), dict), f"翻译缓存缺少 {key}"
        assert cache["feature_en"], "feature_en 为空，Phase C 将回退拼音命名"

    def test_api_defs_snapshot_valid(self):
        """api_defs.json 接口定义快照存在且结构合法（缺失时 Phase C 按 M8 阻断）。"""
        p = os.path.join(PLAN_DIR, "api_defs.json")
        assert os.path.exists(p), (
            f"缺少接口定义快照: {p}（重新执行 Phase B，或从 ChromaDB 迁移导出）")
        with open(p, encoding="utf-8") as f:
            defs = json.load(f)
        assert isinstance(defs, list) and defs, "api_defs.json 为空，Phase C 将被阻断"
        for d in defs:
            assert d.get("url") and d.get("method"), f"定义缺 url/method: {d}"
        print(f"  接口定义快照: {len(defs)} 个")

    def test_excel_assertions_parseable(self):
        """预期结果列的断言关键词可被 C6-1 校验通过（否则 Phase C 会阻断 YAML 生成）。"""
        from openpyxl import load_workbook
        pattern = re.compile(r"\[(eq|contains|ne|db)\]", re.IGNORECASE)
        bad = []
        wb = load_workbook(EXCEL_PATH, read_only=True)
        try:
            ws = wb["测试计划"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                case_id, expected = row[5], row[8]
                if not expected:
                    continue
                for idx, line in enumerate(str(expected).split("\n"), 1):
                    line = line.strip()
                    if not line:
                        continue
                    hits = pattern.findall(line)
                    if len(hits) != 1 or re.search(r"\[\[|\]\]|\[\s+\w+\s*\]|\[\s*\w+\s+\]", line):
                        bad.append(f"{case_id} 第{idx}行: {line[:60]}")
        finally:
            wb.close()
        assert not bad, "断言关键词格式非法（将阻断 Phase C YAML 生成）:\n" + "\n".join(bad[:20])


# ============================================================
# /confirm-plan 异常路径（快速，无 LLM 消耗）
# ============================================================

class TestConfirmPlanValidation:
    """/confirm-plan 提交与异常路径验证。"""

    ENDPOINT = "/confirm-plan"

    def test_nonexistent_excel_fails_gracefully(self, client: httpx.Client):
        """指向不存在的 Excel → 接口受理(200+task_id)，后台任务快速 failed 且带错误信息。

        （不触发 LLM，_read_excel_rows 直接抛 FileNotFoundError）
        """
        fake_path = os.path.join(PLAN_DIR, "__nonexistent__", "test_plan.xlsx")
        resp = client.post(self.ENDPOINT, data={"excel_path": fake_path})
        assert resp.status_code == 200, f"期望 200，得到 {resp.status_code}: {resp.text}"
        data = resp.json()
        assert data.get("success") and data.get("task_id"), f"应返回 task_id: {data}"

        task = poll_task(client, data["task_id"], timeout=60, interval=1.0)
        assert task["status"] == "failed", f"不存在的 Excel 应导致任务失败: {task}"
        assert task.get("error"), f"失败任务应携带 error 信息: {task}"
        print(f"  (符合预期) error={task['error'][:80]}")

    def test_unknown_task_id_returns_404(self, client: httpx.Client):
        """轮询不存在的任务 → 404。"""
        resp = client.get("/task/phase_c_no_such_task_id")
        assert resp.status_code == 404


# ============================================================
# 端到端: Phase C 全量生成 + 产物质量校验
# ============================================================

@pytest.mark.slow
class TestPhaseCEndToEnd:
    """confirm-plan → poll → 磁盘产物校验（真实 LLM 调用）。

    执行前通过 clean_previous_output 删除上次生成记录，保证产物校验
    反映的是本次生成结果而非历史残留。
    """

    def test_full_generation_and_artifacts(self, client: httpx.Client,
                                           clean_previous_output):
        """完整链路: 清理旧产物 → 提交 健身房_4 计划 → 任务完成 → 产物齐备且结构合法。"""
        # ---- 1. 提交（与前端 confirmPlan() 一致，仅传 excel_path）----
        resp = client.post("/confirm-plan", data={"excel_path": EXCEL_PATH})
        assert resp.status_code == 200, f"提交失败 {resp.status_code}: {resp.text}"
        data = resp.json()
        assert data.get("success") and data.get("task_id"), f"应返回 task_id: {data}"
        task_id = data["task_id"]
        print(f"\n  task_id={task_id}")
        print(f"  excel={EXCEL_PATH}")

        # ---- 2. 轮询直到完成 ----
        t0 = time.time()
        task = poll_task(client, task_id)
        elapsed = time.time() - t0
        print(f"  耗时 {elapsed:.0f}s, 最终状态: {task['status']}")

        assert task["status"] == "completed", (
            f"Phase C 任务失败: error={task.get('error')} message={task.get('message')}"
        )
        result = task.get("result") or {}
        print(f"  result: {result.get('message')}")

        # ---- 3. 任务回执校验（含修复循环新字段）----
        assert result.get("py_file"), f"回执缺少 py_file: {result}"
        yaml_total = result.get("yaml_total", 0)
        yaml_success = result.get("yaml_success", 0)
        yaml_failed = result.get("yaml_failed", 0)
        yaml_repaired = result.get("yaml_repaired", 0)
        rounds = result.get("yaml_rounds", 0)
        print(f"  YAML: {yaml_success}/{yaml_total} 成功"
              f"（自查修复 {yaml_repaired}，轮次 {rounds}），仍失败 {yaml_failed}")

        # yaml_total==0 意味着断言校验阻断或无任务 — 均为异常
        assert yaml_total > 0, f"YAML 生成任务数为 0（可能被断言校验阻断）: {result}"
        assert yaml_success + yaml_failed == yaml_total, f"回执计数不闭合: {result}"

        # 终态失败：错误清单必须与回执一致，并在报告中如实列出
        errors_json = os.path.join(PLAN_DIR, "_generation_errors.json")
        if yaml_failed > 0:
            assert os.path.exists(errors_json), (
                f"回执报失败 {yaml_failed} 个但缺少错误清单: {errors_json}")
            with open(errors_json, encoding="utf-8") as f:
                entries = json.load(f)
            assert len(entries) == yaml_failed, (
                f"错误清单条数 {len(entries)} 与回执 failed {yaml_failed} 不一致")
            listing = "\n".join(
                f"  - {e['placeholder_id']} | {e['case_id']} | {e['yaml_path']} | "
                f"{e['error'][:100]}" for e in entries)
            pytest.fail(
                f"Phase C 有 {yaml_failed} 个 YAML 经 {rounds} 轮仍生成失败（已知失败清单）:\n"
                f"{listing}")
        else:
            assert not os.path.exists(errors_json), (
                "回执 failed=0 但存在 _generation_errors.json 残留")

        # ---- 4. 磁盘产物校验 ----
        feature_dirs = [d for d in _load_feature_dirs() if os.path.isdir(d)]
        assert feature_dirs, f"未找到任何 feature 输出目录（translation_cache: {TRANSLATION_CACHE}）"

        problems: list[str] = []
        total_yaml_checked = 0

        for fdir in feature_dirs:
            fname = os.path.basename(fdir)

            # 4a. .py 文件
            py_path = os.path.join(fdir, f"test_{fname}.py")
            if not os.path.exists(py_path):
                problems.append(f"[{fname}] 缺少 test_{fname}.py")
                continue
            with open(py_path, encoding="utf-8") as f:
                py_content = f.read()
            if "class Test" not in py_content:
                problems.append(f"[{fname}] .py 中无测试类")
            if "def test_" not in py_content:
                problems.append(f"[{fname}] .py 中无测试方法")

            # 4b. .py → YAML 引用完整性（./testcase/<feature>/... → 磁盘实际路径）
            refs = re.findall(r"\./testcase/((?:[^'\s/]+/)+[^'\s]+\.yaml)", py_content)
            for ref in refs:
                on_disk = os.path.join(PLAN_DIR, *ref.split("/"))
                if not os.path.exists(on_disk):
                    problems.append(f"[{fname}] .py 引用的 YAML 未生成: {ref}")

            # 4c. 每个用例目录的 test_data.yaml 结构质量
            for entry in sorted(os.listdir(fdir)):
                func_dir = os.path.join(fdir, entry)
                if not os.path.isdir(func_dir) or entry == "setup_data":
                    continue
                yaml_path = os.path.join(func_dir, "test_data.yaml")
                if not os.path.exists(yaml_path):
                    problems.append(f"[{fname}/{entry}] 缺少 test_data.yaml")
                    continue
                total_yaml_checked += 1
                for issue in _validate_test_data_yaml(yaml_path):
                    problems.append(f"[{fname}/{entry}] {issue}")

            # 4d. setup_data（有共享前置时才存在）
            setup_dir = os.path.join(fdir, "setup_data")
            if os.path.isdir(setup_dir):
                for entry in sorted(os.listdir(setup_dir)):
                    if not entry.endswith(".yaml"):
                        continue
                    total_yaml_checked += 1
                    for issue in _validate_setup_yaml(os.path.join(setup_dir, entry)):
                        problems.append(f"[{fname}/setup_data/{entry}] {issue}")

        print(f"  已校验 YAML 文件: {total_yaml_checked} 个")
        assert total_yaml_checked > 0, "未在磁盘上找到任何 YAML 产物"

        if problems:
            report = "\n".join(problems[:40])
            more = f"\n... 共 {len(problems)} 个问题" if len(problems) > 40 else ""
            pytest.fail(f"产物质量校验失败（{len(problems)} 个问题）:\n{report}{more}")
