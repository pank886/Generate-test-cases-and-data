"""Phase B + Phase C 单元测试：资源冲突消解、翻译、断言校验。

不依赖 LLM / ChromaDB / 服务端，纯逻辑测试。
"""
import os
import sys
import json
import tempfile
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from prompts.response_model import (
    TestCaseRow, SharedPrecondition, ExcelPlanV2, TranslationResult
)
from agent_components.generators import GenerationMixin
from agent_components.graph.nodes import ChatTestAgentGraph


# ============================================================
# Phase B — 资源冲突消解
# ============================================================

class TestResolveResourceConflicts:
    """B4: _resolve_resource_conflicts 消解器单元测试。"""

    @staticmethod
    def _make_plan(test_cases: list, shared_pres: list = None) -> ExcelPlanV2:
        return ExcelPlanV2(
            shared_preconditions=shared_pres or [],
            test_cases=test_cases,
        )

    @staticmethod
    def _make_tc(tc_id: str, preconditions: list, steps: str,
                 mutates: bool = False, negative: bool = False) -> TestCaseRow:
        return TestCaseRow(
            id=tc_id, story="测试模块", title=f"测试用例{tc_id}",
            preconditions=preconditions, steps=steps,
            expected="1.[eq]成功", mutates_data=mutates,
            is_negative_test=negative,
        )

    @staticmethod
    def _make_pre(pre_id: str, name: str = "测试前置") -> SharedPrecondition:
        return SharedPrecondition(
            id=pre_id, name=name,
            steps=f"创建{name}", expected="创建成功",
        )

    def test_no_conflict_single_case_per_pre(self):
        """单一用例引用 PRE → 不触发隔离。"""
        agent = ChatTestAgentGraph()
        plan = self._make_plan(
            test_cases=[
                self._make_tc("TC-001", ["PRE-001"], "1.调用修改接口", mutates=True),
                self._make_tc("TC-002", ["PRE-002"], "1.调用删除接口", mutates=True),
            ],
            shared_pres=[
                self._make_pre("PRE-001", "设备A"),
                self._make_pre("PRE-002", "设备B"),
            ],
        )
        pre_count_before = len(plan.shared_preconditions)
        agent._resolve_resource_conflicts(plan)
        # 无冲突 → PRE 数量不变
        assert len(plan.shared_preconditions) == pre_count_before
        # TC 引用不变
        assert plan.test_cases[0].preconditions == ["PRE-001"]
        assert plan.test_cases[1].preconditions == ["PRE-002"]

    def test_conflict_two_write_cases_same_pre(self):
        """两个写操作用例引用同一 PRE → 第二个隔离。"""
        agent = ChatTestAgentGraph()
        plan = self._make_plan(
            test_cases=[
                self._make_tc("TC-001", ["PRE-001"], "1.调用删除设备接口", mutates=True),
                self._make_tc("TC-002", ["PRE-001"], "1.调用修改设备接口", mutates=True),
            ],
            shared_pres=[self._make_pre("PRE-001", "设备A")],
        )
        agent._resolve_resource_conflicts(plan)
        # PRE 从 1 个变成 2 个（克隆）
        assert len(plan.shared_preconditions) == 2
        clone = plan.shared_preconditions[1]
        assert clone.cloned_from == "PRE-001"
        assert "TC-002" in clone.name
        assert clone.id == "PRE-001_isolated_TC-002"
        # TC-001 仍引用原始 PRE
        assert plan.test_cases[0].preconditions == ["PRE-001"]
        # TC-002 引用克隆 PRE
        assert plan.test_cases[1].preconditions == ["PRE-001_isolated_TC-002"]

    def test_conflict_three_write_cases(self):
        """三个写操作用例 → 第1个保留，第2/3个隔离。"""
        agent = ChatTestAgentGraph()
        plan = self._make_plan(
            test_cases=[
                self._make_tc("TC-001", ["PRE-001"], "1.删除", mutates=True),
                self._make_tc("TC-002", ["PRE-001"], "1.修改", mutates=True),
                self._make_tc("TC-003", ["PRE-001"], "1.新增", mutates=True),
            ],
            shared_pres=[self._make_pre("PRE-001", "资源X")],
        )
        agent._resolve_resource_conflicts(plan)
        assert len(plan.shared_preconditions) == 3
        assert plan.test_cases[0].preconditions == ["PRE-001"]
        assert plan.test_cases[1].preconditions == ["PRE-001_isolated_TC-002"]
        assert plan.test_cases[2].preconditions == ["PRE-001_isolated_TC-003"]

    def test_negative_test_excluded(self):
        """负向测试不参与写操作冲突检测。"""
        agent = ChatTestAgentGraph()
        plan = self._make_plan(
            test_cases=[
                self._make_tc("TC-001", ["PRE-001"], "1.删除设备", mutates=True, negative=False),
                self._make_tc("TC-002", ["PRE-001"], "1.删除设备", mutates=True, negative=True),
            ],
            shared_pres=[self._make_pre("PRE-001", "设备A")],
        )
        agent._resolve_resource_conflicts(plan)
        # 只有 1 个正向写操作 → 不触发隔离
        assert len(plan.shared_preconditions) == 1

    def test_keyword_fallback_llm_missed(self):
        """LLM 漏标 mutates_data → 关键词兜底自动标记。"""
        agent = ChatTestAgentGraph()
        plan = self._make_plan(
            test_cases=[
                # LLM 没标 mutates_data，但步骤含"删除"关键词
                self._make_tc("TC-001", ["PRE-001"], "1.调用删除设备接口\n2.校验结果",
                              mutates=False),
                self._make_tc("TC-002", ["PRE-001"], "1.调用修改设备接口\n2.校验结果",
                              mutates=False),
            ],
            shared_pres=[self._make_pre("PRE-001", "设备A")],
        )
        agent._resolve_resource_conflicts(plan)
        # 关键词兜底已标记
        assert plan.test_cases[0].mutates_data is True
        assert plan.test_cases[1].mutates_data is True
        # 触发隔离
        assert len(plan.shared_preconditions) == 2

    def test_keyword_fallback_post_and_put(self):
        """HTTP 方法关键词（POST, DELETE, PUT）有效兜底。"""
        agent = ChatTestAgentGraph()
        plan = self._make_plan(
            test_cases=[
                self._make_tc("TC-001", ["PRE-001"], "1.POST /api/device 创建", mutates=False),
                self._make_tc("TC-002", ["PRE-001"], "1.DELETE /api/device 删除", mutates=False),
            ],
            shared_pres=[self._make_pre("PRE-001", "设备")],
        )
        agent._resolve_resource_conflicts(plan)
        assert plan.test_cases[0].mutates_data is True
        assert plan.test_cases[1].mutates_data is True
        assert len(plan.shared_preconditions) == 2

    def test_readonly_query_not_marked(self):
        """纯查询步骤不被关键词误标。"""
        agent = ChatTestAgentGraph()
        plan = self._make_plan(
            test_cases=[
                self._make_tc("TC-001", ["PRE-001"], "1.调用查询接口获取列表\n2.校验列表非空",
                              mutates=False),
                self._make_tc("TC-002", ["PRE-001"], "1.调用查询接口获取详情\n2.校验详情正确",
                              mutates=False),
            ],
            shared_pres=[self._make_pre("PRE-001", "数据")],
        )
        agent._resolve_resource_conflicts(plan)
        # 查询不含写操作关键词 → 不触发隔离
        assert len(plan.shared_preconditions) == 1

    def test_empty_plan_safe(self):
        """空 plan 不抛异常。"""
        agent = ChatTestAgentGraph()
        plan = ExcelPlanV2(shared_preconditions=[], test_cases=[])
        agent._resolve_resource_conflicts(plan)  # 不应抛异常

    def test_no_shared_preconditions_safe(self):
        """无共享前置时不抛异常。"""
        agent = ChatTestAgentGraph()
        plan = self._make_plan(
            test_cases=[self._make_tc("TC-001", [], "1.查询", mutates=False)],
            shared_pres=[],
        )
        agent._resolve_resource_conflicts(plan)

    def test_multi_pre_conflict_only_isolates_conflicting(self):
        """多 PRE 引用，只隔离冲突的 PRE。"""
        agent = ChatTestAgentGraph()
        plan = self._make_plan(
            test_cases=[
                self._make_tc("TC-001", ["PRE-001", "PRE-002"], "1.删除设备", mutates=True),
                self._make_tc("TC-002", ["PRE-001", "PRE-002"], "1.修改设备", mutates=True),
            ],
            shared_pres=[self._make_pre("PRE-001", "设备"), self._make_pre("PRE-002", "房间")],
        )
        agent._resolve_resource_conflicts(plan)
        # PRE-001 和 PRE-002 各被隔离 1 次
        assert len(plan.shared_preconditions) == 4
        # TC-002 的 preconditions 都变成了隔离版
        assert plan.test_cases[1].preconditions == [
            "PRE-001_isolated_TC-002", "PRE-002_isolated_TC-002"
        ]


# ============================================================
# Phase C — 断言校验 (C6-1)
# ============================================================

class TestAssertionParsing:
    """C6-1: _parse_assertion 断言关键词解析。"""

    def test_eq_keyword(self):
        kw, rest = GenerationMixin._parse_assertion("[eq]接口返回成功，code=0")
        assert kw == "eq"
        assert rest == "接口返回成功，code=0"

    def test_contains_keyword(self):
        kw, rest = GenerationMixin._parse_assertion("[contains]列表包含设备")
        assert kw == "contains"

    def test_ne_keyword(self):
        kw, rest = GenerationMixin._parse_assertion("[ne]id不等于已删除ID")
        assert kw == "ne"

    def test_db_keyword(self):
        kw, rest = GenerationMixin._parse_assertion("[db]数据库中存在记录")
        assert kw == "db"

    def test_case_insensitive(self):
        """不区分大小写。"""
        kw1, _ = GenerationMixin._parse_assertion("[EQ]成功")
        assert kw1 == "eq"
        kw2, _ = GenerationMixin._parse_assertion("[Eq]成功")
        assert kw2 == "eq"
        kw3, _ = GenerationMixin._parse_assertion("[Contains]包含")
        assert kw3 == "contains"

    def test_space_in_brackets_rejected(self):
        """关键词内含空格 → 抛异常。"""
        with pytest.raises(GenerationMixin.AssertionParseError, match="含空格"):
            GenerationMixin._parse_assertion("[ eq ]接口成功")

    def test_double_brackets_rejected(self):
        """双层括号 → 抛异常。"""
        with pytest.raises(GenerationMixin.AssertionParseError, match="双层括号"):
            GenerationMixin._parse_assertion("[[eq]]成功")

    def test_missing_keyword_rejected(self):
        """无断言关键词 → 抛异常。"""
        with pytest.raises(GenerationMixin.AssertionParseError, match="未找到断言关键词"):
            GenerationMixin._parse_assertion("接口返回成功")

    def test_multiple_keywords_takes_first(self):
        """同一步骤多个断言关键词：取第一个，其余作为文本保留（2026-08-03 对齐当前实现）。"""
        kw, rest = GenerationMixin._parse_assertion("[eq]success[contains]data")
        assert kw == "eq"
        assert rest == "success[contains]data"

    def test_parens_rejected(self):
        """圆括号 → 抛异常。"""
        with pytest.raises(GenerationMixin.AssertionParseError):
            GenerationMixin._parse_assertion("(eq)成功")

    def test_no_brackets_rejected(self):
        """无方括号 → 抛异常。"""
        with pytest.raises(GenerationMixin.AssertionParseError):
            GenerationMixin._parse_assertion("eq成功")

    def test_keyword_at_middle(self):
        """关键词在中间位置也正常解析。"""
        kw, rest = GenerationMixin._parse_assertion("返回 [eq]结果一致")
        assert kw == "eq"
        assert rest == "结果一致"


# ============================================================
# Phase C — Sanitize (C4-1)
# ============================================================

class TestSanitizeEn:
    """C4-1: _sanitize_en 标识符清洗。"""

    def test_keeps_alphanumeric(self):
        assert GenerationMixin._sanitize_en("FacilityManagement") == "FacilityManagement"

    def test_removes_special_chars(self):
        # 全中文 → 全移除 → 空字符串补 _ 防非法 identifier
        assert GenerationMixin._sanitize_en("设施管理") == "_"

    def test_replaces_spaces(self):
        assert " " not in GenerationMixin._sanitize_en("Facility Management")

    def test_digit_start_prepends_underscore(self):
        result = GenerationMixin._sanitize_en("123Facility")
        assert result.startswith("_")

    def test_keeps_underscore(self):
        result = GenerationMixin._sanitize_en("facility_add_positive_001")
        assert result == "facility_add_positive_001"


# ============================================================
# Phase C — Translation Cache (C4-1)
# ============================================================

class TestTranslationCache:
    """C4-1: 翻译缓存读写。"""

    def test_load_nonexistent_cache(self):
        cache = GenerationMixin._load_translation_cache("/nonexistent/path/test.xlsx")
        assert cache == {}

    def test_save_and_load(self):
        with tempfile.TemporaryDirectory() as td:
            xlsx_path = os.path.join(td, "test_plan.xlsx")
            # 创建伪 Excel 文件
            with open(xlsx_path, "w") as f:
                f.write("dummy")
            cache_data = {
                "feature_en": {"设施管理": "FacilityManagement"},
                "story_en": {},
                "title_en": {},
            }
            GenerationMixin._save_translation_cache(xlsx_path, cache_data)
            loaded = GenerationMixin._load_translation_cache(xlsx_path)
            assert loaded["feature_en"]["设施管理"] == "FacilityManagement"


# ============================================================
# Phase C — Excel 读取 (C3)
# ============================================================

class TestReadExcelV2:
    """C3: 9 列 Excel V2 读取。"""

    def test_read_9_column_excel(self):
        """创建 V2 格式 Excel 并读取。"""
        from openpyxl import Workbook
        with tempfile.TemporaryDirectory() as td:
            path = os.path.join(td, "test_plan.xlsx")
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "测试计划"
            headers = ["@allure.epic", "@allure.feature", "@allure.story", "@allure.title",
                       "fixture等级", "用例编号", "前置步骤", "执行步骤", "预期结果"]
            for c, h in enumerate(headers, 1):
                ws1.cell(row=1, column=c, value=h)
            ws1.cell(row=2, column=1, value="健身房管理")
            ws1.cell(row=2, column=2, value="设施管理")
            ws1.cell(row=2, column=3, value="设施添加")
            ws1.cell(row=2, column=4, value="设施管理-新增设施-正向")
            ws1.cell(row=2, column=5, value="danyuan")
            ws1.cell(row=2, column=6, value="TC-001")
            ws1.cell(row=2, column=7, value="PRE-001, PRE-002")
            ws1.cell(row=2, column=8, value="1.调用新增接口\n2.查询详情")
            ws1.cell(row=2, column=9, value="1.[eq]创建成功\n2.[eq]信息一致")

            # Sheet2
            ws2 = wb.create_sheet("共享前置")
            h2 = ["前置编号", "前置名称", "详细步骤", "预期结果", "关联用例"]
            for c, h in enumerate(h2, 1):
                ws2.cell(row=1, column=c, value=h)
            ws2.cell(row=2, column=1, value="PRE-001")
            ws2.cell(row=2, column=2, value="已创建测试跑步机")
            ws2.cell(row=2, column=3, value="1.调用新增设施接口\n2.校验创建成功")
            ws2.cell(row=2, column=4, value="设施列表中出现测试跑步机")
            ws2.cell(row=2, column=5, value="TC-001, TC-002")
            wb.save(path)
            wb.close()

            rows = GenerationMixin._read_excel_rows(path)
            assert len(rows) == 1
            r = rows[0]
            assert r["epic"] == "健身房管理"
            assert r["feature"] == "设施管理"
            assert r["story"] == "设施添加"
            assert r["title"] == "设施管理-新增设施-正向"
            assert r["fixture_level"] == "danyuan"
            assert r["case_id"] == "TC-001"
            assert r["preconditions"] == "PRE-001, PRE-002"
            assert "\n" in r["steps"]

            pres = GenerationMixin._read_shared_preconditions(path)
            assert len(pres) == 1
            assert pres[0]["id"] == "PRE-001"
            assert pres[0]["name"] == "已创建测试跑步机"


# ============================================================
# Phase C — 拼音 Fallback (C4-1)
# ============================================================

class TestPinyinFallback:
    """C4-1: _pinyin_fallback 拼音降级。"""

    def test_returns_non_empty(self):
        result = GenerationMixin._pinyin_fallback("设施管理")
        assert len(result) > 0

    def test_consistent_output(self):
        a = GenerationMixin._pinyin_fallback("设施管理")
        b = GenerationMixin._pinyin_fallback("设施管理")
        assert a == b

    def test_real_pinyin_not_hash_fallback(self):
        """pypinyin 已安装时必须输出拼音首字母缩写，而非 hash 兜底命名。

        背景（2026-07-19）：venv 缺 pypinyin 时守护导入静默降级为 hash 命名
        （设施管理 → M36404B7），而上面两条用例只断言非空/一致，缺包照样全绿。
        本用例把"依赖缺失的静默降级"变成显式失败：
          1. pypinyin 必须可导入（requirements.txt 已声明）
          2. 输出必须是纯字母拼音缩写（hash 兜底含数字，isalpha 必假）
        """
        import importlib.util
        assert importlib.util.find_spec("pypinyin") is not None, (
            "pypinyin 未安装 —— requirements.txt 已声明该依赖，"
            "请执行 pip install -r requirements.txt（缺包会静默降级为 hash 文件名）")
        result = GenerationMixin._pinyin_fallback("设施管理")
        assert result == "SSGL", f"应为拼音首字母缩写 SSGL，实际: {result}"
        assert result.isalpha(), f"拼音缩写应为纯字母（hash 兜底含数字）: {result}"


# ============================================================
# Phase B — TestCaseRow 新字段 (B2)
# ============================================================

class TestCaseRowFields:
    """B2: TestCaseRow 新增字段默认值。"""

    def test_defaults(self):
        tc = TestCaseRow(id="TC-001", story="测试", title="测试用例",
                         steps="步骤", expected="预期")
        assert tc.mutates_data is False
        assert tc.is_negative_test is False

    def test_explicit_values(self):
        tc = TestCaseRow(id="TC-002", story="测试", title="修改测试",
                         steps="1.调用修改接口", expected="1.[eq]成功",
                         mutates_data=True, is_negative_test=False)
        assert tc.mutates_data is True
        assert tc.is_negative_test is False


# ============================================================
# Phase C — TranslationResult 模型
# ============================================================

class TestTranslationResult:
    """C4: TranslationResult 模型。"""

    def test_empty_defaults(self):
        tr = TranslationResult()
        assert tr.feature_en == {}
        assert tr.story_en == {}
        assert tr.title_en == {}

    def test_with_data(self):
        tr = TranslationResult(
            feature_en={"设施管理": "FacilityManagement"},
            story_en={"设施添加": "FacilityAdd"},
            title_en={"测试": "test_001"},
        )
        assert tr.feature_en["设施管理"] == "FacilityManagement"


# ============================================================
# SharedPrecondition cloned_from (B4-1)
# ============================================================

class TestSharedPreconditionClone:
    """B4-1: SharedPrecondition cloned_from 字段。"""

    def test_default_none(self):
        pre = SharedPrecondition(id="PRE-001", name="测试", steps="步骤", expected="预期")
        assert pre.cloned_from is None

    def test_clone_marked(self):
        pre = SharedPrecondition(id="PRE-001_isolated_TC-002", name="测试(TC-002专用)",
                                 steps="步骤", expected="预期", cloned_from="PRE-001")
        assert pre.cloned_from == "PRE-001"


# ============================================================
# Phase C — YAML 数据模型兜底（E2E 于 健身房_4 捕获的缺陷）
# ============================================================

from pydantic import ValidationError
from prompts.response_model import (
    TestCase as YamlTestCase,       # 别名避免 pytest 误收集
    TestData as YamlTestData,
    StepData,
)


class TestYamlModelRobustness:
    """TestCase 提取字段强转 + TestData/StepData 空列表拦截。"""

    @staticmethod
    def _make_step(**case_kwargs) -> dict:
        case = {"case_name": "test_x_001", "validation": [{"eq": {"code": 0}}]}
        case.update(case_kwargs)
        return {"baseInfo": {"api_name": "x", "url": "/x", "method": "post",
                             "header": {}},
                "testCase": [case]}

    # ---- 缺陷1(B5/B10): extract 系字段值类型 —— 撤销强转，非 str 一律回炉 ----

    def test_input_extract_int_rejected(self):
        """021 用例失败原型: input_extract: {subscribeId_active: 1} → 校验失败进重生成循环。"""
        with pytest.raises(ValidationError, match="string"):
            YamlTestCase(case_name="t", input_extract={"subscribeId_active": 1})

    def test_extract_float_and_bool_rejected(self):
        with pytest.raises(ValidationError):
            YamlTestCase(case_name="t", extract={"price": 1.5})
        with pytest.raises(ValidationError):
            YamlTestCase(case_name="t", extract_list={"flag": True})

    def test_extract_none_value_rejected(self):
        """B10（用户裁定）: null 条目不静默丢弃，回炉让 LLM 省略字段。"""
        with pytest.raises(ValidationError):
            YamlTestCase(case_name="t", extract={"a": None, "b": "$.data.id"})

    def test_extract_str_values_untouched(self):
        """正常 str 值不受影响。"""
        tc = YamlTestCase(case_name="t", validation=[{"eq": {"code": 0}}],
                          input_extract={"img_url": "$.data.url"})
        assert tc.input_extract == {"img_url": "$.data.url"}

    def test_extract_absent_stays_none(self):
        tc = YamlTestCase(case_name="t", validation=[{"eq": {"code": 0}}])
        assert tc.extract is None
        assert tc.input_extract is None

    # ---- 缺陷2: LLM 输出空列表被静默写成 [] 且计为成功 ----

    def test_testdata_empty_data_rejected(self):
        """data=[] 必须校验失败（触发重试而非写出空 YAML）。"""
        with pytest.raises(ValidationError):
            YamlTestData(data=[])

    def test_stepdata_empty_testcase_rejected(self):
        """testCase=[] 的空块必须校验失败。"""
        with pytest.raises(ValidationError):
            StepData(baseInfo={"url": "/x", "method": "post"}, testCase=[])

    def test_valid_testdata_passes(self):
        """完整合法结构不受两处收紧影响。"""
        td = YamlTestData(data=[self._make_step()])
        assert len(td.data) == 1
        assert td.data[0].testCase[0].case_name == "test_x_001"

    def test_valid_testdata_with_str_extract_passes(self):
        """合法 str 提取值与 min_length 组合场景。"""
        td = YamlTestData(data=[self._make_step(input_extract={"sid": "$.json.sid"})])
        assert td.data[0].testCase[0].input_extract == {"sid": "$.json.sid"}


# ============================================================
# Phase C — YAML 输出卫生（header 缺失 / 空 params / 断言拆分）
# ============================================================

class TestYamlOutputHygiene:
    """用户在 健身房_4 产物评审中反馈的三类问题的代码兜底。"""

    # ---- 问题3: 已有 json 数据仍带空 params ----

    def test_empty_params_dropped_when_json_present(self):
        """params: {} 置 None，model_dump(exclude_none) 后 YAML 不再出现。"""
        tc = YamlTestCase(case_name="t", json={"pageNum": 1}, params={},
                          validation=[{"eq": {"code": 0}}])
        assert tc.params is None
        dumped = tc.model_dump(exclude_none=True, by_alias=True)
        assert "params" not in dumped
        assert dumped["json"] == {"pageNum": 1}

    def test_empty_extract_fields_dropped(self):
        tc = YamlTestCase(case_name="t", extract={}, input_extract={}, extract_list={},
                          validation=[{"eq": {"code": 0}}])
        dumped = tc.model_dump(exclude_none=True, by_alias=True)
        for field in ("extract", "input_extract", "extract_list"):
            assert field not in dumped

    def test_non_empty_params_kept(self):
        """GET 类接口的有效 params 不受影响。"""
        tc = YamlTestCase(case_name="t", params={"id": "1"}, validation=[{"eq": {"code": 0}}])
        assert tc.params == {"id": "1"}

    def test_empty_json_body_not_dropped(self):
        """json: {} 可能是有语义的空请求体，不做剔除。"""
        tc = YamlTestCase(case_name="t", json={}, validation=[{"eq": {"code": 0}}])
        assert tc.request_body == {}

    # ---- 问题2: 同类型断言拆成多条 ----

    def test_same_type_validations_merged(self):
        tc = YamlTestCase(case_name="t", validation=[
            {"eq": {"code": 0}},
            {"eq": {"msg": "success"}},
        ])
        assert tc.validation == [{"eq": {"code": 0, "msg": "success"}}]

    def test_three_way_merge_with_mixed_types(self):
        """eq+contains 混合时各自归并，类型间保持独立。"""
        tc = YamlTestCase(case_name="t", validation=[
            {"eq": {"code": 0}},
            {"contains": {"data.list": "A"}},
            {"eq": {"total": 10}},
        ])
        assert tc.validation == [
            {"eq": {"code": 0, "total": 10}},
            {"contains": {"data.list": "A"}},
        ]

    def test_conflicting_same_field_kept_separate(self):
        """同字段不同期望值是真实冲突（如 contains 两个不同 ID），不能合并丢断言。"""
        tc = YamlTestCase(case_name="t", validation=[
            {"contains": {"data.list[*].id": "A"}},
            {"contains": {"data.list[*].id": "B"}},
        ])
        assert tc.validation == [
            {"contains": {"data.list[*].id": "A"}},
            {"contains": {"data.list[*].id": "B"}},
        ]

    def test_single_validation_untouched(self):
        tc = YamlTestCase(case_name="t", validation=[{"eq": {"code": 0}}])
        assert tc.validation == [{"eq": {"code": 0}}]

    # ---- 2026-08-20 问题5: 多键断言块必须拒绝（_35 全量被执行框架拒绝的根因） ----

    def test_multi_key_validation_block_rejected(self):
        """check/expected/operator 三键写法 → 校验失败，进修复轮重生成。

        执行框架 assert_result 只认单键 {eq|contains|ne|db: {...}}，
        多键块在此处拦截，防止落盘后整体跑不动。
        """
        with pytest.raises(ValueError, match="断言块必须且只能包含一个断言运算符键"):
            YamlTestCase(case_name="t", validation=[
                {"check": "status_code", "expected": {"status_code": 200},
                 "operator": "contains"},
            ])

    def test_multi_key_jsonpath_validation_block_rejected(self):
        """jsonpath/operator/value 三键写法同样拒绝。"""
        with pytest.raises(ValueError, match="断言块必须且只能包含一个断言运算符键"):
            YamlTestCase(case_name="t", validation=[
                {"jsonpath": "$.msg", "operator": "eq", "value": "成功"},
            ])

    def test_opkey_jsonpath_as_block_key_rejected(self):
        """块键是 JSONPath（{$.retCode: {eq: 1}}）→ 校验失败进修复轮。

        2026-08-25 缺口：块键数量==1 但块键不是合法运算符时旧校验整体短路放行。
        块键必须是运算符 eq/contains/ne/db，JSONPath 写在操作数 dict 的键位。
        """
        with pytest.raises(ValueError, match="不是合法运算符"):
            YamlTestCase(case_name="t", validation=[
                {"$.retCode": {"eq": 1}},
            ])

    def test_opkey_bare_field_as_block_key_rejected(self):
        """裸字段名当块键（{status_code: 200}）→ 校验失败进修复轮。"""
        with pytest.raises(ValueError, match="不是合法运算符"):
            YamlTestCase(case_name="t", validation=[
                {"status_code": 200},
            ])

    def test_opkey_legit_operators_pass(self):
        """合法运算符块键 eq/contains/ne/db 不被误拦。"""
        YamlTestCase(case_name="t", validation=[
            {"eq": {"code": 0}},
            {"contains": {"message": "成功"}},
            {"ne": {"code": 1}},
            {"db": {"sql": "select 1"}},
        ])

    # ---- 问题1: baseInfo 无 header（公共头 yq-app-code/token 由框架常量注入，不生成） ----

    def test_default_header_injected_for_json_body(self):
        step = StepData(
            baseInfo={"api_name": "查询", "url": "/x/query", "method": "post"},
            testCase=[{"case_name": "t", "json": {"pageNum": 1},
                       "validation": [{"eq": {"code": 0}}]}],
        )
        assert step.baseInfo["header"] == {
            "Content-Type": "application/json;charset=UTF-8"}

    def test_params_only_no_header_injected(self):
        """仅 params（GET）不注入 Content-Type；显式 header 保持为空。"""
        step = StepData(
            baseInfo={"api_name": "查询", "url": "/x/search", "method": "get",
                      "header": {}},
            testCase=[{"case_name": "t", "params": {"pageNum": 1},
                       "validation": [{"eq": {"code": 0}}]}],
        )
        assert step.baseInfo["header"] == {}

    def test_existing_header_preserved(self):
        """LLM 已输出 header（含鉴权）时不覆盖。"""
        step = StepData(
            baseInfo={"url": "/x", "method": "post",
                      "header": {"Authorization": "${token}"}},
            testCase=[{"case_name": "t", "validation": [{"eq": {"code": 0}}]}],
        )
        assert step.baseInfo["header"] == {"Authorization": "${token}"}

    def test_upload_interface_no_header_injected(self):
        """上传接口（请求体含 file）不注入 Content-Type，multipart 由客户端生成。"""
        step = StepData(
            baseInfo={"url": "/upload/uploadImg", "method": "post", "header": {}},
            testCase=[{"case_name": "t", "json": {"file": "test_upload.png"},
                       "validation": [{"eq": {"code": 0}}]}],
        )
        assert step.baseInfo["header"] == {}

    # ---- data 表单体仅在 x-www-form-urlencoded 下合法 ----

    def test_form_data_with_form_ct_preserved(self):
        """明确表单 Content-Type → data 是合法表单体，输出仍为 data 字段。"""
        step = StepData(
            baseInfo={"api_name": "表单登录", "url": "/api/login", "method": "post",
                      "header": {"Content-Type": "application/x-www-form-urlencoded"}},
            testCase=[{"case_name": "用户名密码登录",
                       "data": {"username": "admin", "password": "123456"},
                       "validation": [{"eq": {"code": 200}}]}],
        )
        case = step.testCase[0]
        assert case.form_data == {"username": "admin", "password": "123456"}
        assert case.request_body is None
        dumped = case.model_dump(exclude_none=True, by_alias=True)
        assert dumped["data"] == {"username": "admin", "password": "123456"}
        assert "json" not in dumped

    def test_data_without_form_ct_migrates_to_json(self):
        """非表单 Content-Type 下 data 仍视为字段漂移 → 迁移为 json（既有行为保持）。"""
        step = StepData(
            baseInfo={"url": "/x", "method": "post",
                      "header": {"Content-Type": "application/json;charset=UTF-8"}},
            testCase=[{"case_name": "t", "data": {"k": "v"},
                       "validation": [{"eq": {"code": 0}}]}],
        )
        assert step.testCase[0].request_body == {"k": "v"}
        assert step.testCase[0].form_data is None

    def test_data_without_header_treated_as_json_drift(self):
        """无 header 时 data 视为漂移：注入 json Content-Type 并迁移为 json。"""
        step = StepData(
            baseInfo={"url": "/x", "method": "post"},
            testCase=[{"case_name": "t", "data": {"k": "v"},
                       "validation": [{"eq": {"code": 0}}]}],
        )
        assert step.baseInfo["header"] == {
            "Content-Type": "application/json;charset=UTF-8"}
        assert step.testCase[0].request_body == {"k": "v"}

    # ---- method 小写 / url 不含域名 ----

    def test_uppercase_method_lowered(self):
        step = StepData(
            baseInfo={"url": "/x", "method": "POST", "header": {}},
            testCase=[{"case_name": "t", "validation": [{"eq": {"code": 0}}]}],
        )
        assert step.baseInfo["method"] == "post"

    def test_full_url_stripped_to_path(self):
        step = StepData(
            baseInfo={"url": "https://park.example.com/gymFacility/add", "method": "post",
                      "header": {}},
            testCase=[{"case_name": "t", "validation": [{"eq": {"code": 0}}]}],
        )
        assert step.baseInfo["url"] == "/gymFacility/add"

    # ---- 三选一 — 无请求体的 GET 保留空 params ----

    def test_empty_params_kept_when_no_body(self):
        """无 json/data 时空 params 保留（满足 json/params/data 三选一必填）。"""
        tc = YamlTestCase(case_name="t", params={}, validation=[{"eq": {"code": 0}}])
        assert tc.params == {}

    # ---- 三问题组合的端到端序列化 ----

    def test_full_dump_matches_expected_yaml_shape(self):
        """组合场景：dump 后无空字段、断言归并、header 就位。"""
        td = YamlTestData(data=[{
            "baseInfo": {"api_name": "分页查询", "url": "/getPage", "method": "post"},
            "testCase": [{
                "case_name": "查询列表",
                "json": {"pageNum": 1, "pageSize": 10},
                "params": {},
                "extract": {},
                "validation": [{"eq": {"code": 0}}, {"eq": {"msg": "ok"}}],
            }],
        }])
        block = td.data[0].model_dump(exclude_none=True, by_alias=True)
        assert block["baseInfo"]["header"]["Content-Type"].startswith("application/json")
        case = block["testCase"][0]
        assert "params" not in case and "extract" not in case
        assert case["validation"] == [{"eq": {"code": 0, "msg": "ok"}}]


# ============================================================
# B9 — json/params/data 三选一（回炉类）
# ============================================================

class TestBodyExclusivity:
    """B9: 多个请求体字段非空并存 → 校验失败进重生成循环。"""

    def test_json_and_params_both_nonempty_rejected(self):
        with pytest.raises(ValidationError, match="三选一"):
            YamlTestCase(case_name="t", json={"a": 1}, params={"b": "2"})

    def test_json_and_form_data_rejected(self):
        with pytest.raises(ValidationError, match="三选一"):
            YamlTestCase(case_name="t", json={"a": 1}, form_data={"u": "x"})

    def test_single_body_each_passes(self):
        assert YamlTestCase(case_name="t", json={"a": 1},
                            validation=[{"eq": {"code": 0}}]).request_body == {"a": 1}
        assert YamlTestCase(case_name="t", params={"p": "1"},
                            validation=[{"eq": {"code": 0}}]).params == {"p": "1"}
        assert YamlTestCase(case_name="t", form_data={"u": "x"},
                            validation=[{"eq": {"code": 0}}]).form_data == {"u": "x"}

    def test_json_with_empty_params_passes(self):
        """空 params 占位先被 A7 剔除，不触发三选一。"""
        tc = YamlTestCase(case_name="t", json={"a": 1}, params={},
                          validation=[{"eq": {"code": 0}}])
        assert tc.params is None


# ============================================================
# 数据工厂注册表（data_factory/registry.py + methods.yaml v2）
# ============================================================

from data_factory import registry as df_registry


class TestFactoryRegistry:
    """注册表加载 / 渲染 / 校验规则三处同源。"""

    def setup_method(self):
        df_registry.reset_cache()

    def teardown_method(self):
        df_registry.reset_cache()

    def test_load_methods_covers_all_seven(self):
        names = {m["name"] for m in df_registry.load_methods()}
        assert names == {
            "random_plates", "random_code", "get_extract_data",
            "get_extract_data_list", "get_current_time", "get_offset_time",
            "split_extract_data",
        }

    def test_every_method_has_category(self):
        cats = {m["category"] for m in df_registry.load_methods()}
        assert cats == {"基础类", "数据生成类", "时间类"}

    def test_render_has_catalog_and_details(self):
        text = df_registry.render_for_prompt()
        assert "【数据工厂方法目录】" in text
        assert "【方法详情】" in text
        assert "get_offset_time" in text
        # validation 块不渲染进 prompt
        assert "min_args" not in text and "arg0_enum" not in text

    def test_validation_rules_from_registry(self):
        rules = df_registry.get_validation_rules()
        assert rules["get_offset_time"]["max_args"] == 5
        assert rules["get_current_time"]["arg0_enum"] == ["ydm", "hms"]
        assert rules["get_extract_data"]["min_args"] == 1

    def test_legacy_flat_structure_compatible(self, tmp_path, monkeypatch):
        """旧版扁平 methods: 结构仍可加载（迁移保险）。"""
        p = tmp_path / "methods.yaml"
        p.write_text(
            "methods:\n  - name: old_func\n    syntax: '${old_func(x)}'\n"
            "    description: 旧条目\n", encoding="utf-8")
        monkeypatch.setattr(df_registry, "_registry_path", lambda: str(p))
        df_registry.reset_cache()
        methods = df_registry.load_methods()
        assert [m["name"] for m in methods] == ["old_func"]
        assert methods[0]["category"] == "默认"


# ============================================================
# B1-B4 — 动态占位符校验（注册表驱动）
# ============================================================

class TestPlaceholderValidation:
    """validate_placeholders：{{}} / 运算 / 非注册表函数 / 实参规则。"""

    @staticmethod
    def _td(**case_kwargs) -> YamlTestData:
        case = {"case_name": "t", "validation": [{"eq": {"code": 0}}]}
        case.update(case_kwargs)
        return YamlTestData(data=[{
            "baseInfo": {"api_name": "x", "url": "/x", "method": "post",
                         "header": {}},
            "testCase": [case],
        }])

    def test_user_reported_sample_rejected(self):
        """现场样本: '{{(get_current_time(ymd) + 1day)}} 11:00:00'。"""
        with pytest.raises(ValidationError, match="双花括号"):
            self._td(json={"endTime": "{{(get_current_time(ymd) + 1day)}} 11:00:00"})

    def test_arithmetic_in_placeholder_rejected(self):
        with pytest.raises(ValidationError, match="运算"):
            self._td(json={"t": "${get_current_time(ydm) + 1day}"})

    def test_unknown_function_rejected(self):
        with pytest.raises(ValidationError, match="未知占位符函数"):
            self._td(json={"t": "${tomorrow()}"})

    def test_wrong_fmt_enum_rejected(self):
        """ymd 是 ydm 的错序写法 → 按注册表 arg0_enum 拦截。"""
        with pytest.raises(ValidationError, match="ydm"):
            self._td(json={"t": "${get_current_time(ymd)}"})

    def test_arity_overflow_rejected(self):
        with pytest.raises(ValidationError, match="实参个数"):
            self._td(json={"t": "${get_offset_time(ydm, 1, 2, 3, 4, 5)}"})

    def test_unclosed_placeholder_rejected(self):
        with pytest.raises(ValidationError, match="未闭合"):
            self._td(json={"t": "${get_current_time(ydm"})

    def test_offset_time_concat_literal_passes(self):
        """推荐写法：日期占位符拼固定时分秒。"""
        td = self._td(json={"startTime": "${get_offset_time(ydm, 1)} 10:00:00"})
        assert td.data[0].testCase[0].request_body["startTime"].startswith("${get_offset_time")

    def test_negative_offset_passes(self):
        self._td(json={"t": "${get_offset_time(hms, 0, 0, -30)}"})

    def test_fixed_literal_and_plain_strings_pass(self):
        self._td(json={"endTime": "2029-12-31 10:00:00", "note": "单{括号}无碍"})

    def test_placeholder_in_validation_value_checked(self):
        with pytest.raises(ValidationError, match="未知占位符函数"):
            self._td(validation=[{"contains": {"data.id": "${no_such(x)}"}}])

    def test_registry_sweep_all_methods_valid(self):
        """遍历注册表构造每个方法的最小合法调用，防注册表与校验器脱节。"""
        rules = df_registry.get_validation_rules()
        assert rules, "注册表为空"
        for func, rule in rules.items():
            n = rule.get("min_args") or 1
            args = []
            for i in range(n):
                if i == 0 and rule.get("arg0_enum"):
                    args.append(str(rule["arg0_enum"][0]))
                else:
                    args.append("1")
            expr = "${" + func + "(" + ", ".join(args) + ")}"
            self._td(json={"v": expr})  # 不抛异常即通过


# ============================================================
# 修复循环（_run_yaml_rounds，注入假生成函数，不依赖 LLM）
# ============================================================

class TestYamlRepairLoop:
    """§4: 登记占位 → 修复轮自查 → 终态错误清单。"""

    @staticmethod
    def _tasks(base: str, n: int = 3) -> list:
        return [
            ({"steps": f"步骤{i}", "expected": "", "case_id": f"TC-{i:03d}"},
             os.path.join(base, f"case_{i}", "test_data.yaml"))
            for i in range(n)
        ]

    @staticmethod
    def _ok_gen(row, api, ctx, path, repair_ctx=None):
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            f.write("- ok\n")
        return path

    def test_all_success_single_round(self, tmp_path):
        agent = ChatTestAgentGraph()
        result = agent._run_yaml_rounds(
            self._tasks(str(tmp_path)), "[]", "", str(tmp_path),
            gen_func=self._ok_gen, repair_rounds=1)
        assert result["total"] == 3 and result["success"] == 3
        assert result["failed"] == 0 and result["repaired"] == 0
        assert result["rounds"] == 1 and result["errors_file"] is None
        assert not os.path.exists(os.path.join(str(tmp_path), "_generation_errors.json"))

    def test_fail_once_repaired_in_round2(self, tmp_path):
        """首轮失败 → 修复轮携带自查上下文 → 成功；repair_ctx 内容完整。"""
        seen_ctx = {}

        def flaky(row, api, ctx, path, repair_ctx=None):
            if "case_1" in path and repair_ctx is None:
                raise RuntimeError(
                    'Failed to parse TestData from completion {"data": []}. '
                    "Got: 含 '{{}}' 双花括号")
            if repair_ctx is not None:
                seen_ctx.update(repair_ctx)
            return self._ok_gen(row, api, ctx, path)

        agent = ChatTestAgentGraph()
        result = agent._run_yaml_rounds(
            self._tasks(str(tmp_path)), "[]", "", str(tmp_path),
            gen_func=flaky, repair_rounds=1)
        assert result["success"] == 3
        assert result["repaired"] == 1
        assert result["failed"] == 0
        assert result["rounds"] == 2
        assert result["errors_file"] is None
        # 自查上下文三要素齐备
        assert "双花括号" in seen_ctx["error_detail"]
        assert '{"data": []}' in seen_ctx["prior_output"]
        assert "B1" in seen_ctx["error_pattern_summary"]

    def test_persistent_failure_final_state(self, tmp_path):
        """修复轮仍失败 → 计 failed + _generation_errors.json，不写占位文件。"""
        def always_fail_one(row, api, ctx, path, repair_ctx=None):
            if "case_2" in path:
                raise RuntimeError("json/params/data 三选一，检测到并存: json + params")
            return self._ok_gen(row, api, ctx, path)

        agent = ChatTestAgentGraph()
        result = agent._run_yaml_rounds(
            self._tasks(str(tmp_path)), "[]", "", str(tmp_path),
            gen_func=always_fail_one, repair_rounds=1)
        assert result["success"] == 2
        assert result["failed"] == 1
        assert result["rounds"] == 2
        errors_file = result["errors_file"]
        assert errors_file and os.path.exists(errors_file)
        entries = json.loads(open(errors_file, encoding="utf-8").read())
        assert len(entries) == 1
        e = entries[0]
        assert e["case_id"] == "TC-002"
        assert e["placeholder_id"].startswith("GEN-FAIL-R2-")
        assert e["rounds_attempted"] == 2
        assert "三选一" in e["error"]
        # 失败项不写占位假文件
        assert not os.path.exists(os.path.join(str(tmp_path), "case_2", "test_data.yaml"))

    def test_zero_repair_rounds(self, tmp_path):
        """repair_rounds=0 → 只跑全量轮，失败直接终态，不重试。"""
        def fail_all(row, api, ctx, path, repair_ctx=None):
            raise RuntimeError("boom")

        agent = ChatTestAgentGraph()
        result = agent._run_yaml_rounds(
            self._tasks(str(tmp_path), n=2), "[]", "", str(tmp_path),
            gen_func=fail_all, repair_rounds=0)
        assert result["rounds"] == 1
        assert result["failed"] == 2
        assert result["success"] == 0

    # 注：2026-08-03 起 _run_yaml_rounds 已移除熔断逻辑（circuit_breaker_threshold 参数
    #     及对应 RuntimeError 分支均不存在，阈值配置 YAML_FAILURE_CIRCUIT_BREAKER 仅残留），
    #     原 test_circuit_breaker_triggers_on_mass_failure 测试的对象已不存在，故删除。


# ============================================================
# M7b — YAML 单节点生成（恢复）+ 思考内容监测
# ============================================================

class TestYamlSingleNodeFlag:
    """YAML_SINGLE_NODE 开关 → _generate_all_yamls 传入的 gen_func 路由。"""

    @staticmethod
    def _agent(monkeypatch):
        from unittest.mock import Mock
        import infrastructure.observability as observability
        agent = object.__new__(ChatTestAgentGraph)
        agent._read_excel_rows = lambda path: [{
            "feature": "设施管理", "story": "设施添加", "title": "新增设施-正向",
            "steps": "1.调用新增接口", "expected": "", "preconditions": "",
            "case_id": "TC-001",
        }]
        agent._translate_to_en = lambda path, rows: {
            "feature_en": {}, "story_en": {}, "title_en": {}}
        agent._read_shared_preconditions = lambda path: []
        captured = {}
        agent._run_yaml_rounds = Mock(
            side_effect=lambda *a, **kw: captured.update(kw) or {
                "total": 1, "success": 1, "failed": 0, "repaired": 0,
                "rounds": 1, "errors_file": None})
        agent._log_node_output = Mock()
        monkeypatch.setattr(observability, "log_thinking", Mock())
        return agent, captured

    def test_true_selects_single_node(self, monkeypatch, tmp_path):
        """YAML_SINGLE_NODE=True → gen_func=_generate_one_yaml_single。"""
        import infrastructure.config as config
        monkeypatch.setattr(config, "YAML_SINGLE_NODE", True)
        agent, captured = self._agent(monkeypatch)
        agent._generate_all_yamls(
            os.path.join(str(tmp_path), "test_plan.xlsx"), "[]", "ctx")
        assert captured["gen_func"] == agent._generate_one_yaml_single

    # 两段式已注释（2026-08-24）：_generate_all_yamls 恒走单节点，不再存在
    # YAML_SINGLE_NODE=False → gen_func=None 的路径。
    # def test_false_selects_two_stage(self, monkeypatch, tmp_path):
    #     """YAML_SINGLE_NODE=False → gen_func=None（两段式 _generate_one_yaml）。"""
    #     import infrastructure.config as config
    #     monkeypatch.setattr(config, "YAML_SINGLE_NODE", False)
    #     agent, captured = self._agent(monkeypatch)
    #     agent._generate_all_yamls(
    #         os.path.join(str(tmp_path), "test_plan.xlsx"), "[]", "ctx")
    #     assert captured["gen_func"] is None


class TestYamlSingleNodeGenerate:
    """单节点：thinking + json_object 一次调用，绕开 _invoke_structured。"""

    @staticmethod
    def _agent(monkeypatch, llm_text: str):
        from unittest.mock import Mock
        import infrastructure.observability as observability
        agent = object.__new__(ChatTestAgentGraph)
        agent.llm = Mock()
        agent._invoke_think = Mock(return_value=llm_text)
        agent._invoke_structured = Mock()
        agent._load_factory_methods = lambda: ""
        monkeypatch.setattr(observability, "log_thinking", Mock())
        return agent

    def test_single_node_writes_yaml_bypasses_structured(self, monkeypatch, tmp_path):
        """单节点一次调用生成合法 TestData JSON → 落盘；_invoke_structured 不被调用。"""
        import infrastructure.config as config
        monkeypatch.setattr(config, "YAML_SINGLE_NODE", True)
        yaml_json = json.dumps({
            "data": [{
                "baseInfo": {"api_name": "新增", "url": "/meterDevice/add",
                             "method": "post", "header": {}},
                "testCase": [{"case_name": "新增单一费率电表",
                              "validation": [{"eq": {"code": 0}}]}],
            }]
        }, ensure_ascii=False)
        agent = self._agent(monkeypatch, yaml_json)
        out = os.path.join(str(tmp_path), "feature_en", "func_en", "test_data.yaml")
        ret = agent._generate_one_yaml_single(
            {"steps": "1.调用新增接口", "expected": "1.[eq]成功", "case_id": "TC-001"},
            "[]", "用户意图", out)
        assert ret == out
        assert os.path.exists(out)
        assert "/meterDevice/add" in open(out, encoding="utf-8").read()
        agent._invoke_structured.assert_not_called()
        _, kwargs = agent._invoke_think.call_args
        assert kwargs.get("reasoning_label") == "generate_yaml_data_single"

    def test_single_node_messages_schema_in_system_not_human(self, monkeypatch, tmp_path):
        """单节点 prompt 分层：schema 固定入 system，human 只放可变内容、无示例字面量。"""
        import infrastructure.config as config
        monkeypatch.setattr(config, "YAML_SINGLE_NODE", True)
        yaml_json = json.dumps({
            "data": [{
                "baseInfo": {"api_name": "新增", "url": "/meterDevice/add",
                             "method": "post", "header": {}},
                "testCase": [{"case_name": "新增电表",
                              "validation": [{"eq": {"code": 0}}]}],
            }]
        }, ensure_ascii=False)
        agent = self._agent(monkeypatch, yaml_json)
        out = os.path.join(str(tmp_path), "feature_en", "func_en", "test_data.yaml")
        agent._generate_one_yaml_single(
            {"steps": "1.调用新增接口", "expected": "1.[eq]成功", "case_id": "TC-001"},
            "[]", "用户意图", out)

        msgs = agent._invoke_think.call_args[0][1]
        by_type = {m.type: m.content for m in msgs}
        sys_text, hum_text = by_type["system"], by_type["human"]
        # schema 固定入 system；human 无 {json_schema} 占位
        assert "$defs" in sys_text and "StepData" in sys_text
        assert "{json_schema}" not in hum_text
        # human 三节可变来源
        for k in ("B 用例内容", "A 数据分析", "接口详情文档"):
            assert k in hum_text
        # 无可照抄断言字面量（防 LLM 照抄）
        text = sys_text + hum_text
        assert "retCode: 0" not in text
        assert "$.retCode" not in text
        # 19号 prompt 可修缺陷铁律
        assert "接口返回定义" in sys_text
        assert "数据唯一化" in sys_text
        assert "delete 参数按接口定义" in sys_text


class TestYamlSingleNodePromptRender:
    """generate_yaml_data_single_prompt 直接渲染：schema 只在 system、无示例字面量。"""

    @staticmethod
    def _render() -> dict:
        from prompts.extraction_prompts import generate_yaml_data_single_prompt
        from prompts.response_model import TestData
        schema = json.dumps(TestData.model_json_schema(), ensure_ascii=False, indent=2)
        msgs = generate_yaml_data_single_prompt().format_messages(
            data_factory_methods="factory_text",
            api_definitions="api_defs",
            data_analysis="analysis",
            test_case_logic="test_case_logic",
            user_context="user_ctx",
            db_schema="db_schema",
            json_schema=schema,
        )
        return {m.type: m.content for m in msgs}

    def test_schema_in_system_only(self):
        """{json_schema} 占位替换进 system；human 不含占位。"""
        msgs = self._render()
        assert "$defs" in msgs["system"] and "StepData" in msgs["system"]
        assert "{json_schema}" not in msgs["system"]
        assert "{json_schema}" not in msgs["human"]

    def test_no_copyable_assertion_literals(self):
        """整体无 retCode: 0 / $.retCode 可照抄断言字面量。"""
        text = self._render()["system"] + self._render()["human"]
        assert "retCode: 0" not in text
        assert "$.retCode" not in text

    def test_sources_and_rules_present(self):
        """system 含铁律；human 含 B/A/接口 三节来源。"""
        msgs = self._render()
        assert "接口返回定义" in msgs["system"]
        assert "数据唯一化" in msgs["system"]
        assert "delete 参数按接口定义" in msgs["system"]
        for k in ("B 用例内容", "A 数据分析", "接口详情文档"):
            assert k in msgs["human"]

    def test_no_project_specific_retcode(self):
        """不硬编码任何项目特例取值（retCode 1/0）——断言值委托接口返回定义。"""
        text = self._render()["system"]
        assert "retCode: 1" not in text
        assert "retCode: 0" not in text
        assert "成功=1" not in text


# 两段式 format_yaml_data_prompt 已注释（2026-08-24），本类测试对象不存在。
# class TestFormatYamlPromptNoProjectRetcode:
#     """两段式 format_yaml_data_prompt 同思路：不硬编码项目特例 retCode 取值。"""
#
#     @staticmethod
#     def _render() -> str:
#         from prompts.extraction_prompts import format_yaml_data_prompt
#         m = format_yaml_data_prompt().format_messages(
#             data_factory_methods="factory_text",
#             api_definitions="api_defs",
#             data_analysis="analysis",
#             test_case_logic="test_case_logic",
#             user_context="user_ctx",
#             db_schema="db_schema",
#         )
#         return {x.type: x.content for x in m}["system"]
#
#     def test_no_retcode_zero_hardcode(self):
#         """示例与规则均无 retCode: 0 / $.retCode 可照抄字面量。"""
#         text = self._render()
#         assert "retCode: 0" not in text
#         assert "$.retCode" not in text
#
#     def test_assertion_value_from_return_definition(self):
#         """铁律 19：断言期望值取自接口返回定义（同单节点思路）。"""
#         assert "返回定义中真实给出的字段" in self._render()


# ============================================================
# M8 — Phase C 接口定义解析与缺失阻断（web/tasks._resolve_api_defs）
# ============================================================

from web.tasks import _resolve_api_defs


class TestResolveApiDefs:
    """接口定义优先级: 显式入参 > SQL(documents 表) 全部接口详情 > None(阻断)。

    2026-08-18 改：Phase C 不再用 Phase B 快照（ChromaDB 检索缺 body/return，
    导致 YAML 生成字段瞎编 meterName/meterNo），直接用 SQL 读全部接口详情。
    """

    @staticmethod
    def _excel(tmp_path) -> str:
        p = tmp_path / "test_plan.xlsx"
        p.write_text("dummy", encoding="utf-8")
        return str(p)

    def test_param_takes_priority(self, tmp_path):
        """显式入参信任原样，不查 SQL。"""
        excel = self._excel(tmp_path)
        assert _resolve_api_defs(excel, '[{"url": "/from_param"}]') == '[{"url": "/from_param"}]'

    def test_empty_param_reads_all_api_details_from_sql(self):
        """空入参 → SQL 读全部接口详情（body 六字段完整，字段名真实）。"""
        resolved = _resolve_api_defs("/nonexistent/test_plan.xlsx", "")
        assert resolved
        apis = json.loads(resolved)
        assert len(apis) >= 60  # 全部接口
        add = next((a for a in apis if a.get("url") == "/electricMeter/add"), None)
        assert add is not None
        body = add.get("body", [])
        assert len(body) >= 60  # body 字段齐全
        names = {f.get("name") for f in body}
        assert {"name", "code", "meterDeviceType", "meterTypeCode"} <= names
        assert len(add.get("return", [])) >= 1

    def test_empty_list_param_treated_as_missing(self):
        """入参 "[]"（_build_response 的空占位）不算有效定义，继续读 SQL。"""
        resolved = _resolve_api_defs("", "[]")
        assert resolved and "/electricMeter/add" in resolved

    def test_annotations_included_from_sql(self):
        """annotations 从 SQL 列带出（has_path_params / is_export）。"""
        resolved = _resolve_api_defs("", "")
        apis = json.loads(resolved)
        pp = next(a for a in apis if a.get("url") == "/electricMeter/getEle/{code}")
        assert pp["annotations"]["has_path_params"]["active"] is True
        imp = next(a for a in apis if a.get("url") == "/electricMeter/importTemplate")
        assert imp["annotations"]["is_export"]["active"] is True

    def test_no_api_docs_returns_none(self, tmp_path, monkeypatch):
        """SQL 无接口定义 → None 阻断（M8：宁可失败，不带残缺数据续跑）。"""
        import web.tasks as _tasks
        monkeypatch.setattr(_tasks, "_load_all_api_defs", lambda: [])
        assert _resolve_api_defs(self._excel(tmp_path), "") is None

    # ---- 2026-08-20 模块作用域收敛（模块绑定 + 关联模块绑定）----

    def test_module_scope_returns_only_bound_apis(self):
        """module_name → 仅模块绑定接口（详情形式，非全量 72）。"""
        resolved = _resolve_api_defs("", "", module_name="智慧用电")
        assert resolved
        apis = json.loads(resolved)
        urls = {a.get("url") for a in apis}
        assert urls == {
            "/electricMeter/add", "/electricMeter/delete",
            "/electricMeter/getList", "/electricMeter/getParentList",
        }
        assert len(apis) < 60  # 已收敛，非全部
        add = next(a for a in apis if a.get("url") == "/electricMeter/add")
        names = {f.get("name") for f in add.get("body", [])}
        assert {"name", "code", "meterDeviceType", "meterTypeCode"} <= names  # 详情完整
        assert len(add.get("return", [])) >= 1

    def test_module_scope_reads_module_scope_file(self, tmp_path):
        """计划目录 module_scope.json → 免传 module_name 也作用域化（B→C 产物传递）。"""
        scope = tmp_path / "module_scope.json"
        scope.write_text(json.dumps(
            {"module": "智慧用电", "related_modules": []},
            ensure_ascii=False), encoding="utf-8")
        resolved = _resolve_api_defs(self._excel(tmp_path), "")
        assert resolved
        urls = {a.get("url") for a in json.loads(resolved)}
        assert "/electricMeter/add" in urls
        assert len(urls) == 4

    def test_module_scope_empty_falls_back_to_all(self):
        """无绑定接口的模块 → 回退全部接口定义（不空定义续跑）。"""
        resolved = _resolve_api_defs("", "", module_name="不存在的模块")
        assert resolved
        apis = json.loads(resolved)
        assert len(apis) >= 60  # 回退全量

    def test_module_scope_includes_related_modules(self, monkeypatch):
        """关联模块绑定接口纳入作用域（monkeypatch 关联发现）。"""
        # 园区管理本身绑定 0 接口；关联模块智慧用电绑定 4 个 → 全部纳入
        monkeypatch.setattr(
            "database.operations.bindings.discover_related_modules",
            lambda session, name: ["智慧用电"],
        )
        resolved = _resolve_api_defs("", "", module_name="园区管理")
        assert resolved
        urls = {a.get("url") for a in json.loads(resolved)}
        assert "/electricMeter/add" in urls  # 关联模块绑定接口被纳入


# ============================================================
# discover_related_modules（关联模块三路召回，共享函数）
# ============================================================

from database.operations.bindings import discover_related_modules as _discover_related


class TestDiscoverRelatedModules:
    """绑定图三路召回：module↔module + product/axure/api→module；去重、排除自身、排序。"""

    class _FakeSession:
        pass

    def test_module_module_direct(self, monkeypatch):
        """路径1 module↔module：取关联、跳过自身、去重排序。"""
        import database.operations.bindings as _b
        monkeypatch.setattr(_b.BindingOps, "get_partners",
                            lambda *a, **k: [("module", "B"), ("module", "智慧用电"),
                                             ("module", "A"), ("module", "B")])
        monkeypatch.setattr(_b.BindingOps, "get_bound_docs", lambda *a, **k: [])
        monkeypatch.setattr(_b.BindingOps, "get_partners_batch", lambda *a, **k: {})
        assert _discover_related(self._FakeSession(), "智慧用电") == ["A", "B"]

    def test_doc_to_module_paths(self, monkeypatch):
        """路径2+3：product/axure/api 文档 → 共享模块。"""
        import database.operations.bindings as _b

        class _D:
            def __init__(self, did, doc_type):
                self.id, self.doc_type = did, doc_type

        monkeypatch.setattr(_b.BindingOps, "get_partners", lambda *a, **k: [])
        monkeypatch.setattr(_b.BindingOps, "get_bound_docs",
                            lambda *a, **k: [_D("p1", "product"), _D("a1", "axure"),
                                             _D("api1", "api")])
        monkeypatch.setattr(_b.BindingOps, "get_partners_batch",
                            lambda *a, **k: {"p1": [("module", "M2")],
                                             "a1": [("module", "M1")],
                                             "api1": [("module", "M3")]})
        # a1 → M1 是自身，被排除
        assert _discover_related(self._FakeSession(), "M1") == ["M2", "M3"]

    def test_empty_module_returns_empty(self, monkeypatch):
        """空模块名 / 无绑定 → 空列表。"""
        import database.operations.bindings as _b
        monkeypatch.setattr(_b.BindingOps, "get_partners", lambda *a, **k: [])
        monkeypatch.setattr(_b.BindingOps, "get_bound_docs", lambda *a, **k: [])
        monkeypatch.setattr(_b.BindingOps, "get_partners_batch", lambda *a, **k: {})
        assert _discover_related(self._FakeSession(), "") == []
        assert _discover_related(self._FakeSession(), "孤立模块") == []
